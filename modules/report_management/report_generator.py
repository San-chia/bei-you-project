# modules/report_management/report_generator.py
"""报表生成器模块"""
import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List, Any
import plotly.graph_objects as go
import plotly.io as pio
import base64
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.drawing.image import Image as XLImage

class ReportGenerator:
    """报表生成器"""
    
    def __init__(self):
        self.default_font = "Arial"
        
    def generate_pdf_html(self, template: dict, data: dict) -> str:
        """生成PDF用的HTML内容（可以用浏览器打印为PDF）"""
        html_parts = []
        
        # HTML头部
        html_parts.append("""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>{title}</title>
            <style>
                @media print {{
                    .no-print {{ display: none; }}
                    .page-break {{ page-break-after: always; }}
                }}
                body {{
                    font-family: Arial, sans-serif;
                    margin: 20px;
                    color: #333;
                }}
                h1 {{ color: #1f77b4; margin-bottom: 10px; }}
                h2 {{ color: #333; margin-top: 30px; margin-bottom: 15px; }}
                .info {{ color: #666; font-size: 14px; margin-bottom: 20px; }}
                table {{
                    border-collapse: collapse;
                    width: 100%;
                    margin: 20px 0;
                }}
                th, td {{
                    border: 1px solid #ddd;
                    padding: 8px;
                    text-align: left;
                }}
                th {{
                    background-color: #1f77b4;
                    color: white;
                    font-weight: bold;
                }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                .summary-cards {{
                    display: grid;
                    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                    gap: 20px;
                    margin: 20px 0;
                }}
                .card {{
                    border: 1px solid #ddd;
                    padding: 20px;
                    text-align: center;
                    background: #f8f9fa;
                    border-radius: 8px;
                }}
                .card h3 {{ margin: 0; color: #1f77b4; }}
                .card .label {{ color: #666; font-size: 14px; margin-bottom: 10px; }}
                .chart-container {{
                    margin: 20px 0;
                    text-align: center;
                }}
                .chart-container img {{
                    max-width: 100%;
                    height: auto;
                }}
            </style>
        </head>
        <body>
        """.format(title=template['name']))
        
        # 标题和信息
        html_parts.append(f"""
            <h1>{template['name']}</h1>
            <div class="info">生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
        """)
        
        # 生成各个section
        for section in template.get('default_sections', []):
            section_html = self._generate_html_section(section, data.get(section['id'], {}))
            html_parts.append(section_html)
        
        # HTML结尾
        html_parts.append("""
        </body>
        </html>
        """)
        
        return ''.join(html_parts)
    
    def generate_excel(self, template: dict, data: dict, output_path: str = None) -> bytes:
        """生成Excel报表"""
        wb = openpyxl.Workbook()
        
        # 创建总览工作表
        ws = wb.active
        ws.title = "报表总览"
        
        # 设置标题
        ws['A1'] = template['name']
        ws['A2'] = f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # 标题样式
        title_font = Font(size=16, bold=True, color="1f77b4")
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells('A1:F1')
        ws.merge_cells('A2:F2')
        
        current_row = 4
        
        # 为每个section创建内容
        for section in template.get('default_sections', []):
            section_type = section['type']
            section_data = data.get(section['id'])
            
            if section_type == 'table' and isinstance(section_data, pd.DataFrame):
                # 为表格数据创建新的工作表
                sheet_name = section['title'][:31]  # Excel工作表名称限制
                if sheet_name in wb.sheetnames:
                    sheet_name = f"{sheet_name}_{len(wb.sheetnames)}"
                
                ws_table = wb.create_sheet(title=sheet_name)
                self._add_dataframe_to_sheet(ws_table, section_data, section['title'])
                
                # 在总览表中添加链接
                ws.cell(row=current_row, column=1, value=f"查看: {section['title']}")
                ws.cell(row=current_row, column=1).hyperlink = f"#{sheet_name}!A1"
                ws.cell(row=current_row, column=1).font = Font(color="0000FF", underline="single")
                current_row += 2
                
            elif section_type == 'summary_cards':
                current_row = self._add_summary_to_excel(
                    ws, section, data.get(section['id'], {}), current_row
                )
                current_row += 1
                
            elif section_type in ['chart', 'kpi_cards']:
                # 添加图表预览说明
                ws.cell(row=current_row, column=1, value=section['title'])
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
                current_row += 1
                
                # 如果有图表数据，创建Excel图表
                if section_type == 'chart' and isinstance(section_data, pd.DataFrame):
                    self._add_chart_to_excel(ws, section, section_data, current_row)
                    current_row += 15  # 为图表预留空间
                else:
                    ws.cell(row=current_row, column=1, value="（图表数据见网页版报表）")
                    current_row += 2
        
        # 调整列宽
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存或返回
        if output_path:
            wb.save(output_path)
        else:
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
    
    def generate_charts_as_images(self, template: dict, data: dict) -> Dict[str, str]:
        """将图表生成为base64编码的图片"""
        charts = {}
        
        for section in template.get('default_sections', []):
            if section['type'] == 'chart':
                section_data = data.get(section['id'])
                if section_data:
                    fig = self._create_plotly_chart(section, section_data)
                    if fig:
                        # 将图表转换为base64图片
                        img_bytes = fig.to_image(format='png', width=800, height=400)
                        img_base64 = base64.b64encode(img_bytes).decode()
                        charts[section['id']] = f"data:image/png;base64,{img_base64}"
        
        return charts
    
    def _generate_html_section(self, section: dict, data: Any) -> str:
        """生成HTML章节内容"""
        html_parts = []
        
        # 章节标题
        html_parts.append(f'<h2>{section["title"]}</h2>')
        
        # 根据类型生成内容
        if section['type'] == 'summary_cards' and isinstance(data, dict):
            html_parts.append('<div class="summary-cards">')
            for metric in section.get('metrics', []):
                value = data.get(metric['field'], 'N/A')
                if metric['format'] == 'currency' and isinstance(value, (int, float)):
                    value = f"¥{value:,.2f}"
                elif metric['format'] == 'percentage' and isinstance(value, (int, float)):
                    value = f"{value*100:.1f}%"
                
                html_parts.append(f'''
                    <div class="card">
                        <div class="label">{metric['name']}</div>
                        <h3>{value}</h3>
                    </div>
                ''')
            html_parts.append('</div>')
            
        elif section['type'] == 'table' and isinstance(data, pd.DataFrame):
            html_parts.append('<table>')
            # 表头
            html_parts.append('<thead><tr>')
            for col in section.get('columns', []):
                html_parts.append(f'<th>{col["title"]}</th>')
            html_parts.append('</tr></thead>')
            
            # 表体
            html_parts.append('<tbody>')
            for _, row in data.iterrows():
                html_parts.append('<tr>')
                for col in section.get('columns', []):
                    field = col['field']
                    value = row.get(field, '')
                    
                    # 格式化值
                    if col.get('format') == 'currency' and isinstance(value, (int, float)):
                        value = f"¥{value:,.2f}"
                    elif col.get('format') == 'percentage' and isinstance(value, (int, float)):
                        value = f"{value*100:.1f}%"
                    
                    html_parts.append(f'<td>{value}</td>')
                html_parts.append('</tr>')
            html_parts.append('</tbody>')
            html_parts.append('</table>')
            
        elif section['type'] == 'chart':
            # 这里可以插入图表图片
            html_parts.append(f'<div class="chart-container">')
            html_parts.append(f'<p>（图表：{section["title"]}）</p>')
            html_parts.append(f'</div>')
        
        return '\n'.join(html_parts)
    
    def _add_dataframe_to_sheet(self, ws, df: pd.DataFrame, title: str):
        """将DataFrame添加到Excel工作表"""
        # 添加标题
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells(f'A1:{chr(65 + len(df.columns) - 1)}1')
        
        # 添加表头
        header_row = 3
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 添加数据
        for row_idx, row in enumerate(df.values, header_row + 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # 添加边框
        thin = Side(border_style="thin", color="000000")
        for row in ws.iter_rows(min_row=header_row, max_row=header_row + len(df), 
                               min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        # 添加汇总行
        if any(col.get('format') in ['currency', 'numeric'] for col in df.columns):
            total_row = header_row + len(df) + 1
            ws.cell(row=total_row, column=1, value="合计")
            ws.cell(row=total_row, column=1).font = Font(bold=True)
            
            for col_idx in range(2, len(df.columns) + 1):
                col_name = df.columns[col_idx - 1]
                if df[col_name].dtype in ['int64', 'float64']:
                    total_value = df[col_name].sum()
                    cell = ws.cell(row=total_row, column=col_idx, value=total_value)
                    cell.number_format = '#,##0.00'
                    cell.font = Font(bold=True)
    
    def _add_summary_to_excel(self, ws, section: dict, data: dict, start_row: int) -> int:
        """将汇总数据添加到Excel"""
        # 标题
        ws.cell(row=start_row, column=1, value=section['title'])
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)
        
        current_row = start_row + 2
        
        # 创建一个表格样式的汇总
        for i, metric in enumerate(section.get('metrics', [])):
            # 指标名称
            name_cell = ws.cell(row=current_row, column=1, value=metric['name'])
            name_cell.font = Font(bold=True)
            name_cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
            
            # 指标值
            value = data.get(metric['field'], 0)
            value_cell = ws.cell(row=current_row, column=2)
            
            if metric['format'] == 'currency':
                value_cell.value = value
                value_cell.number_format = '¥#,##0.00'
            elif metric['format'] == 'percentage':
                value_cell.value = value
                value_cell.number_format = '0.0%'
            else:
                value_cell.value = value
            
            # 添加边框
            thin = Side(border_style="thin", color="CCCCCC")
            for col in [1, 2]:
                cell = ws.cell(row=current_row, column=col)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            
            current_row += 1
        
        return current_row + 1
    
    def _add_chart_to_excel(self, ws, section: dict, data: pd.DataFrame, start_row: int):
        """在Excel中添加图表"""
        chart_type = section.get('chart_type')
        
        # 先将数据添加到工作表（隐藏区域）
        data_start_col = 10  # 从J列开始放置数据
        
        # 添加数据
        for col_idx, col_name in enumerate(data.columns):
            ws.cell(row=start_row, column=data_start_col + col_idx, value=col_name)
            for row_idx, value in enumerate(data[col_name]):
                ws.cell(row=start_row + 1 + row_idx, column=data_start_col + col_idx, value=value)
        
        # 创建图表
        if str(chart_type).endswith('bar'):
            chart = BarChart()
        elif str(chart_type).endswith('line'):
            chart = LineChart()
        elif str(chart_type).endswith('pie'):
            chart = PieChart()
        else:
            chart = LineChart()  # 默认线图
        
        # 设置图表数据
        data_ref = Reference(ws, min_col=data_start_col + 1, min_row=start_row,
                           max_col=data_start_col + len(data.columns) - 1,
                           max_row=start_row + len(data))
        
        cats = Reference(ws, min_col=data_start_col, min_row=start_row + 1,
                        max_row=start_row + len(data))
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.title = section.get('title', '')
        
        # 添加图表到工作表
        ws.add_chart(chart, f"A{start_row}")
    
    def _create_plotly_chart(self, section: dict, data: Any) -> go.Figure:
        """创建Plotly图表"""
        chart_type = str(section.get('chart_type', '')).lower()
        
        if isinstance(data, pd.DataFrame) and not data.empty:
            if 'line' in chart_type:
                fig = go.Figure()
                for col in data.columns[1:]:  # 假设第一列是x轴
                    fig.add_trace(go.Scatter(
                        x=data.iloc[:, 0],
                        y=data[col],
                        mode='lines+markers',
                        name=col
                    ))
            elif 'bar' in chart_type:
                fig = go.Figure()
                for col in data.columns[1:]:
                    fig.add_trace(go.Bar(
                        x=data.iloc[:, 0],
                        y=data[col],
                        name=col
                    ))
            elif 'pie' in chart_type:
                fig = go.Figure(data=[go.Pie(
                    labels=data.iloc[:, 0],
                    values=data.iloc[:, 1]
                )])
            else:
                return None
                
            fig.update_layout(
                title=section.get('title', ''),
                template='plotly_white',
                height=400
            )
            return fig
        
        return None

def export_report(template_id: str, template: dict, data: dict, format: str = 'pdf') -> dict:
    """导出报表的统一接口"""
    generator = ReportGenerator()
    
    if format == 'pdf':
        # 生成HTML，让用户通过浏览器打印为PDF
        html_content = generator.generate_pdf_html(template, data)
        return {
            'content': html_content,
            'filename': f"{template_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html",
            'type': 'text/html'
        }
    elif format == 'xlsx':
        excel_bytes = generator.generate_excel(template, data)
        return {
            'content': excel_bytes,
            'filename': f"{template_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            'type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
    else:
        # 默认返回Excel
        excel_bytes = generator.generate_excel(template, data)
        return {
            'content': excel_bytes,
            'filename': f"{template_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            'type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }