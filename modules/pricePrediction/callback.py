# project_20250427_V3/modules/price/callback.py
import io
from datetime import datetime, timedelta  # 添加 timedelta 导入
import dash
from dash import dcc, html, dash_table, callback_context
import dash_bootstrap_components as dbc
from dash.dependencies import Input, Output, State
import pandas as pd
import numpy as np
from dash.exceptions import PreventUpdate
import mysql.connector  # 替换sqlite3
import logging
import json
import uuid
import traceback  # 新增：用于错误追踪

# 导入预测系统类
from .data import CostPredictionSystem
from .change import process_and_update_databases
from .indicator_mapping import (
    INDICATOR_FIELD_MAPPING,
    get_all_indicators_for_mode,
    get_fields_by_indicator,
    get_indicator_config
)
# 导入模态窗口创建函数
from .modals import (
    create_steel_reinforcement_parameter_modal,
    create_custom_mode_parameter_modal,
    create_price_modification_modal,
    create_steel_lining_parameter_modal,
)

# 导入MySQL配置和连接函数
# MySQL数据库连接配置
MYSQL_CONFIG = {
    'host': 'localhost',  # 修改为您的MySQL服务器地址
    'user': 'dash',  # 修改为您的MySQL用户名
    'password': '123456',  # 修改为您的MySQL密码
    'database': 'dash_project',  # 修改为您的MySQL数据库名
    'charset': 'utf8mb4',
    'autocommit': True  # 默认启用自动提交
}

# 在MYSQL_CONFIG之后添加表名映射
TABLE_NAME_MAPPING = {
    '价格基准1': 'price_baseline_1',
    '价格基准2': 'price_baseline_2',
    '关键因素1': 'key_factors_1',
    '关键因素2': 'key_factors_2'
}

def get_english_table_name(chinese_name):
    """将中文表名转换为英文表名"""
    return TABLE_NAME_MAPPING.get(chinese_name, chinese_name)
def get_connection():
    """获取MySQL数据库连接"""
    try:
        conn = mysql.connector.connect(**MYSQL_CONFIG)
        return conn
    except Exception as e:
        print(f"MySQL连接失败: {e}")
        raise

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

PRIMARY_COLOR = "#2C3E50"

global_predictors = {
    'steel_cage': None,
    'steel_lining': None,
}

def safe_float(value):
    try:
        if value is None or str(value).strip() == "" :
            return 0.0
        return float(value)
    except (ValueError, TypeError):
        return 0.0

def initialize_prediction_system(mode):
    """初始化或重新初始化指定模式的全局预测系统"""
    global global_predictors
    if global_predictors[mode] is None:
        logger.info(f"初始化/重新初始化机器学习预测系统 for mode: {mode}...")
        try:
            current_predictor = CostPredictionSystem(mode=mode)
            if current_predictor.load_data_from_database(mode):
                if current_predictor.train_system():
                    global_predictors[mode] = current_predictor
                    logger.info(f"✓ 机器学习预测系统初始化/重新初始化成功 for mode: {mode}")
                    return True
                else:
                    logger.error(f"✗ 机器学习模型训练失败 for mode: {mode}")
                    global_predictors[mode] = None
                    return False
            else:
                logger.error(f"✗ 历史数据加载失败 for mode: {mode}")
                global_predictors[mode] = None
                return False
        except Exception as e:
            logger.error(f"✗ 预测系统初始化/重新初始化异常 for mode: {mode}: {e}", exc_info=True)
            global_predictors[mode] = None
            return False
    else:
        logger.info(f"机器学习预测系统已初始化 for mode: {mode}，跳过重复初始化。")
    return True

def save_report_to_database(report_data, report_type):
    """
    保存报告数据到数据库
    
    Args:
        report_data: 报告数据字典
        report_type: 报告类型 ('steel_cage', 'steel_lining', 'custom_mode')
    
    Returns:
        bool: 保存成功返回True，失败返回False
    """
    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()
        
        # 生成唯一报告ID
        report_id = str(uuid.uuid4())
        
        # 提取基本信息
        mode = report_data.get('模式', '')
        generation_time = report_data.get('生成时间', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        # 根据不同模式提取预测结果 - 适配新的数据结构
        total_cost = 0
        ml_ensemble_prediction = None
        ratio_method_prediction = None
        prediction_status_summary = {}
        
        if report_type in ['steel_cage', 'steel_lining']:
            prediction_results = report_data.get('预测结果', {})
            if prediction_results:
                # 提取主要预测结果
                main_ai_result = report_data.get('主要AI预测', {})
                main_ratio_result = report_data.get('主要比率法预测', {})
                
                # 获取有效的预测值
                if main_ai_result.get('status') == 'success':
                    ml_ensemble_prediction = main_ai_result.get('value')
                elif main_ai_result.get('status') == 'display_disabled':
                    ml_ensemble_prediction = "显示被禁用"
                else:
                    ml_ensemble_prediction = None
                    
                if main_ratio_result.get('status') == 'success':
                    ratio_method_prediction = main_ratio_result.get('value')
                elif main_ratio_result.get('status') == 'display_disabled':
                    ratio_method_prediction = "显示被禁用"
                else:
                    ratio_method_prediction = None
                
                # 计算总成本（优先使用数值结果）
                if isinstance(ml_ensemble_prediction, (int, float)):
                    total_cost = ml_ensemble_prediction
                elif isinstance(ratio_method_prediction, (int, float)):
                    total_cost = ratio_method_prediction
                else:
                    total_cost = 0
                
                # 获取状态汇总信息用于Excel
                prediction_status_summary = format_prediction_status_for_export(prediction_results)
                
        elif report_type == 'steel_lining':
            prediction_results = report_data.get('预测结果', {})
            if prediction_results:
                ml_predictions = prediction_results.get('机器学习预测结果', {})
                ml_ensemble_prediction = ml_predictions.get('集成平均预测')
                ratio_method_prediction = prediction_results.get('比率法预测总价')
                total_cost = ml_ensemble_prediction or ratio_method_prediction or 0
                
        elif report_type == 'custom_mode':
            prediction_results = report_data.get('预测结果', {})
            if prediction_results:
                total_cost = prediction_results.get('total_predicted_cost', 0)
        
        # 检查并创建报告主表（如果不存在）
        if not ensure_reports_table_exists(cursor):
            logger.error("无法确保报告表存在，取消保存操作")
            return False
        
        # 插入报告主记录
        insert_query = """
        INSERT INTO `prediction_reports` 
        (report_id, report_type, mode_name, generation_time, total_predicted_cost, 
         ml_ensemble_prediction, ratio_method_prediction, raw_data)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """
        
        # 将完整的report_data转换为JSON存储
        raw_data_json = json.dumps(report_data, ensure_ascii=False, default=str)
        
        cursor.execute(insert_query, (
            report_id,
            report_type,
            mode,
            generation_time,
            total_cost,
            ml_ensemble_prediction,
            ratio_method_prediction,
            raw_data_json
        ))
        
        conn.commit()
        logger.info(f"✅ 报告已保存到数据库: ID={report_id}, 类型={report_type}, 总价={total_cost:,.2f}")
        return True
        
    except mysql.connector.Error as e:
        logger.error(f"❌ MySQL数据库保存失败: {e}")
        if conn:
            conn.rollback()
        return False
    except Exception as e:
        logger.error(f"❌ 保存报告到数据库时发生错误: {e}")
        if conn:
            conn.rollback()
        return False
    finally:
        if conn:
            conn.close()

def ensure_reports_table_exists(cursor):
    """
    确保报告表存在的辅助函数
    
    Args:
        cursor: 数据库游标
    
    Returns:
        bool: 表存在或创建成功返回True，否则返回False
    """
    create_table_query = """
    CREATE TABLE IF NOT EXISTS `prediction_reports` (
        `report_id` VARCHAR(36) PRIMARY KEY,
        `report_type` VARCHAR(50) NOT NULL,
        `mode_name` VARCHAR(100),
        `generation_time` DATETIME,
        `total_predicted_cost` DECIMAL(15,2),
        `ml_ensemble_prediction` DECIMAL(15,2),
        `ratio_method_prediction` DECIMAL(15,2),
        `raw_data` JSON,
        `created_at` TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        INDEX idx_report_type (report_type),
        INDEX idx_generation_time (generation_time),
        INDEX idx_created_at (created_at)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='预测报告数据表'
    """
    
    try:
        cursor.execute(create_table_query)
        logger.debug("✅ 报告表检查/创建成功")
        return True
        
    except mysql.connector.Error as table_error:
        logger.warning(f"⚠️ 创建或检查报告表时出现问题: {table_error}")
        
        # 尝试检查表是否已存在
        try:
            cursor.execute("SHOW TABLES LIKE 'prediction_reports'")
            result = cursor.fetchone()
            if result:
                logger.info("✅ 报告表已存在，继续执行保存操作")
                return True
            else:
                logger.error("❌ 报告表不存在且无法创建")
                return False
                
        except mysql.connector.Error as check_error:
            logger.error(f"❌ 检查表存在性失败: {check_error}")
            return False
            
    except Exception as unexpected_error:
        logger.error(f"❌ 创建表时发生意外错误: {unexpected_error}")
        return False
    
def add_measures_cost_to_predictions(ml_results, measures_cost_value, mode_name):
    """
    安全地将措施费添加到预测结果中
    
    Args:
        ml_results: 机器学习预测结果字典
        measures_cost_value: 措施费数值
        mode_name: 模式名称（用于日志）
    
    Returns:
        修改后的预测结果字典
    """
    if not ml_results or measures_cost_value <= 0:
        return ml_results
    
    try:
        # 处理机器学习预测结果
        if "机器学习预测结果" in ml_results and ml_results["机器学习预测结果"]:
            for model_key, prediction_value in ml_results["机器学习预测结果"].items():
                # 只为数值类型的预测结果添加措施费，跳过停用算法的字典状态
                if isinstance(prediction_value, (int, float)) and prediction_value is not None:
                    ml_results["机器学习预测结果"][model_key] += measures_cost_value
                    logger.debug(f"为 {mode_name} 模式的 {model_key} 添加措施费 {measures_cost_value}")
                elif isinstance(prediction_value, dict):
                    # 停用算法的状态信息保持不变
                    logger.debug(f"{mode_name} 模式的 {model_key} 已停用，跳过措施费添加")
                    continue
                else:
                    logger.warning(f"{mode_name} 模式的 {model_key} 预测值类型异常: {type(prediction_value)}")

        # 处理比率法预测结果
        ratio_prediction = ml_results.get("比率法预测总价")
        if ratio_prediction is not None and isinstance(ratio_prediction, (int, float)):
            ml_results["比率法预测总价"] += measures_cost_value
            logger.debug(f"为 {mode_name} 模式的比率法预测添加措施费 {measures_cost_value}")
        elif ratio_prediction is not None:
            logger.warning(f"{mode_name} 模式的比率法预测值类型异常: {type(ratio_prediction)}")

        return ml_results
        
    except Exception as e:
        logger.error(f"为 {mode_name} 模式添加措施费时发生错误: {e}", exc_info=True)
        return ml_results


def generate_prediction_confirmation_message(ml_prediction_results, mode_name):
    """
    生成基于预测结果状态的用户友好确认消息
    
    Args:
        ml_prediction_results (dict): 预测结果数据
        mode_name (str): 模式名称
        
    Returns:
        html.Div: 确认消息组件
    """
    if not ml_prediction_results:
        return html.Div([
            dbc.Alert([
                html.I(className="fas fa-exclamation-triangle me-2"),
                html.Strong("预测未完成！"),
                html.Br(),
                "预测系统返回空结果，请检查输入数据和系统配置。"
            ], color="danger", className="mb-0")
        ])
    
    # 检查预测方法状态
    method_status = ml_prediction_results.get("预测方法状态", {})
    algorithm_status = ml_prediction_results.get("算法执行状态", {})
    
    # 统计可用方法数量
    available_methods = 0
    total_methods = len(method_status)
    execution_issues = []
    display_issues = []
    
    for method_key, status_info in method_status.items():
        final_status = status_info.get('final_status', 'unknown')
        if final_status == 'fully_available':
            available_methods += 1
        elif final_status == 'execution_failed' or final_status == 'display_error':
            execution_issues.append(status_info.get('name', method_key))
        elif final_status == 'execute_only':
            display_issues.append(status_info.get('name', method_key))
    
    # 确定消息类型和内容
    if available_methods == total_methods:
        # 全部方法都可用
        alert_color = "success"
        icon = "fas fa-check-circle"
        title = "预测计算完成！"
        message = f"所有{mode_name}预测方法都已成功运行，智能预测结果已显示在下方表格中。"
        
    elif available_methods > 0:
        # 部分方法可用
        alert_color = "warning" 
        icon = "fas fa-exclamation-triangle"
        title = "预测部分完成！"
        message = f"{available_methods}/{total_methods}个{mode_name}预测方法可用，部分结果已显示在下方表格中。"
        
        if execution_issues:
            message += f" 执行失败的方法：{', '.join(execution_issues)}。"
        if display_issues:
            message += f" 显示被禁用的方法：{', '.join(display_issues)}。"
            
    else:
        # 没有可用方法
        alert_color = "danger"
        icon = "fas fa-times-circle"
        title = "预测失败！"
        
        algo_can_execute = algorithm_status.get('can_execute_ai', False)
        if not algo_can_execute:
            message = f"{mode_name}预测失败：{algorithm_status.get('message', 'AI算法不可用')}。请检查算法配置。"
        else:
            message = f"{mode_name}预测失败：所有预测结果的显示都被禁用。请检查综合指标设置。"
    
    return html.Div([
        dbc.Alert([
            html.I(className=f"{icon} me-2"),
            html.Strong(title),
            html.Br(),
            message
        ], color=alert_color, className="mb-0")
    ])


def extract_main_prediction_result(prediction_data):
    """
    从预测数据中提取主要结果用于报告
    
    Args:
        prediction_data: 预测数据（可能是字典、数值或None）
        
    Returns:
        dict: 标准化的预测结果信息
    """
    if prediction_data is None:
        return {
            "status": "no_data",
            "message": "无预测数据",
            "value": None
        }
    
    if isinstance(prediction_data, (int, float)):
        return {
            "status": "success",
            "message": "预测成功",
            "value": prediction_data
        }
    
    if isinstance(prediction_data, dict):
        if "error" in prediction_data:
            return {
                "status": "error",
                "message": prediction_data.get("error", "预测出错"),
                "details": prediction_data.get("reason", ""),
                "value": None
            }
        elif "status" in prediction_data:
            status = prediction_data.get("status")
            if status == "display_disabled":
                return {
                    "status": "display_disabled",
                    "message": "显示被禁用",
                    "details": prediction_data.get("message", ""),
                    "value": "hidden"
                }
            elif status in ["execution_failed", "fully_disabled"]:
                return {
                    "status": status,
                    "message": prediction_data.get("message", "预测不可用"),
                    "details": prediction_data.get("reason", ""),
                    "value": None
                }
        elif "集成平均预测" in prediction_data:
            ensemble_pred = prediction_data["集成平均预测"]
            if isinstance(ensemble_pred, (int, float)):
                return {
                    "status": "success",
                    "message": "AI预测成功",
                    "value": ensemble_pred,
                    "method": "ensemble_average"
                }
    
    return {
        "status": "unknown",
        "message": "预测结果格式异常",
        "value": None,
        "raw_data": str(prediction_data)[:100] + "..." if len(str(prediction_data)) > 100 else str(prediction_data)
    }


def format_prediction_status_for_export(prediction_results):
    """
    为导出功能格式化预测状态信息
    
    Args:
        prediction_results (dict): 完整的预测结果
        
    Returns:
        dict: 格式化的状态信息，适合导出到Excel
    """
    if not prediction_results:
        return {
            "状态概览": "预测结果为空",
            "可用方法数量": 0,
            "算法状态": "未知",
            "显示权限": "未知"
        }
    
    # 提取状态信息
    method_status = prediction_results.get("预测方法状态", {})
    algorithm_status = prediction_results.get("算法执行状态", {})
    display_status = prediction_results.get("显示权限状态", {})
    
    # 统计方法状态
    status_counts = {
        'fully_available': 0,
        'execute_only': 0, 
        'display_error': 0,
        'fully_disabled': 0
    }
    
    for method_info in method_status.values():
        final_status = method_info.get('final_status', 'unknown')
        if final_status in status_counts:
            status_counts[final_status] += 1
    
    # 格式化输出
    formatted_status = {
        "状态概览": f"完全可用:{status_counts['fully_available']}个, 可执行但显示禁用:{status_counts['execute_only']}个, 显示启用但执行失败:{status_counts['display_error']}个, 完全禁用:{status_counts['fully_disabled']}个",
        "可用方法数量": status_counts['fully_available'],
        "算法执行能力": "可用" if algorithm_status.get('can_execute_ai', False) else "不可用",
        "算法状态详情": algorithm_status.get('message', '未知'),
        "启用算法数量": algorithm_status.get('enabled_count', 0),
        "总算法数量": algorithm_status.get('total_count', 0)
    }
    
    # 添加各方法的详细状态
    for method_key, method_info in method_status.items():
        method_name = method_info.get('name', method_key)
        formatted_status[f"{method_name}_状态"] = method_info.get('final_status', 'unknown')
        formatted_status[f"{method_name}_执行消息"] = method_info.get('execution_message', '')
        formatted_status[f"{method_name}_显示消息"] = method_info.get('display_message', '')
    
    return formatted_status


def create_best_algorithm_params_section(ml_prediction_results):
    """
    创建最佳算法参数信息展示部分
    
    Args:
        ml_prediction_results (dict): 包含最佳算法信息的预测结果
        
    Returns:
        html.Div: 最佳算法参数展示组件
    """
    if not ml_prediction_results or "最佳算法信息" not in ml_prediction_results:
        return html.Div()
    
    best_algo_info = ml_prediction_results["最佳算法信息"]
    
    if not best_algo_info.get("best_algorithm_name"):
        return html.Div([
            html.H5("🎯 最佳算法选择", className="text-warning mb-3"),
            dbc.Alert([
                html.I(className="fas fa-exclamation-triangle me-2"),
                f"未能确定最佳算法: {best_algo_info.get('selection_reason', '原因未知')}"
            ], color="warning")
        ], className="mb-4")
    
    # 创建参数配置表格
    params_table_data = []
    parameter_details = best_algo_info.get("parameter_details", {})
    
    for param_name, param_info in parameter_details.items():
        chinese_name = param_info.get("chinese_name", param_name)
        current_value = param_info.get("current_value", "未知")
        param_type = param_info.get("type", "")
        suggested_range = param_info.get("suggested_range", "")
        adjustment_tips = param_info.get("adjustment_tips", "")
        
        params_table_data.append(html.Tr([
            html.Td(param_name, className="fw-bold"),
            html.Td(chinese_name),
            html.Td([
                html.Span(str(current_value), className="badge bg-primary"),
                html.Small(f" ({param_type})", className="text-muted ms-1") if param_type else ""
            ]),
            html.Td(suggested_range, className="text-muted"),
            html.Td(adjustment_tips, className="small")
        ]))
    
    return html.Div([
        html.H5([
            html.I(className="fas fa-trophy me-2", style={"color": "#FFD700"}),
            "🎯 最佳算法及参数配置"
        ], className="text-success mb-3"),
        
        # 算法选择信息
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader([
                        html.I(className="fas fa-robot me-2"),
                        "最佳算法信息"
                    ]),
                    dbc.CardBody([
                        html.P([
                            html.Strong("算法名称: "),
                            html.Span(best_algo_info.get("best_algorithm_name", ""), className="badge bg-success ms-2")
                        ], className="mb-2"),
                        html.P([
                            html.Strong("预测值: "),
                            f"¥{best_algo_info.get('best_prediction_value', 0):,.2f}"
                        ], className="mb-2"),
                        html.P([
                            html.Strong("选择原因: "),
                            best_algo_info.get("selection_reason", "")
                        ], className="mb-2"),
                        html.P([
                            html.Strong("与比率法差异: "),
                            f"¥{best_algo_info.get('difference_to_ratio', 0):,.2f} ({best_algo_info.get('difference_percentage', 0):.2f}%)"
                        ], className="mb-0")
                    ])
                ])
            ], width=6),
            
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader([
                        html.I(className="fas fa-cogs me-2"),
                        "参数配置统计"
                    ]),
                    dbc.CardBody([
                        html.P([
                            html.Strong("参数总数: "),
                            html.Span(len(parameter_details), className="badge bg-info")
                        ], className="mb-2"),
                        html.P([
                            html.Strong("算法类型: "),
                            "机器学习回归算法"
                        ], className="mb-2"),
                        html.P([
                            html.Strong("参数来源: "),
                            "数据库配置表"
                        ], className="mb-2"),
                        html.P([
                            html.Strong("最后更新: "),
                            "动态加载"
                        ], className="mb-0")
                    ])
                ])
            ], width=6)
        ], className="mb-4"),
        
        # 详细参数配置表格
        html.Div([
            html.H6("📊 详细参数配置", className="text-secondary mb-3"),
            dbc.Table([
                html.Thead([
                    html.Tr([
                        html.Th("参数名", style={"width": "15%"}),
                        html.Th("中文名称", style={"width": "15%"}),
                        html.Th("当前值", style={"width": "20%"}),
                        html.Th("建议范围", style={"width": "20%"}),
                        html.Th("调优提示", style={"width": "30%"})
                    ])
                ]),
                html.Tbody(params_table_data)
            ], striped=True, bordered=True, hover=True, size="sm", className="mb-0")
        ], className="border rounded p-3 bg-light") if params_table_data else html.Div([
            html.P("暂无参数配置信息", className="text-muted text-center")
        ])
        
    ], className="mb-4 p-3 border rounded shadow-sm", style={"backgroundColor": "#f8f9fa"})
    
def register_price_prediction_callbacks(app):
    
    """注册价格预测页面的回调函数"""

    initialize_prediction_system('steel_cage')
    initialize_prediction_system('steel_lining')


    @app.callback(
        [Output("steel-reinforcement-parameter-modal", "is_open"),
        Output("steel-lining-parameter-modal2", "is_open"),
        Output("custom-mode-parameter-modal", "is_open"),
        Output("steel-cage-results-section", "style"),
        Output("steel-lining-results-section", "style"),
        Output("custom-mode-results-section", "style")],
        [Input("steel-cage-mode-div", "n_clicks"),
        Input("steel-lining-mode-div", "n_clicks"),
        Input("custom-mode-div", "n_clicks"),
        Input("close-steel-reinforcement-modal", "n_clicks"),
        Input("close-steel-lining-modal", "n_clicks"),
        Input("close-custom-mode-modal", "n_clicks"),
        Input("confirm-steel-reinforcement", "n_clicks"),
        Input("confirm-steel-lining", "n_clicks"),
        Input("confirm-custom-mode", "n_clicks")],
        [State("steel-reinforcement-parameter-modal", "is_open"),
        State("steel-lining-parameter-modal2", "is_open"),
        State("custom-mode-parameter-modal", "is_open")],
        prevent_initial_call=True
    )
    def toggle_mode_modals(
        n_clicks_steel_cage, n_clicks_steel_lining, n_clicks_custom_mode,
        n_clicks_close_steel_cage, n_clicks_close_steel_lining, n_clicks_close_custom_mode,
        n_clicks_confirm_steel_cage, n_clicks_confirm_steel_lining, n_clicks_confirm_custom_mode,
        is_open_steel_cage, is_open_steel_lining, is_open_custom_mode
    ):
        ctx = dash.callback_context
        if not ctx.triggered:
            raise PreventUpdate

        trigger_id = ctx.triggered[0]["prop_id"].split(".")[0]

        # 初始化所有结果显示样式为隐藏
        steel_cage_results_style = {"display": "none"}
        steel_lining_results_style = {"display": "none"}
        custom_mode_results_style = {"display": "none"}

        # 点击模式选择卡片 - 在这里添加状态检查逻辑
        if trigger_id == "steel-cage-mode-div":
            # 新增：记录模态窗口打开，后续的状态回调会自动处理字段状态
            logger.info("钢筋笼模式模态窗口打开，将自动检查字段状态")
            return True, False, False, steel_cage_results_style, steel_lining_results_style, custom_mode_results_style
            
        elif trigger_id == "steel-lining-mode-div":
            # 新增：记录模态窗口打开，后续的状态回调会自动处理字段状态
            logger.info("钢衬里模式模态窗口打开，将自动检查字段状态")
            return False, True, False, steel_cage_results_style, steel_lining_results_style, custom_mode_results_style
            
        elif trigger_id == "custom-mode-div":
            # 新增：记录模态窗口打开，后续的状态回调会自动处理字段状态
            logger.info("自定义模式模态窗口打开，将自动检查字段状态")
            return False, False, True, steel_cage_results_style, steel_lining_results_style, custom_mode_results_style

        # 点击确认按钮时，显示对应模式的结果，隐藏其他模式的结果
        if trigger_id == "confirm-steel-reinforcement":
            steel_cage_results_style = {"display": "block"}
            return False, is_open_steel_lining, is_open_custom_mode, steel_cage_results_style, steel_lining_results_style, custom_mode_results_style
        elif trigger_id == "confirm-steel-lining":
            steel_lining_results_style = {"display": "block"}
            return is_open_steel_cage, False, is_open_custom_mode, steel_cage_results_style, steel_lining_results_style, custom_mode_results_style
        elif trigger_id == "confirm-custom-mode":
            custom_mode_results_style = {"display": "block"}
            return is_open_steel_cage, is_open_steel_lining, False, steel_cage_results_style, steel_lining_results_style, custom_mode_results_style

        # 点击取消按钮时，隐藏所有结果
        if trigger_id == "close-steel-reinforcement-modal":
            return False, is_open_steel_lining, is_open_custom_mode, {"display": "none"}, {"display": "none"}, {"display": "none"}
        elif trigger_id == "close-steel-lining-modal":
            return is_open_steel_cage, False, is_open_custom_mode, {"display": "none"}, {"display": "none"}, {"display": "none"}
        elif trigger_id == "close-custom-mode-modal":
            return is_open_steel_cage, is_open_steel_lining, False, {"display": "none"}, {"display": "none"}, {"display": "none"}

        return is_open_steel_cage, is_open_steel_lining, is_open_custom_mode, steel_cage_results_style, steel_lining_results_style, custom_mode_results_style

    # 控制钢筋笼模式字段覆盖层的显示/隐藏 - 修复版本
    @app.callback(
        [Output(f"{field_id}-overlay", "style") for field_id in [
            "tower-crane-category-param", "tower-crane-1500", "heavy-tower-crane",
            "steel-production-category-param",
            "lifting-equipment-category-param", "steel-wire-a36-1500", "shackle-55t", "basket-bolt-3128",
            "sleeve-category-param", "straight-threaded-sleeve", "cone-steel-sleeve", "module-vertical-connection", 
            "steel-tonnage-category-param", "steel-price-category-param"
        ]] +
        [Output(f"{field_id}-warning", "style") for field_id in [
            "tower-crane-category-param", "tower-crane-1500", "heavy-tower-crane",
            "steel-production-category-param",
            "lifting-equipment-category-param", "steel-wire-a36-1500", "shackle-55t", "basket-bolt-3128",
            "sleeve-category-param", "straight-threaded-sleeve", "cone-steel-sleeve", "module-vertical-connection",
            "steel-tonnage-category-param", "steel-price-category-param"
        ]],
        [Input("steel-cage-field-status-store", "data")],
        prevent_initial_call=True
    )
    def update_steel_cage_field_overlays(field_status_data):
        """更新钢筋笼模式字段覆盖层的显示状态 - 与钢衬里模式保持一致"""
        if not field_status_data:
            # 模态窗口关闭时，返回默认状态
            field_count = 14  # 上面字段的数量
            return [False] * field_count + [{}]
        
        # ===== 新增：检查算法状态对界面的影响 =====
        try:
            from .data import CostPredictionSystem
            temp_system = CostPredictionSystem(mode='steel_cage')
            temp_system.load_algorithm_configs()
            temp_system.load_comprehensive_indicators_status()
            
            algo_capability = temp_system.check_algorithm_execution_capability()
            
            # 如果算法完全不可用，在日志中记录警告
            if not algo_capability.get('can_execute_ai', True):
                logger.warning(f"⚠️ 钢筋笼模式算法执行能力受限: {algo_capability.get('message', '')}")
                
            # 检查显示权限状态
            display_permissions = {}
            for method_key in ['ml_prediction_raw', 'ml_prediction_final', 'ratio_method_raw', 'ratio_method_final']:
                perm_info = temp_system.check_result_display_permission(method_key)
                display_permissions[method_key] = perm_info
                if not perm_info['can_display']:
                    logger.info(f"📋 钢筋笼模式 {method_key} 显示权限: {perm_info['message']}")
            
        except Exception as e:
            logger.warning(f"检查钢筋笼模式算法和显示状态时出错: {e}")
        
        # 获取字段状态
        field_status = get_field_status_for_mode('steel_cage')
        # 添加调试日志
        logger.info(f"收到字段状态数据: {field_status_data}")
        
        if not field_status_data:
            # 没有状态数据时，隐藏所有覆盖层
            field_count = 14
            hidden_style = {'display': 'none'}
            logger.info("字段状态数据为空，隐藏所有覆盖层")
            return [hidden_style] * (field_count * 2)  # 覆盖层 + 警告提示
        
        field_ids = [
            "tower-crane-category-param", "tower-crane-1500", "heavy-tower-crane",
            "steel-production-category-param", 
            "lifting-equipment-category-param", "steel-wire-a36-1500", "shackle-55t", "basket-bolt-3128",
            "sleeve-category-param", "straight-threaded-sleeve", "cone-steel-sleeve", "module-vertical-connection",
            "steel-tonnage-category-param", "steel-price-category-param"
        ]
        
        overlay_styles = []
        warning_styles = []
        
        for field_id in field_ids:
            field_info = field_status_data.get(field_id, {'status': 'enabled'})
            
            # 添加详细的调试信息
            logger.info(f"字段 {field_id} 的状态信息: {field_info}")
            
            if field_info['status'] == 'disabled':
                # 显示覆盖层和警告 - 与钢衬里模式样式保持一致
                overlay_style = {
                    'position': 'absolute',
                    'top': '0',
                    'left': '0', 
                    'right': '0',
                    'bottom': '0',
                    'backgroundColor': 'rgba(248, 249, 250, 0.9)',
                    'borderRadius': '4px',
                    'display': 'flex',
                    'alignItems': 'center',
                    'justifyContent': 'center',
                    'zIndex': '10'
                }
                warning_style = {'display': 'block'}
                logger.info(f"字段 {field_id} 被禁用，显示覆盖层")
            else:
                # 隐藏覆盖层和警告
                overlay_style = {'display': 'none'}
                warning_style = {'display': 'none'}
                logger.info(f"字段 {field_id} 已启用，隐藏覆盖层")
            
            overlay_styles.append(overlay_style)
            warning_styles.append(warning_style)
        
        logger.info(f"返回 {len(overlay_styles)} 个覆盖层样式和 {len(warning_styles)} 个警告样式")
        return overlay_styles + warning_styles

    # 控制钢衬里模式字段覆盖层的显示/隐藏 - 与钢筋笼模式保持一致
    @app.callback(
        [Output(f"{field_id}-overlay", "style") for field_id in [
            "assembly-site-category-param", "fixture-making-category-param",
            "steel-support-embedded-category-param", "steel-support-concrete-chiseling", 
            "steel-support-concrete-backfill", "steel-support-installation", "steel-support-depreciation",
            "buttress-category-param", "buttress-installation", "buttress-removal", "buttress-component-depreciation",
            "walkway-platform-category-param", "walkway-platform-manufacturing", "walkway-platform-erection", "walkway-platform-removal",
            "steel-grid-beam-category-param", "steel-grid-manufacturing", "steel-grid-installation", "steel-grid-removal",
            "steel-lining-measures-category-param"
        ]] +
        [Output(f"{field_id}-warning", "style") for field_id in [
            "assembly-site-category-param", "fixture-making-category-param",
            "steel-support-embedded-category-param", "steel-support-concrete-chiseling",
            "steel-support-concrete-backfill", "steel-support-installation", "steel-support-depreciation",
            "buttress-category-param", "buttress-installation", "buttress-removal", "buttress-component-depreciation", 
            "walkway-platform-category-param", "walkway-platform-manufacturing", "walkway-platform-erection", "walkway-platform-removal",
            "steel-grid-beam-category-param", "steel-grid-manufacturing", "steel-grid-installation", "steel-grid-removal",
            "steel-lining-measures-category-param"
        ]],
        [Input("steel-lining-field-status-store", "data")],
        prevent_initial_call=True
    )
    def update_steel_lining_field_overlays(field_status_data):
        """更新钢衬里模式字段覆盖层的显示状态 - 与钢筋笼模式保持一致"""
        if not field_status_data:
            # 模态窗口关闭时，返回默认状态
            field_count = 14  # 上面字段的数量
            return [False] * field_count + [{}]
        
        # ===== 新增：检查算法状态对界面的影响 =====
        try:
            from .data import CostPredictionSystem
            temp_system = CostPredictionSystem(mode='steel_lining')
            temp_system.load_algorithm_configs()
            temp_system.load_comprehensive_indicators_status()
            
            algo_capability = temp_system.check_algorithm_execution_capability()
            
            # 如果算法完全不可用，在日志中记录警告
            if not algo_capability.get('can_execute_ai', True):
                logger.warning(f"⚠️ 钢筋笼模式算法执行能力受限: {algo_capability.get('message', '')}")
                
            # 检查显示权限状态
            display_permissions = {}
            for method_key in ['ml_prediction_raw', 'ml_prediction_final', 'ratio_method_raw', 'ratio_method_final']:
                perm_info = temp_system.check_result_display_permission(method_key)
                display_permissions[method_key] = perm_info
                if not perm_info['can_display']:
                    logger.info(f"📋 钢筋笼模式 {method_key} 显示权限: {perm_info['message']}")
            
        except Exception as e:
            logger.warning(f"检查钢筋笼模式算法和显示状态时出错: {e}")
        
        # 获取字段状态
        field_status = get_field_status_for_mode('steel_cage')
        # 添加调试日志
        logger.info(f"收到钢衬里字段状态数据: {field_status_data}")
        
        if not field_status_data:
            # 没有状态数据时，隐藏所有覆盖层
            field_count = 20
            hidden_style = {'display': 'none'}
            logger.info("钢衬里字段状态数据为空，隐藏所有覆盖层")
            return [hidden_style] * (field_count * 2)  # 覆盖层 + 警告提示
        
        field_ids = [
            "assembly-site-category-param", "fixture-making-category-param",
            "steel-support-embedded-category-param", "steel-support-concrete-chiseling",
            "steel-support-concrete-backfill", "steel-support-installation", "steel-support-depreciation",
            "buttress-category-param", "buttress-installation", "buttress-removal", "buttress-component-depreciation", 
            "walkway-platform-category-param", "walkway-platform-manufacturing", "walkway-platform-erection", "walkway-platform-removal",
            "steel-grid-beam-category-param", "steel-grid-manufacturing", "steel-grid-installation", "steel-grid-removal",
            "steel-lining-measures-category-param"
        ]
        
        overlay_styles = []
        warning_styles = []
        
        for field_id in field_ids:
            field_info = field_status_data.get(field_id, {'status': 'enabled'})
            
            # 添加详细的调试信息
            logger.info(f"钢衬里字段 {field_id} 的状态信息: {field_info}")
            
            if field_info['status'] == 'disabled':
                # 显示覆盖层和警告 - 与钢筋笼模式样式保持一致
                overlay_style = {
                    'position': 'absolute',
                    'top': '0',
                    'left': '0', 
                    'right': '0',
                    'bottom': '0',
                    'backgroundColor': '#f8f9fa',  # 灰色背景
                    'border': '1px solid #dee2e6',
                    'borderRadius': '0.375rem',
                    'display': 'flex',
                    'alignItems': 'center',
                    'justifyContent': 'center',
                    'zIndex': '10',
                    'cursor': 'not-allowed'
                }
                warning_style = {'display': 'block'}
                logger.info(f"钢衬里字段 {field_id} 被禁用，显示覆盖层")
            else:
                # 隐藏覆盖层和警告
                overlay_style = {'display': 'none'}
                warning_style = {'display': 'none'}
                logger.info(f"钢衬里字段 {field_id} 已启用，隐藏覆盖层")
            
            overlay_styles.append(overlay_style)
            warning_styles.append(warning_style)
        
        logger.info(f"返回钢衬里 {len(overlay_styles)} 个覆盖层样式和 {len(warning_styles)} 个警告样式")
        return overlay_styles + warning_styles

    @app.callback(
        [Output("steel-reinforcement-calculation-result5", "children"),
         Output('cost-comparison-container2', 'style'),
         Output('cost-comparison-table2', 'children'),
         Output("steel-cage-report-data", "data")],
        Input("confirm-steel-reinforcement", "n_clicks"),
        [State('tower-crane-category-param', 'value'),
         State('steel-production-category-param', 'value'),
         State('lifting-equipment-category-param', 'value'),
         State('sleeve-category-param', 'value'),
         State('steel-tonnage-category-param', 'value'),
         State('steel-price-category-param', 'value'),
         State('tower-crane-1500', 'value'),
         State('heavy-tower-crane', 'value'),
         State('steel-wire-a36-1500', 'value'),
         State('shackle-55t', 'value'),
         State('basket-bolt-3128', 'value'),
         State('straight-threaded-sleeve', 'value'),
         State('cone-steel-sleeve', 'value'),
         State('module-vertical-connection', 'value')],
        prevent_initial_call=True
    )
    def handle_steel_reinforcement_with_ml_prediction(confirm_clicks,
                                                     tower_crane_qty_category, steel_production_qty_category,
                                                     lifting_equipment_qty_category, sleeve_qty_category,
                                                     steel_tonnage_category, measures_qty_category,
                                                     tower_crane_1500, heavy_tower_crane,
                                                     steel_wire_a36, shackle_55t, basket_bolt,
                                                     straight_threaded_sleeve, cone_steel_sleeve,
                                                     module_vertical_connection):
        if not confirm_clicks:
            raise PreventUpdate

        # 使用MySQL配置替换原来的SQLite路径
        # 这些配置应该根据实际的MySQL表结构进行调整
        db1_table = 'price_baseline_1'
        db1_proj_col = 'project'
        db1_param_col = 'parameter_category'
        db1_qty_col = 'modular_material_quantity'
        db1_target_val = '钢筋吨数'

        db2_table = 'key_factors_1'
        db2_proj_col = 'project_id'
        db2_target_sum_col = '钢筋总吨数'

        output_target_table = 'final_project_summary1'
        output_if_exists = 'replace'

        try:
            success = process_and_update_databases(
                table1_name=db1_table,
                project_col_db1=db1_proj_col,
                param_category_col_db1=db1_param_col,
                material_quantity_col_db1=db1_qty_col,
                target_param_category_value=db1_target_val,
                table2_name=db2_table,
                project_col_db2=db2_proj_col,
                target_sum_col_db2=db2_target_sum_col,
                output_table_name=output_target_table,
                if_exists_behavior=output_if_exists
                )
            logger.info("钢筋笼模式的关键因素数据库已更新。")

            global global_predictors
            global_predictors['steel_cage'] = None
            init_success = initialize_prediction_system('steel_cage')
            if not init_success:
                error_message = "机器学习预测系统重新初始化失败，请检查数据或配置（钢筋笼模式）。"
                return html.Div(dbc.Alert(error_message, color="danger")), {"display": "block"}, "", {}
        except Exception as e:
            logger.error(f"更新数据或重新初始化钢筋笼模式预测系统异常: {e}", exc_info=True)
            
            # 创建详细的错误报告
            error_report = create_detailed_error_report(e, "钢筋笼", "system_initialization")
            error_display = create_error_display_component(error_report)
            
            # 生成错误状态的报告数据
            error_report_data = {
                "模式": "钢筋笼施工模式",
                "生成时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "状态": "系统错误",
                "错误信息": error_report,
                "工程量数据": quantities,
                "预测结果": {"error": "系统初始化失败"},
                "措施费": measures_cost_value
            }
            
            return error_display, {"display": "block"}, "", error_report_data
        quantities = collect_user_quantities(
            tower_crane_qty_category, steel_production_qty_category,
            lifting_equipment_qty_category, sleeve_qty_category,
            steel_tonnage_category, measures_qty_category,
            tower_crane_1500, heavy_tower_crane, steel_wire_a36,
            shackle_55t, basket_bolt, straight_threaded_sleeve,
            cone_steel_sleeve, module_vertical_connection, safe_float
        )

        measures_cost_value = quantities.get('措施费工程量', 0.0)
        ml_prediction_results = None

        if global_predictors['steel_cage'] and global_predictors['steel_cage'].is_trained:
            try:
                user_inputs_for_ml = prepare_ml_inputs(quantities)
                if user_inputs_for_ml:
                    ml_results_from_system = global_predictors['steel_cage'].predict(user_inputs_for_ml, {})

                    # 【修复】只为数值类型的预测结果添加措施费
                    if "机器学习预测结果" in ml_results_from_system and ml_results_from_system["机器学习预测结果"]:
                        for model_key in ml_results_from_system["机器学习预测结果"]:
                            prediction_value = ml_results_from_system["机器学习预测结果"][model_key]
                            # 只对数值类型的预测结果添加措施费，跳过停用算法的字典状态
                            if prediction_value is not None and isinstance(prediction_value, (int, float)):
                                ml_results_from_system["机器学习预测结果"][model_key] += measures_cost_value

                    # 【修复】比率法预测也需要检查
                    if ml_results_from_system.get("比率法预测总价") is not None and isinstance(ml_results_from_system.get("比率法预测总价"), (int, float)):
                        ml_results_from_system["比率法预测总价"] += measures_cost_value

                    ml_prediction_results = ml_results_from_system
                    logger.info(f"✓ 钢筋笼机器学习预测完成 (已包含措施费): {user_inputs_for_ml}")
                else:
                    ml_prediction_results = {"error": "无法从输入中提取有效的工程量参数 for 钢筋笼模式"}
            except Exception as e:
                ml_prediction_results = {"error": f"机器学习预测失败 for 钢筋笼模式: {e}"}
                logger.error(f"✗ 钢筋笼机器学习预测异常: {e}", exc_info=True)
        else:
            ml_prediction_results = {"error": "机器学习预测系统未就绪或训练失败（钢筋笼模式）"}

        simple_confirmation = html.Div([
            dbc.Alert([
                html.I(className="fas fa-check-circle me-2"),
                html.Strong("预测计算完成！"),
                html.Br(),
                "智能预测结果已显示在下方表格中，请查看详细分析。"
            ], color="success", className="mb-0")
        ])

        # 修改后的代码：处理新的预测结果结构
        detailed_results_table = create_prediction_results_table(
            quantities, 
            ml_prediction_results, 
            measures_cost_value
        )
        results_container_style = {"display": "block"}

        # 生成用户友好的确认消息
        confirmation_message = generate_prediction_confirmation_message(ml_prediction_results, "钢筋笼")
        
        # 新增：保存报告数据到Store - 适配新的数据结构
        report_data = {
            "模式": "钢筋笼施工模式",
            "生成时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "工程量数据": quantities,
            "预测结果": ml_prediction_results,
            "措施费": measures_cost_value,
            # 新增：状态信息
            "算法执行状态": ml_prediction_results.get("算法执行状态", {}),
            "预测方法状态": ml_prediction_results.get("预测方法状态", {}),
            "显示权限状态": ml_prediction_results.get("显示权限状态", {}),
            # 新增：主要预测结果（用于报告）
            "主要AI预测": extract_main_prediction_result(ml_prediction_results.get("机器学习预测结果")),
            "主要比率法预测": extract_main_prediction_result(ml_prediction_results.get("比率法预测总价")),
            # 【新增】最佳算法信息
            "最佳算法信息": ml_prediction_results.get("最佳算法信息", {})
        }

        return confirmation_message, results_container_style, detailed_results_table, report_data


    @app.callback(
        [Output('steel-lining-calculation-result-output', 'children'),
         Output('steel-lining-cost-comparison-container', 'style'),
         Output('steel-lining-cost-comparison-table', 'children'),
         Output("steel-lining-report-data", "data")],
        Input("confirm-steel-lining", "n_clicks"),
        [State("assembly-site-category-param", "value"),
         State("fixture-making-category-param", "value"),
         State("steel-support-embedded-category-param", "value"),
         State("steel-support-concrete-chiseling", "value"),
         State("steel-support-concrete-backfill", "value"),
         State("steel-support-installation", "value"),
         State("steel-support-depreciation", "value"),
         State("buttress-category-param", "value"),
         State("buttress-installation", "value"),
         State("buttress-removal", "value"),
         State("buttress-component-depreciation", "value"),
         State("walkway-platform-category-param", "value"),
         State("walkway-platform-manufacturing", "value"),
         State("walkway-platform-erection", "value"),
         State("walkway-platform-removal", "value"),
         State("steel-grid-beam-category-param", "value"),
         State("steel-grid-manufacturing", "value"),
         State("steel-grid-installation", "value"),
         State("steel-grid-removal", "value"),
         State("steel-lining-measures-category-param", "value")
        ],
        prevent_initial_call=True
    )
    def handle_steel_lining_prediction(confirm_clicks,
                                        assembly_site_qty, fixture_making_qty, steel_support_embedded_category_qty,
                                        steel_support_concrete_chiseling_qty, steel_support_concrete_backfill_qty,
                                        steel_support_installation_qty, steel_support_depreciation_qty,
                                        buttress_category_qty, buttress_installation_qty, buttress_removal_qty,
                                        buttress_component_depreciation_qty, walkway_platform_category_qty,
                                        walkway_platform_manufacturing_qty, walkway_platform_erection_qty,
                                        walkway_platform_removal_qty, steel_grid_beam_category_qty,
                                        steel_grid_manufacturing_qty, steel_grid_installation_qty,
                                        steel_grid_removal_qty, measures_cost_qty):
        if not confirm_clicks:
            raise PreventUpdate

        # MySQL配置替换SQLite路径
        db1_table = 'price_baseline_2'
        db1_proj_col = 'project'
        db1_param_col = 'parameter_category'
        db1_qty_col = 'modular_material_quantity'
        db1_target_val = '拼装场地'

        db2_table = 'key_factors_2'
        db2_proj_col = 'project_id'
        db2_target_sum_col = '拼装场地费用'

        output_target_table = 'final_project_summary2'
        output_if_exists = 'replace'
        
        try:
            success = process_and_update_databases(
                table1_name=db1_table,
                project_col_db1=db1_proj_col,
                param_category_col_db1=db1_param_col,
                material_quantity_col_db1=db1_qty_col,
                target_param_category_value=db1_target_val,
                table2_name=db2_table,
                project_col_db2=db2_proj_col,
                target_sum_col_db2=db2_target_sum_col,
                output_table_name=output_target_table,
                if_exists_behavior=output_if_exists
            )
            logger.info("钢衬里模式的关键因素数据库已更新。")

            global global_predictors
            global_predictors['steel_lining'] = None
            init_success = initialize_prediction_system('steel_lining')
            if not init_success:
                error_message = "机器学习预测系统重新初始化失败，请检查数据或配置（钢衬里模式）。"
                return html.Div(dbc.Alert(error_message, color="danger")), {"display": "block"}, "", {}
        except Exception as e:
            logger.error(f"更新数据或重新初始化钢衬里模式预测系统异常: {e}", exc_info=True)
            return html.Div(dbc.Alert(f"预测失败：更新数据或模型训练出错 - {e}", color="danger")), {"display": "block"}, "", {}

        quantities = {}

        quantities['拼装场地总工程量'] = safe_float(assembly_site_qty)
        quantities['制作胎具总工程量'] = safe_float(fixture_making_qty)

        if steel_support_embedded_category_qty is not None and str(steel_support_embedded_category_qty).strip() != "":
            quantities['钢支墩埋件总工程量'] = safe_float(steel_support_embedded_category_qty)
        else:
            sub_total = safe_float(steel_support_concrete_chiseling_qty) + safe_float(steel_support_concrete_backfill_qty) + \
                        safe_float(steel_support_installation_qty) + safe_float(steel_support_depreciation_qty)
            quantities['钢支墩埋件总工程量'] = sub_total

        if buttress_category_qty is not None and str(buttress_category_qty).strip() != "":
            quantities['扶壁柱总工程量'] = safe_float(buttress_category_qty)
        else:
            sub_total = safe_float(buttress_installation_qty) + safe_float(buttress_removal_qty) + safe_float(buttress_component_depreciation_qty)
            quantities['扶壁柱总工程量'] = sub_total

        if walkway_platform_category_qty is not None and str(walkway_platform_category_qty).strip() != "":
            quantities['走道板操作平台总工程量'] = safe_float(walkway_platform_category_qty)
        else:
            sub_total = safe_float(walkway_platform_manufacturing_qty) + safe_float(walkway_platform_erection_qty) + safe_float(walkway_platform_removal_qty)
            quantities['走道板操作平台总工程量'] = sub_total

        if steel_grid_beam_category_qty is not None and str(steel_grid_beam_category_qty).strip() != "":
            quantities['钢网梁总工程量'] = safe_float(steel_grid_beam_category_qty)
        else:
            sub_total = safe_float(steel_grid_manufacturing_qty) + safe_float(steel_grid_installation_qty) + safe_float(steel_grid_removal_qty)
            quantities['钢网梁总工程量'] = sub_total

        measures_cost_value = safe_float(measures_cost_qty)

        ml_prediction_results = None

        if global_predictors['steel_lining'] and global_predictors['steel_lining'].is_trained:
            try:
                ml_results_from_system = global_predictors['steel_lining'].predict(quantities, {})

                # 【修复】只为数值类型的预测结果添加措施费
                if "机器学习预测结果" in ml_results_from_system and ml_results_from_system["机器学习预测结果"]:
                    for model_key in ml_results_from_system["机器学习预测结果"]:
                        prediction_value = ml_results_from_system["机器学习预测结果"][model_key]
                        # 只对数值类型的预测结果添加措施费，跳过停用算法的字典状态
                        if prediction_value is not None and isinstance(prediction_value, (int, float)):
                            ml_results_from_system["机器学习预测结果"][model_key] += measures_cost_value
                            
                # 【修复】比率法预测也需要检查
                if ml_results_from_system.get("比率法预测总价") is not None and isinstance(ml_results_from_system.get("比率法预测总价"), (int, float)):
                    ml_results_from_system["比率法预测总价"] += measures_cost_value

                ml_prediction_results = ml_results_from_system
                logger.info(f"✓ 钢衬里机器学习预测完成 (已包含措施费): {quantities}")
            except Exception as e:
                ml_prediction_results = {"error": f"机器学习预测失败 for 钢衬里模式: {e}"}
                logger.error(f"✗ 钢衬里机器学习预测异常: {e}", exc_info=True)
        else:
            ml_prediction_results = {"error": "机器学习预测系统未就绪或训练失败（钢衬里模式）"}

        simple_confirmation = html.Div([
            dbc.Alert([
                html.I(className="fas fa-check-circle me-2"),
                html.Strong("预测计算完成！"),
                html.Br(),
                "智能预测结果已显示在下方表格中，请查看详细分析。"
            ], color="success", className="mb-0")
        ])

        # 修改后的代码：处理新的预测结果结构
        detailed_results_table = create_steel_lining_prediction_results_table(
            quantities, 
            ml_prediction_results, 
            measures_cost_value
        )
        results_container_style = {"display": "block"}

        # 生成用户友好的确认消息
        confirmation_message = generate_prediction_confirmation_message(ml_prediction_results, "钢衬里")

        # 新增：保存报告数据到Store - 适配新的数据结构
        report_data = {
            "模式": "钢衬里施工模式",
            "生成时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "工程量数据": quantities,
            "预测结果": ml_prediction_results,
            "措施费": measures_cost_value,
            # 新增：状态信息
            "算法执行状态": ml_prediction_results.get("算法执行状态", {}),
            "预测方法状态": ml_prediction_results.get("预测方法状态", {}),
            "显示权限状态": ml_prediction_results.get("显示权限状态", {}),
            # 新增：主要预测结果（用于报告）
            "主要AI预测": extract_main_prediction_result(ml_prediction_results.get("机器学习预测结果")),
            "主要比率法预测": extract_main_prediction_result(ml_prediction_results.get("比率法预测总价")),
            # 【新增】最佳算法信息
            "最佳算法信息": ml_prediction_results.get("最佳算法信息", {})
        }
        
        return confirmation_message, results_container_style, detailed_results_table, report_data


    @app.callback(
        Output("price-modification-modal", "is_open", allow_duplicate=True),
        [Input("open-price-modification-modal-button", "n_clicks"),
         Input("close-price-modification-modal-button", "n_clicks")],
        [State("price-modification-modal", "is_open")],
        prevent_initial_call=True
    )
    def toggle_price_modification_modal(open_clicks, close_clicks, is_open):
        if open_clicks or close_clicks:
            return not is_open
        return is_open

    @app.callback(
        [Output("price-modification-table", "data"),
         Output("price-modification-table", "columns")],
        [Input("price-modification-modal", "is_open"),
         Input("price-modification-mode-radio", "value")],
        prevent_initial_call=True
    )
    def load_price_modification_data(is_open, selected_mode):
        if not is_open:
            raise PreventUpdate

        conn = None
        try:
            # 使用MySQL连接替换SQLite
            conn = get_connection()
            
            if selected_mode == "钢筋笼施工模式":
                query = 'SELECT * FROM `price_baseline_1`'
            elif selected_mode == "钢衬里施工模式":
                query = 'SELECT * FROM `price_baseline_2`'
            else:
                return [], [{"name": "错误", "id": "错误"}]

            df = pd.read_sql_query(query, conn)

            if df.empty:
                return [], []

            editable_columns_map = {
                "钢筋笼施工模式": [
                    "modular_labor_unit_price", "modular_material_unit_price", "modular_machinery_unit_price",
                ],
                "钢衬里施工模式": [
                    "modular_labor_unit_price", "modular_material_unit_price", "modular_machinery_unit_price",
                ]
            }
            editable_cols = editable_columns_map.get(selected_mode, [])

            cols_to_display = [
                "project", "sequence_number", "parameter_category", "unit",
                "modular_labor_unit_price", "modular_labor_quantity", "modular_labor_total",
                "modular_material_unit_price", "modular_material_quantity", "modular_material_total",
                "modular_machinery_unit_price", "modular_machinery_quantity", "modular_machinery_total",
                "total_price"
            ]

            actual_cols_in_db = [col for col in cols_to_display if col in df.columns]
            df_display = df[actual_cols_in_db].copy()

            columns_for_table = []
            for col_name in actual_cols_in_db:
                col_def = {
                    "name": col_name.replace("模块化施工", "模块化").replace("直接施工", "直接") if "施工" in col_name else col_name,
                    "id": col_name,
                    "editable": (col_name in editable_cols)
                }
                if any(k in col_name for k in ["单价", "合价", "工程量", "总价"]):
                    col_def["type"] = "numeric"
                    col_def["format"] = {"specifier": ".2f"}
                columns_for_table.append(col_def)

            for col in editable_cols:
                if col in df_display.columns:
                    df_display[col] = pd.to_numeric(df_display[col], errors='coerce').fillna(0)
            df_display = df_display.fillna('')

            return df_display.to_dict('records'), columns_for_table

        except mysql.connector.Error as e:
            logger.error(f"MySQL database error loading data for {selected_mode} mode: {e}", exc_info=True)
            error_data = [{"错误": f"数据库加载错误: {e}", "建议": "请检查MySQL数据库连接和表名"}]
            error_columns = [{"name": "错误", "id": "错误"}, {"name": "建议", "id": "建议"}]
            return error_data, error_columns
        except Exception as e:
            logger.error(f"Error loading data for price modification ({selected_mode} mode): {e}", exc_info=True)
            error_data = [{"错误": f"加载数据时出错: {e}", "建议": "请检查数据处理逻辑"}]
            error_columns = [{"name": "错误", "id": "错误"}, {"name": "建议", "id": "建议"}]
            return error_data, error_columns
        finally:
            if conn:
                conn.close()

    @app.callback(
        Output("price-modification-feedback", "children", allow_duplicate=True),
        [Input("price-modification-table", "data_timestamp")],
        [State("price-modification-table", "data"),
         State("price-modification-table", "data_previous"),
         State("price-modification-mode-radio", "value")],
        prevent_initial_call=True
    )
    def save_modified_prices(timestamp, current_data, previous_data, current_mode):
        if previous_data is None or current_data is None:
            raise PreventUpdate

        feedback_messages = []
        conn = None

        try:
            # 使用MySQL连接替换SQLite
            conn = get_connection()
            cursor = conn.cursor()

            # 修改为英文列名
            editable_db_cols = [
                "modular_labor_unit_price", "modular_material_unit_price", "modular_machinery_unit_price",
                "modular_labor_total", "modular_material_total", "modular_machinery_total", "total_price",
                "modular_labor_quantity", "modular_material_quantity", "modular_machinery_quantity",
            ]

            changes_detected = False
            for i, row_data in enumerate(current_data):
                if i >= len(previous_data):
                    continue

                prev_row_data = previous_data[i]

                record_id = row_data.get("sequence_number")
                if record_id is None:
                    feedback_messages.append(dbc.Alert(f"警告: 发现缺少 'sequence_number' 的行，无法更新。", color="warning", duration=4000))
                    continue

                for db_col_name in editable_db_cols:
                    current_val = row_data.get(db_col_name)
                    prev_val = prev_row_data.get(db_col_name)

                    try:
                        current_val_processed = float(current_val) if current_val is not None and str(current_val).strip() != '' else None
                    except ValueError:
                        current_val_processed = None

                    try:
                        prev_val_processed = float(prev_val) if prev_val is not None and str(prev_val).strip() != '' else None
                    except ValueError:
                        prev_val_processed = None

                    if current_val_processed != prev_val_processed:
                        changes_detected = True
                        update_value = current_val_processed

                        try:
                            if current_mode == "钢筋笼施工模式":
                                cursor.execute(f'UPDATE `price_baseline_1` SET `{db_col_name}` = %s WHERE `sequence_number` = %s', (update_value, record_id))
                            elif current_mode == "钢衬里施工模式":
                                cursor.execute(f'UPDATE `price_baseline_2` SET `{db_col_name}` = %s WHERE `sequence_number` = %s', (update_value, record_id))
                            
                            feedback_messages.append(dbc.Alert(f"记录 {record_id} (模式: {current_mode}) 的 '{db_col_name}' 已更新。", color="success", duration=3000))
                        except mysql.connector.Error as err:
                            feedback_messages.append(dbc.Alert(f"更新记录 {record_id} 的 '{db_col_name}' 失败: {err}", color="danger", duration=5000))
                            conn.rollback()
                            raise err

            if changes_detected:
                conn.commit()
                feedback_messages.insert(0, dbc.Alert(f"数据已成功保存到 '{current_mode}' MySQL数据库。", color="success"))
            else:
                feedback_messages.append(dbc.Alert("没有检测到有效的数据更改以保存。", color="info", duration=3000))

        except mysql.connector.Error as e:
            feedback_messages.append(dbc.Alert(f"MySQL数据库操作失败: {e}", color="danger"))
            if conn:
                conn.rollback()
        except Exception as e:
            feedback_messages.append(dbc.Alert(f"保存数据时发生意外错误: {e}", color="danger"))
            if conn:
                conn.rollback()
        finally:
            if conn:
                conn.close()

        return feedback_messages if feedback_messages else None



    @app.callback(
        [Output("custom-params-feedback", "children"),
         Output("custom-param-name", "value"),
         Output("custom-param-quantity", "value"),
         Output("custom-quantity-ratio", "value"),
         Output("custom-price-amount", "value"),
         Output("custom-price-ratio", "value"),
         Output("custom-key-factor", "value")],
        Input("add-custom-param-btn", "n_clicks"),
        [State("custom-param-name", "value"),
         State("custom-param-quantity", "value"),
         State("custom-quantity-ratio", "value"),
         State("custom-price-amount", "value"),
         State("custom-price-ratio", "value"),
         State("custom-key-factor", "value")],
        prevent_initial_call=True
    )
    def add_custom_parameter(n_clicks, param_name, param_quantity, quantity_ratio, 
                           price_amount, price_ratio, key_factor):
        """添加自定义参数到数据库"""
        if not n_clicks:
            raise PreventUpdate
        
        # 验证输入数据
        if not param_name or param_name.strip() == "":
            return [
                dbc.Alert("参数名称不能为空！", color="danger", duration=3000),
                dash.no_update, dash.no_update, dash.no_update, 
                dash.no_update, dash.no_update, dash.no_update
            ]
        
        try:
            param_quantity = safe_float(param_quantity)
            quantity_ratio = safe_float(quantity_ratio) 
            price_amount = safe_float(price_amount)
            price_ratio = safe_float(price_ratio)
            
            # 验证占比范围
            if quantity_ratio < 0 or quantity_ratio > 100:
                return [
                    dbc.Alert("工程量占比必须在0-100之间！", color="danger", duration=3000),
                    dash.no_update, dash.no_update, dash.no_update, 
                    dash.no_update, dash.no_update, dash.no_update
                ]
            
            if price_ratio < 0 or price_ratio > 100:
                return [
                    dbc.Alert("价格占比必须在0-100之间！", color="danger", duration=3000),
                    dash.no_update, dash.no_update, dash.no_update, 
                    dash.no_update, dash.no_update, dash.no_update
                ]
            
        except (ValueError, TypeError):
            return [
                dbc.Alert("请输入有效的数值！", color="danger", duration=3000),
                dash.no_update, dash.no_update, dash.no_update, 
                dash.no_update, dash.no_update, dash.no_update
            ]
        
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            # 插入新参数
            insert_query = """
            INSERT INTO `price_custom_parameters` 
            (`param_name`, `param_quantity`, `quantity_ratio`, `price_amount`, `price_ratio`, `key_factor`, `user_id`) 
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            
            cursor.execute(insert_query, (
                param_name.strip(),
                param_quantity,
                quantity_ratio,
                price_amount,
                price_ratio,
                key_factor.strip() if key_factor else '',
                'default'  # 默认用户ID
            ))
            
            conn.commit()
            
            success_alert = dbc.Alert([
                html.I(className="fas fa-check-circle me-2"),
                f"参数 '{param_name}' 添加成功！"
            ], color="success", duration=3000)
            
            # 清空输入框
            return [success_alert, "", None, None, None, None, ""]
            
        except mysql.connector.Error as e:
            error_alert = dbc.Alert([
                html.I(className="fas fa-exclamation-triangle me-2"),
                f"数据库操作失败: {str(e)}"
            ], color="danger", duration=5000)
            
            return [
                error_alert,
                dash.no_update, dash.no_update, dash.no_update, 
                dash.no_update, dash.no_update, dash.no_update
            ]
            
        except Exception as e:
            error_alert = dbc.Alert([
                html.I(className="fas fa-exclamation-triangle me-2"),
                f"添加参数时发生错误: {str(e)}"
            ], color="danger", duration=5000)
            
            return [
                error_alert,
                dash.no_update, dash.no_update, dash.no_update, 
                dash.no_update, dash.no_update, dash.no_update
            ]
            
        finally:
            if conn:
                conn.close()

    @app.callback(
        [Output("custom-params-table", "data"),
         Output("custom-params-feedback", "children", allow_duplicate=True)],
        [Input("refresh-custom-table-btn", "n_clicks"),
         Input("add-custom-param-btn", "n_clicks")],  # 添加参数后也自动刷新表格
        prevent_initial_call=True
    )
    def refresh_custom_params_table(refresh_clicks, add_clicks):
        """刷新自定义参数表格数据"""
        # 无论是点击刷新按钮还是添加参数，都执行刷新
        if not refresh_clicks and not add_clicks:
            raise PreventUpdate
        
        conn = None
        try:
            conn = get_connection()
            
            # 查询所有有效的自定义参数
            query = """
            SELECT 
                `id`,
                `param_name`,
                `param_quantity`,
                `quantity_ratio`,
                `price_amount`,
                `price_ratio`,
                `key_factor`,
                DATE_FORMAT(`create_time`, '%Y-%m-%d %H:%i:%s') as create_time
            FROM `price_custom_parameters`
            WHERE `status` = 1 AND `user_id` = 'default'
            ORDER BY `create_time` DESC
            """
            
            df = pd.read_sql_query(query, conn)
            
            if df.empty:
                return [], dbc.Alert([
                    html.I(className="fas fa-info-circle me-2"),
                    "暂无自定义参数数据，请先添加参数。"
                ], color="info", duration=3000)
            
            # 转换数据格式供DataTable使用
            table_data = df.to_dict('records')
            
            # 格式化数值显示
            for row in table_data:
                row['param_quantity'] = round(float(row['param_quantity']), 2) if row['param_quantity'] else 0.0
                row['quantity_ratio'] = round(float(row['quantity_ratio']), 2) if row['quantity_ratio'] else 0.0
                row['price_amount'] = round(float(row['price_amount']), 2) if row['price_amount'] else 0.0
                row['price_ratio'] = round(float(row['price_ratio']), 2) if row['price_ratio'] else 0.0
            
            success_alert = dbc.Alert([
                html.I(className="fas fa-sync-alt me-2"),
                f"表格已更新，共加载 {len(table_data)} 条参数记录。"
            ], color="success", duration=3000)
            
            return table_data, success_alert
            
        except mysql.connector.Error as e:
            error_alert = dbc.Alert([
                html.I(className="fas fa-exclamation-triangle me-2"),
                f"数据库查询失败: {str(e)}"
            ], color="danger", duration=5000)
            
            return [], error_alert
            
        except Exception as e:
            error_alert = dbc.Alert([
                html.I(className="fas fa-exclamation-triangle me-2"),
                f"加载数据时发生错误: {str(e)}"
            ], color="danger", duration=5000)
            
            return [], error_alert
            
        finally:
            if conn:
                conn.close()

    @app.callback(
        Output("custom-params-feedback", "children", allow_duplicate=True),
        [Input("custom-params-table", "data_timestamp")],
        [State("custom-params-table", "data"),
         State("custom-params-table", "data_previous")],
        prevent_initial_call=True
    )
    def save_table_edits(timestamp, current_data, previous_data):
        """保存表格中的编辑内容到数据库"""
        if not timestamp or not current_data or not previous_data:
            raise PreventUpdate
        
        if len(current_data) != len(previous_data):
            raise PreventUpdate
        
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            changes_count = 0
            error_messages = []
            
            # 比较数据变化并更新
            for i, (current_row, previous_row) in enumerate(zip(current_data, previous_data)):
                row_id = current_row.get('id')
                if not row_id:
                    continue
                
                # 检查每个字段是否有变化
                fields_to_check = ['param_name', 'param_quantity', 'quantity_ratio', 
                                 'price_amount', 'price_ratio', 'key_factor']
                
                updates = {}
                for field in fields_to_check:
                    current_val = current_row.get(field)
                    previous_val = previous_row.get(field)
                    
                    if current_val != previous_val:
                        # 验证数值字段
                        if field in ['param_quantity', 'quantity_ratio', 'price_amount', 'price_ratio']:
                            try:
                                numeric_val = safe_float(current_val)
                                if field in ['quantity_ratio', 'price_ratio'] and (numeric_val < 0 or numeric_val > 100):
                                    error_messages.append(f"第{i+1}行 {field} 值必须在0-100之间")
                                    continue
                                updates[field] = numeric_val
                            except (ValueError, TypeError):
                                error_messages.append(f"第{i+1}行 {field} 必须为有效数值")
                                continue
                        else:
                            # 文本字段
                            if field == 'param_name' and (not current_val or current_val.strip() == ""):
                                error_messages.append(f"第{i+1}行参数名称不能为空")
                                continue
                            updates[field] = str(current_val).strip() if current_val else ''
                
                # 如果有有效的更新，执行数据库更新
                if updates:
                    set_clause = ', '.join([f"`{field}` = %s" for field in updates.keys()])
                    update_query = f"UPDATE `price_custom_parameters` SET {set_clause} WHERE `id` = %s"
                    
                    try:
                        cursor.execute(update_query, list(updates.values()) + [row_id])
                        changes_count += 1
                    except mysql.connector.Error as e:
                        error_messages.append(f"更新第{i+1}行失败: {str(e)}")
            
            if changes_count > 0:
                conn.commit()
            
            # 构建反馈消息
            feedback_alerts = []
            
            if changes_count > 0:
                feedback_alerts.append(
                    dbc.Alert([
                        html.I(className="fas fa-check-circle me-2"),
                        f"成功更新 {changes_count} 条记录。"
                    ], color="success", duration=3000)
                )
            
            if error_messages:
                feedback_alerts.append(
                    dbc.Alert([
                        html.I(className="fas fa-exclamation-triangle me-2"),
                        html.Div([
                            "以下更新失败：",
                            html.Ul([html.Li(msg) for msg in error_messages])
                        ])
                    ], color="warning", duration=5000)
                )
            
            return feedback_alerts if feedback_alerts else None
            
        except mysql.connector.Error as e:
            if conn:
                conn.rollback()
            return dbc.Alert([
                html.I(className="fas fa-exclamation-triangle me-2"),
                f"数据库操作失败: {str(e)}"
            ], color="danger", duration=5000)
            
        except Exception as e:
            if conn:
                conn.rollback()
            return dbc.Alert([
                html.I(className="fas fa-exclamation-triangle me-2"),
                f"保存数据时发生错误: {str(e)}"
            ], color="danger", duration=5000)
            
        finally:
            if conn:
                conn.close()

    @app.callback(
        Output("custom-params-feedback", "children", allow_duplicate=True),
        Input("custom-params-table", "selected_rows"),
        State("custom-params-table", "data"),
        prevent_initial_call=True
    )
    def handle_row_deletion(selected_rows, table_data):
        """处理表格行删除（软删除）"""
        if not selected_rows or not table_data:
            raise PreventUpdate
        
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            deleted_count = 0
            error_messages = []
            
            for row_index in selected_rows:
                if row_index < len(table_data):
                    row_id = table_data[row_index].get('id')
                    param_name = table_data[row_index].get('param_name', '未知参数')
                    
                    if row_id:
                        try:
                            # 软删除：将状态设置为0
                            delete_query = "UPDATE `price_custom_parameters` SET `status` = 0 WHERE `id` = %s"
                            cursor.execute(delete_query, (row_id,))
                            deleted_count += 1
                        except mysql.connector.Error as e:
                            error_messages.append(f"删除参数 '{param_name}' 失败: {str(e)}")
            
            if deleted_count > 0:
                conn.commit()
            
            # 构建反馈消息
            feedback_alerts = []
            
            if deleted_count > 0:
                feedback_alerts.append(
                    dbc.Alert([
                        html.I(className="fas fa-trash-alt me-2"),
                        f"成功删除 {deleted_count} 个参数。刷新表格以查看更新。"
                    ], color="success", duration=3000)
                )
            
            if error_messages:
                feedback_alerts.append(
                    dbc.Alert([
                        html.I(className="fas fa-exclamation-triangle me-2"),
                        html.Div([
                            "以下删除操作失败：",
                            html.Ul([html.Li(msg) for msg in error_messages])
                        ])
                    ], color="warning", duration=5000)
                )
            
            return feedback_alerts if feedback_alerts else None
            
        except mysql.connector.Error as e:
            if conn:
                conn.rollback()
            return dbc.Alert([
                html.I(className="fas fa-exclamation-triangle me-2"),
                f"删除操作失败: {str(e)}"
            ], color="danger", duration=5000)
            
        except Exception as e:
            if conn:
                conn.rollback()
            return dbc.Alert([
                html.I(className="fas fa-exclamation-triangle me-2"),
                f"删除数据时发生错误: {str(e)}"
            ], color="danger", duration=5000)
            
        finally:
            if conn:
                conn.close()

    @app.callback(
        [Output("custom-mode-calculation-result", "children"),
         Output("custom-mode-cost-comparison-container", "style"),
         Output("custom-mode-cost-comparison-table", "children"),
         Output("custom-mode-report-data", "data")],
        Input("confirm-custom-mode", "n_clicks"),
        [State("custom-params-table", "data"),
         State("custom-params-table", "selected_rows")],
        prevent_initial_call=True
    )
    def calculate_custom_mode_prediction(n_clicks, table_data, selected_rows):
        """处理自定义模式的预测计算"""
        if not n_clicks:
            raise PreventUpdate
        
        if not table_data or not selected_rows:
            error_message = dbc.Alert(
                [html.I(className="fas fa-exclamation-triangle me-2"),
                 "请先添加参数并选择要计算的参数行！"],
                color="warning"
            )
            return error_message, {"display": "none"}, ""
        
        try:
            # 获取选中的参数数据
            selected_params = [table_data[i] for i in selected_rows]
            
            # 执行自定义模式预测计算
            calculation_result = perform_custom_mode_calculation(selected_params)
            
            if "error" in calculation_result:
                error_message = dbc.Alert(
                    [html.I(className="fas fa-exclamation-triangle me-2"),
                     f"计算失败：{calculation_result['error']}"],
                    color="danger"
                )
                return error_message, {"display": "none"}, ""
            
            # 生成成功的确认消息
            success_message = html.Div([
                dbc.Alert([
                    html.I(className="fas fa-check-circle me-2"),
                    html.Strong("自定义模式预测计算完成！"),
                    html.Br(),
                    "智能预测结果已显示在下方表格中，请查看详细分析。"
                ], color="success", className="mb-0")
            ])
            
            # 生成详细结果表格
            detailed_results_table = create_custom_mode_results_table(
                calculation_result["input_params"],
                calculation_result["estimated_quantities"], 
                calculation_result["estimated_prices"],
                calculation_result["total_prediction"]
            )
            
            results_container_style = {"display": "block"}


            # 新增：保存报告数据到Store
            report_data = {
                "模式": "自定义模式",
                "生成时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "输入参数": calculation_result["input_params"],
                "预测结果": calculation_result["total_prediction"],
                "估算工程量": calculation_result["estimated_quantities"],
                "估算价格": calculation_result["estimated_prices"]
            }
            return success_message, results_container_style, detailed_results_table, report_data
            
        except Exception as e:
            logger.error(f"自定义模式计算异常: {e}", exc_info=True)
            error_message = dbc.Alert(
                [html.I(className="fas fa-exclamation-triangle me-2"),
                 f"计算过程中发生错误：{str(e)}"],
                color="danger"
            )
            return error_message, {"display": "none"}, ""




    def create_steel_lining_prediction_results_table(quantities, ml_prediction_results, measures_cost_value):
        """创建钢衬里预测结果表格显示 - 四种独立方法版本"""
        
        title_section = html.Div([
            html.H3("钢衬里施工智能预测分析报告", className="text-center text-primary mb-4"),
            html.Hr()
        ])

        input_summary_rows = [
            html.Tr([html.Td("拼装场地总工程量"), html.Td(f"{quantities.get('拼装场地总工程量', 0):.2f}"), html.Td("立方米")]),
            html.Tr([html.Td("制作胎具总工程量"), html.Td(f"{quantities.get('制作胎具总工程量', 0):.2f}"), html.Td("吨")]),
            html.Tr([html.Td("钢支墩埋件总工程量"), html.Td(f"{quantities.get('钢支墩埋件总工程量', 0):.2f}"), html.Td("吨")]),
            html.Tr([html.Td("扶壁柱总工程量"), html.Td(f"{quantities.get('扶壁柱总工程量', 0):.2f}"), html.Td("吨")]),
            html.Tr([html.Td("走道板操作平台总工程量"), html.Td(f"{quantities.get('走道板操作平台总工程量', 0):.2f}"), html.Td("吨")]),
            html.Tr([html.Td("钢网梁总工程量"), html.Td(f"{quantities.get('钢网梁总工程量', 0):.2f}"), html.Td("吨")]),
        ]
        if measures_cost_value > 0:
            input_summary_rows.append(
                html.Tr([html.Td("措施费", style={'fontWeight': 'bold'}), html.Td(f"{measures_cost_value:,.2f}", style={'fontWeight': 'bold'}), html.Td("元", style={'fontWeight': 'bold'})])
            )

        input_summary = html.Div([
            html.H5("📋 项目工程量输入汇总", className="text-info mb-3"),
            dbc.Row([
                dbc.Col(dbc.Card([
                    dbc.CardHeader("📊 数据概览"),
                    dbc.CardBody([
                        html.P(f"工程量类型: {len(input_summary_rows)}项", className="mb-1"),
                        html.P(f"有效输入: {sum(1 for k, v in quantities.items() if v is not None and v > 0) + (1 if measures_cost_value > 0 else 0)}项", className="mb-1"),
                        html.P(f"预测状态: {'✅ 可执行' if any(v is not None and v > 0 for v in quantities.values()) else '❌ 数据不足'}", className="mb-0")
                    ])
                ]), width=4)
            ], className="mb-4")
        ])

        # 预测方法状态总览
        status_summary_section = create_prediction_status_summary(ml_prediction_results, mode='steel_lining')

        # 【在这里插入】最佳算法参数信息部分
        best_algorithm_section = create_best_algorithm_params_section(ml_prediction_results)
            
        prediction_section = html.Div()

        if ml_prediction_results and "error" not in ml_prediction_results:
            # 获取状态信息
            method_status = ml_prediction_results.get("预测方法状态", {})
            
            # 获取四种独立的预测结果
            ai_raw = ml_prediction_results.get("AI预测-原始值")
            ai_final = ml_prediction_results.get("AI预测-最终值") 
            ratio_raw = ml_prediction_results.get("比率法-原始值")
            ratio_final = ml_prediction_results.get("比率法-最终值")
            
            matched_rules_source = ml_prediction_results.get("匹配到的规则来源", "未知")
            estimated_quantities_from_ml = ml_prediction_results.get("估算的工程量", {})
            estimated_costs_from_ml = ml_prediction_results.get("估算的各项成本 (用于ML的特征)", {})

            prediction_table_rows = []

            # 匹配规则来源
            prediction_table_rows.append(
                html.Tr([
                    html.Td("🎯 匹配规则来源", className="fw-bold"),
                    html.Td(matched_rules_source, className="text-info"),
                    html.Td("-"), html.Td("-"), html.Td("-"),
                    html.Td("基于历史项目聚类分析匹配最相似的项目类型")
                ])
            )

            # 智能工程量补全
            if estimated_quantities_from_ml:
                prediction_table_rows.append(
                    html.Tr([html.Td("📊 智能工程量补全 (机器学习模型估算)", className="fw-bold", colSpan=6)], className="table-info")
                )
                for param, qty in estimated_quantities_from_ml.items():
                    if quantities.get(param) is None or quantities.get(param) == 0:
                        unit = get_quantity_unit(param)
                        prediction_table_rows.append(
                            html.Tr([html.Td(f"  └ {param}"), html.Td(f"{qty:.2f}"), html.Td(unit), html.Td("-"), html.Td("-"), html.Td("基于历史数据比例关系智能推算")])
                        )

            # 各项成本预测
            if estimated_costs_from_ml:
                prediction_table_rows.append(
                    html.Tr([html.Td("💰 各项成本预测 (机器学习模型估算)", className="fw-bold", colSpan=6)], className="table-warning")
                )
                # 钢衬里模式的成本项目
                cost_items = [
                    "拼装场地费用", "制作胎具费用", "钢支墩、埋件费用", 
                    "扶壁柱费用", "走道板及操作平台费用", "钢网架费用"
                ]
                for item in cost_items:
                    if item in estimated_costs_from_ml:
                        prediction_table_rows.append(
                            html.Tr([html.Td(f"  └ {item}"), html.Td(f"{estimated_costs_from_ml[item]:,.2f}"), html.Td("元"), html.Td("-"), html.Td("-"), html.Td("基于估算工程量和历史单价计算")])
                        )

            # 关键新增：解析最佳算法信息（钢衬里模式）
            best_algorithm_info = get_best_algorithm_info_safe(ml_prediction_results)

            # 四种独立的预测方法显示 - 显示具体算法名称（钢衬里模式）
            prediction_methods = [
                {
                    'key': 'ml_prediction_raw',
                    'title': f'🤖 {best_algorithm_info["raw"]["display_name"]}',
                    'data': ai_raw,
                    'description': f'{best_algorithm_info["raw"]["description"]} (不含措施费)',
                    'row_class': 'table-primary',
                    'algorithm_details': best_algorithm_info["raw"]
                },
                {
                    'key': 'ml_prediction_final', 
                    'title': f'🤖 {best_algorithm_info["final"]["display_name"]}',
                    'data': ai_final,
                    'description': f'{best_algorithm_info["final"]["description"]} (含措施费 ¥{measures_cost_value:,.2f})',
                    'row_class': 'table-warning',
                    'algorithm_details': best_algorithm_info["final"]
                },
                {
                    'key': 'ratio_method_raw',
                    'title': '📈 比率法-原始值', 
                    'data': ratio_raw,
                    'description': '比率法预测，不含措施费',
                    'row_class': 'table-info'
                },
                {
                    'key': 'ratio_method_final',
                    'title': '📈 比率法-最终值',
                    'data': ratio_final, 
                    'description': f'比率法预测，已包含措施费 ¥{measures_cost_value:,.2f}',
                    'row_class': 'table-success'
                }
            ]

            for method in prediction_methods:
                        method_data = method['data']
                        method_title = method['title']
                        method_description = method['description']
                        method_row_class = method['row_class']
                        
                        # 检查预测结果的状态
                        if method_data is None:
                            # 数据为空
                            prediction_table_rows.append(
                                html.Tr([
                                    html.Td(f"❌ {method_title} (无数据)", style={'color': '#dc3545', 'fontWeight': 'bold'}),
                                    html.Td("无预测结果", style={'color': '#dc3545'}),
                                    html.Td("元", style={'color': '#6c757d'}),
                                    html.Td("-", style={'color': '#6c757d'}),
                                    html.Td("-", style={'color': '#6c757d'}),
                                    html.Td("预测计算未执行或失败", style={'color': '#6c757d'})
                                ], style={'backgroundColor': '#fff5f5'})
                            )
                        elif isinstance(method_data, dict):
                            if method_data.get('status') == 'display_disabled':
                                # 可以执行但显示被禁用
                                prediction_table_rows.append(
                                    html.Tr([
                                        html.Td(f"🚫 {method_title} (显示已禁用)", style={'color': '#ffc107', 'fontWeight': 'bold'}),
                                        html.Td("🚫 显示被禁用", style={'color': '#ffc107', 'fontWeight': 'bold'}),
                                        html.Td("元", style={'color': '#6c757d'}),
                                        html.Td("-", style={'color': '#6c757d'}),
                                        html.Td("-", style={'color': '#6c757d'}),
                                        html.Td(method_data.get('message', '显示权限已禁用'), style={'color': '#6c757d'})
                                    ], style={'backgroundColor': '#fffbf0'})
                                )
                            elif method_data.get('status') == 'execution_failed':
                                # 想显示但无法执行
                                prediction_table_rows.append(
                                    html.Tr([
                                        html.Td(f"❌ {method_title} (执行失败)", style={'color': '#dc3545', 'fontWeight': 'bold'}),
                                        html.Td("执行失败", style={'color': '#dc3545', 'fontWeight': 'bold'}),
                                        html.Td("元", style={'color': '#6c757d'}),
                                        html.Td("-", style={'color': '#6c757d'}),
                                        html.Td("-", style={'color': '#6c757d'}),
                                        html.Td(method_data.get('reason', '执行条件不满足'), style={'color': '#6c757d'})
                                    ], style={'backgroundColor': '#fff5f5'})
                                )
                            elif method_data.get('status') == 'fully_disabled':
                                # 既不能执行也不能显示
                                prediction_table_rows.append(
                                    html.Tr([
                                        html.Td(f"🚫 {method_title} (完全禁用)", style={'color': '#6c757d', 'fontWeight': 'bold'}),
                                        html.Td("🚫 完全禁用", style={'color': '#6c757d', 'fontWeight': 'bold'}),
                                        html.Td("元", style={'color': '#6c757d'}),
                                        html.Td("-", style={'color': '#6c757d'}),
                                        html.Td("-", style={'color': '#6c757d'}),
                                        html.Td("执行和显示都已禁用", style={'color': '#6c757d'})
                                    ], style={'backgroundColor': '#f8f9fa', 'opacity': '0.7'})
                                )
                            elif 'error' in method_data:
                                # 执行过程中出现错误
                                prediction_table_rows.append(
                                    html.Tr([
                                        html.Td(f"⚠️ {method_title} (计算错误)", style={'color': '#dc3545', 'fontWeight': 'bold'}),
                                        html.Td("计算异常", style={'color': '#dc3545', 'fontWeight': 'bold'}),
                                        html.Td("元", style={'color': '#6c757d'}),
                                        html.Td("-", style={'color': '#6c757d'}),
                                        html.Td("-", style={'color': '#6c757d'}),
                                        html.Td(str(method_data.get('error', '未知错误'))[:50] + "...", style={'color': '#6c757d'})
                                    ], style={'backgroundColor': '#fff5f5'})
                                )
                            elif method['key'].startswith('ml_') and method_data.get('集成平均预测'):
                                # AI预测方法：正常显示集成结果，添加算法详情
                                ensemble_prediction = method_data['集成平均预测']
                                algorithm_details = method.get('algorithm_details', {})
                                confidence_info = algorithm_details.get('confidence', '高可信度')
                                
                                prediction_table_rows.append(
                                    html.Tr([
                                        html.Td([
                                            html.Div(method_title, className="fw-bold"),
                                            html.Small([
                                            html.Span("🏆 最佳预测算法 | ", className="text-success fw-bold", style={'fontSize': '11px'}),
                                            html.Span(f"算法可信度: {confidence_info}", className="text-muted", style={'fontSize': '11px'})
                                        ], className="d-block")
                                        ], style={'color': 'black'}),
                                        html.Td(f"{ensemble_prediction:,.0f}", className="fw-bold", style={'color': 'black'}),
                                        html.Td("元", className="fw-bold", style={'color': 'black'}),
                                        html.Td("-", className="fw-bold", style={'color': 'black'}),
                                        html.Td("-", className="fw-bold", style={'color': 'black'}),
                                        html.Td([
                                            html.Div(method_description, className="mb-1"),
                                            html.Small(f"基于算法: {algorithm_details.get('algorithm_name', '未知算法')}", className="text-info", style={'fontSize': '11px'})
                                        ], style={'color': 'black'})
                                    ], className=method_row_class)
                                )
                            else:
                                # 其他字典类型的处理
                                prediction_table_rows.append(
                                    html.Tr([
                                        html.Td(f"❓ {method_title} (状态未知)", style={'color': '#6c757d', 'fontWeight': 'bold'}),
                                        html.Td("状态未知", style={'color': '#6c757d'}),
                                        html.Td("元", style={'color': '#6c757d'}),
                                        html.Td("-", style={'color': '#6c757d'}),
                                        html.Td("-", style={'color': '#6c757d'}),
                                        html.Td("请检查预测系统配置", style={'color': '#6c757d'})
                                    ], style={'backgroundColor': '#f8f9fa'})
                                )
                        elif isinstance(method_data, (int, float)) and method['key'].startswith('ratio_'):
                            # 比率法：正常的数值结果
                            prediction_table_rows.append(
                                html.Tr([
                                    html.Td(method_title, className="fw-bold", style={'color': 'black'}),
                                    html.Td(f"{method_data:,.0f}", className="fw-bold", style={'color': 'black'}),
                                    html.Td("元", className="fw-bold", style={'color': 'black'}),
                                    html.Td("-", className="fw-bold", style={'color': 'black'}),
                                    html.Td("-", className="fw-bold", style={'color': 'black'}),
                                    html.Td(method_description, className="fw-bold", style={'color': 'black'})
                                ], className=method_row_class)
                            )
                        else:
                            # 其他未知类型的处理
                            prediction_table_rows.append(
                                html.Tr([
                                    html.Td(f"❓ {method_title} (类型异常)", style={'color': '#dc3545', 'fontWeight': 'bold'}),
                                    html.Td("数据类型错误", style={'color': '#dc3545'}),
                                    html.Td("元", style={'color': '#6c757d'}),
                                    html.Td("-", style={'color': '#6c757d'}),
                                    html.Td("-", style={'color': '#6c757d'}),
                                    html.Td(f"预期数据类型异常: {type(method_data)}", style={'color': '#6c757d'})
                                ], style={'backgroundColor': '#fff5f5'})
                            )

            prediction_section = html.Div([
                html.H5("智能预测分析结果", className="text-primary mb-3"),
                dbc.Table([
                    html.Thead(html.Tr([
                        html.Th("预测项目", style={"width": "25%"}),
                        html.Th("预测总价(元)", style={"width": "15%"}),
                        html.Th("单位", style={"width": "5%"}),
                        html.Th("与比例法偏差(元)", style={"width": "15%"}),
                        html.Th("与比例法偏差(%)", style={"width": "15%"}),
                        html.Th("算法说明", style={"width": "25%"})
                    ])),
                    html.Tbody(prediction_table_rows)
                ], bordered=True, hover=True, striped=True, size="sm", className="mb-4"),
                create_independent_prediction_summary(method_status, ai_raw, ai_final, ratio_raw, ratio_final, measures_cost_value, mode='steel_lining')
            ])

        elif ml_prediction_results and "error" in ml_prediction_results:
            prediction_section = html.Div([
                html.H5("智能预测结果", className="text-warning mb-3"),
                dbc.Alert([
                    html.I(className="fas fa-exclamation-triangle me-2"),
                    html.Strong("⚠️ 预测失败: "),
                    html.Span(ml_prediction_results['error'])
                ], color="warning", className="mb-4")
            ])

        instruction_section = html.Div([
            html.Hr(),
            html.H6("📖 技术说明", className="text-muted mb-3"),
            dbc.Row([
                dbc.Col(dbc.Card([dbc.CardHeader("🔬 算法原理"), dbc.CardBody(html.Ul([html.Li("K-Means聚类：自动识别项目类型"), html.Li("多算法集成：提高预测准确性"), html.Li("交叉验证：四种独立方法相互验证")], className="small mb-0"))]), width=4),
                dbc.Col(dbc.Card([dbc.CardHeader("📊 数据来源"), dbc.CardBody(html.Ul([html.Li("历史项目数据库：真实施工记录"), html.Li("成本规律提取：自动学习价格模式"), html.Li("持续优化：新数据不断完善模型")], className="small mb-0"))]), width=4),
                dbc.Col(dbc.Card([dbc.CardHeader("💡 使用建议"), dbc.CardBody(html.Ul([html.Li("独立方法：每种方法可单独参考"), html.Li("灵活选择：根据项目需要选择合适方法"), html.Li("综合判断：多方法结果综合考虑")], className="small mb-0"))]), width=4)
            ])
        ])

        return html.Div([title_section, input_summary, status_summary_section,best_algorithm_section, prediction_section, instruction_section], className="container-fluid")
   
    def get_best_algorithm_info_safe(ml_prediction_results):
        """
        安全版本：解析预测结果，获取最佳算法信息，处理停用算法的字典状态
        
        Args:
            ml_prediction_results: 预测结果字典
            
        Returns:
            dict: 包含原始值和最终值的最佳算法信息
        """
        # 算法显示名称映射
        algorithm_display_names = {
            "岭回归 (RidgeCV)": "岭回归算法预测",
            "决策树 (Decision Tree)": "决策树算法预测", 
            "随机森林 (Random Forest)": "随机森林算法预测",
            "支持向量回归 (SVR)": "支持向量机算法预测",
            "神经网络 (MLPRegressor)": "神经网络算法预测"
        }
        
        # 获取比率法预测值用于计算偏差
        ratio_raw = ml_prediction_results.get("比率法-原始值", 0)
        if isinstance(ratio_raw, dict):
            ratio_raw = 0  # 如果比率法也是字典状态，设为0
        
        # 分析原始值的最佳算法
        ai_raw_data = ml_prediction_results.get("AI预测-原始值", {})
        raw_best_algo = find_best_single_algorithm_safe(ai_raw_data, ratio_raw)
        
        # 分析最终值的最佳算法
        ai_final_data = ml_prediction_results.get("AI预测-最终值", {})
        final_best_algo = find_best_single_algorithm_safe(ai_final_data, ratio_raw)
        
        return {
            "raw": {
                "display_name": algorithm_display_names.get(raw_best_algo["algorithm"], "集成算法预测"),
                "algorithm_name": raw_best_algo["algorithm"],
                "confidence": raw_best_algo["confidence"],
                "description": f"基于{raw_best_algo['algorithm']}的预测结果"
            },
            "final": {
                "display_name": algorithm_display_names.get(final_best_algo["algorithm"], "集成算法预测"),
                "algorithm_name": final_best_algo["algorithm"], 
                "confidence": final_best_algo["confidence"],
                "description": f"基于{final_best_algo['algorithm']}的预测结果"
            }
        }


    def find_best_single_algorithm_safe(ai_prediction_data, ratio_value):
        """
        安全版本：从AI预测数据中找出表现最佳的单一算法，处理停用算法的字典状态
        
        Args:
            ai_prediction_data: AI预测数据字典
            ratio_value: 比率法预测值
            
        Returns:
            dict: 最佳算法信息
        """
        if not isinstance(ai_prediction_data, dict) or not isinstance(ratio_value, (int, float)) or ratio_value <= 0:
            return {
                "algorithm": "集成平均",
                "confidence": "中等",
                "deviation": 0
            }
        
        best_algorithm = "集成平均"
        min_deviation = float('inf')
        best_confidence = "中等"
        valid_algorithms_found = False
        
        # 遍历所有算法预测结果，跳过停用的算法
        for algo_name, prediction_value in ai_prediction_data.items():
            if algo_name == '集成平均预测':
                continue
                
            # 检查是否是有效的数值结果（不是停用算法的字典状态）
            if isinstance(prediction_value, (int, float)) and prediction_value > 0:
                valid_algorithms_found = True
                
                # 计算与比率法的偏差
                try:
                    deviation = abs(prediction_value - ratio_value)
                    
                    if deviation < min_deviation:
                        min_deviation = deviation
                        best_algorithm = algo_name
                        
                        # 根据偏差程度确定可信度
                        deviation_pct = (deviation / ratio_value * 100) if ratio_value > 0 else 100
                        if deviation_pct < 5:
                            best_confidence = "极高可信度"
                        elif deviation_pct < 10:
                            best_confidence = "高可信度"
                        elif deviation_pct < 20:
                            best_confidence = "中等可信度"
                        else:
                            best_confidence = "低可信度"
                            
                except (TypeError, ValueError) as e:
                    logger.warning(f"计算算法 {algo_name} 偏差时出错: {e}")
                    continue
                    
            elif isinstance(prediction_value, dict):
                # 记录停用的算法，但不参与最佳算法选择
                logger.info(f"算法 {algo_name} 已停用: {prediction_value.get('status', '未知状态')}")
                continue
        
        # 如果没有找到有效的算法，检查是否有集成平均预测
        if not valid_algorithms_found:
            ensemble_prediction = ai_prediction_data.get('集成平均预测')
            if isinstance(ensemble_prediction, (int, float)) and ensemble_prediction > 0:
                try:
                    deviation = abs(ensemble_prediction - ratio_value) if ratio_value > 0 else 0
                    deviation_pct = (deviation / ratio_value * 100) if ratio_value > 0 else 0
                    
                    if deviation_pct < 10:
                        best_confidence = "高可信度"
                    elif deviation_pct < 20:
                        best_confidence = "中等可信度"
                    else:
                        best_confidence = "低可信度"
                        
                    return {
                        "algorithm": "集成平均",
                        "confidence": best_confidence,
                        "deviation": deviation
                    }
                except (TypeError, ValueError):
                    pass
        
        return {
            "algorithm": best_algorithm,
            "confidence": best_confidence,
            "deviation": min_deviation if min_deviation != float('inf') else 0
        }
    # 需要修改之前的辅助函数，添加模式参数支持

    def create_enhanced_prediction_summary_with_status(ml_predictions, ratio_prediction_value, method_status, mode='steel_cage'):
        """创建增强版预测结果总结 - 支持状态控制和模式参数"""
        ml_enabled = method_status.get('ml_prediction_raw', {}).get('enabled', True)
        ratio_enabled = method_status.get('ratio_method_raw', {}).get('enabled', True)
        
        mode_name = "钢筋笼" if mode == 'steel_cage' else "钢衬里"
        
        if not ml_enabled and not ratio_enabled:
            return html.Div([
                dbc.Alert([
                    html.I(className="fas fa-exclamation-triangle me-2"),
                    html.Strong("⚠️ 所有预测方法都已禁用"),
                    html.Br(),
                    f"请前往数据管理模块启用{mode_name}模式相关预测方法指标"
                ], color="danger")
            ])
        
        if not ml_enabled or not ratio_enabled:
            missing_method = "机器学习预测" if not ml_enabled else "比率法预测"
            return html.Div([
                dbc.Alert([
                    html.I(className="fas fa-info-circle me-2"),
                    html.Strong(f"部分预测方法不可用：{missing_method}已禁用"),
                    html.Br(),
                    f"建议启用完整的{mode_name}模式预测方法以获得更准确的分析结果"
                ], color="warning")
            ])
        
        # 如果两种方法都启用，使用原来的逻辑
        return create_enhanced_prediction_summary(ml_predictions, ratio_prediction_value, mode)

    def create_enhanced_prediction_summary(ml_predictions, ratio_prediction_value, mode='steel_cage'):
        """创建增强版预测结果总结"""
        if not ml_predictions and not ratio_prediction_value:
            return html.Div()

        summary_content = []

        if ml_predictions and ml_predictions.get('集成平均预测') is not None and ratio_prediction_value is not None:
            ml_pred = ml_predictions['集成平均预测']

            diff_pct = abs(ml_pred - ratio_prediction_value) / max(ml_pred, ratio_prediction_value) * 100 if max(ml_pred, ratio_prediction_value) != 0 else 0

            if diff_pct < 10:
                confidence = "高"
                color = "success"
                icon = "✅"
                recommendation = "两种方法预测总价接近，建议采用集成平均预测值。"
            elif diff_pct < 20:
                confidence = "中"
                color = "warning"
                icon = "⚠️"
                recommendation = "两种方法预测总价存在一定差异，建议结合项目特点综合判断。"
            else:
                confidence = "低"
                color = "danger"
                icon = "❌"
                recommendation = "两种方法预测总价差异较大，建议重新检查输入数据或寻求专家意见。"

            summary_content.append(
                dbc.Row([
                    dbc.Col(dbc.Card([
                        dbc.CardHeader(html.H6(f"{icon} 预测可信度评估: {confidence}", className="mb-0")),
                        dbc.CardBody([
                            html.P(f"🤖 机器学习集成平均预测总价: {ml_pred:,.0f} 元", className="mb-1"),
                            html.P(f"📈 比率法预测总价: {ratio_prediction_value:,.0f} 元", className="mb-1"),
                            html.P(f"📊 预测总价差异: {diff_pct:.1f}%", className="mb-2"),
                            html.Hr(),
                            html.P(recommendation, className="small text-muted mb-0")
                        ])
                    ], color=color, outline=True), width=6),
                    dbc.Col(dbc.Card([
                        dbc.CardHeader(html.H6("📈 预测总价区间建议", className="mb-0")),
                        dbc.CardBody([
                            html.P(f"预测总价中值: {(ml_pred + ratio_prediction_value) / 2:,.0f} 元", className="mb-1"),
                            html.P(f"预测总价范围: {min(ml_pred, ratio_prediction_value):,.0f} - {max(ml_pred, ratio_prediction_value):,.0f} 元", className="mb-1"),
                            html.P(f"建议预算 (含10%缓冲): {(ml_pred + ratio_prediction_value) / 2 * 1.1:,.0f} 元", className="mb-2"),
                            html.Hr(),
                            html.P("建议在预测总价基础上增加10%的风险缓冲。", className="small text-muted mb-0")
                        ])
                    ], color="info", outline=True), width=6)
                ], className="mb-3")
            )

        return html.Div(summary_content)
    

# 在 register_price_prediction_callbacks(app) 函数的最后，添加以下三个回调函数
    # 它们应该与其他回调函数有相同的缩进级别
    @app.callback(
        Output("download-steel-cage-report", "data"),
        Input("export-steel-cage-report-btn", "n_clicks"),
        State("steel-cage-report-data", "data"),
        prevent_initial_call=True
    )
    def export_steel_cage_excel(n_clicks, report_data):
        """导出钢筋笼模式报告为Excel并保存到数据库"""
        if not n_clicks or not report_data:
            raise PreventUpdate

            
        try:
            from openpyxl import Workbook
            from openpyxl.chart import PieChart, Reference
            from openpyxl.chart.label import DataLabelList
            from openpyxl.utils import get_column_letter
            import tempfile
            import os
            
            # 数据完整性检查
            if not report_data:
                raise Exception("报告数据为空，无法生成Excel文件")
            
            quantities = report_data.get('工程量数据', {})
            ml_prediction_results = report_data.get('预测结果', {})
            measures_cost_value = report_data.get('措施费', 0)
            
            # 获取预测结果
            ml_predictions = ml_prediction_results.get("机器学习预测结果", {}) if ml_prediction_results else {}
            ratio_prediction = ml_prediction_results.get("比率法预测总价") if ml_prediction_results else None
            estimated_costs = ml_prediction_results.get("估算的各项成本 (用于ML的特征)", {}) if ml_prediction_results else {}
            
            # 计算总预测价格
            ensemble_prediction = ml_predictions.get('集成平均预测') if ml_predictions else None
            total_cost = ensemble_prediction or ratio_prediction or 0
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "钢筋笼模式预测报告"
            
            # 设置标题
            ws['A1'] = "钢筋笼施工模式智能预测分析报告"
            ws['A1'].font = ws['A1'].font.copy(bold=True, size=16)
            ws.merge_cells('A1:E1')
            
            # 基本信息
            row = 3
            ws[f'A{row}'] = "报告概述"
            ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
            row += 1
            
            ws[f'A{row}'] = "施工模式:"
            ws[f'B{row}'] = report_data.get('模式', '钢筋笼施工模式')
            row += 1
            ws[f'A{row}'] = "生成时间:"
            ws[f'B{row}'] = report_data.get('生成时间', '')
            row += 1
            ws[f'A{row}'] = "预测总价(元):"
            ws[f'B{row}'] = f"{total_cost:,.0f}"
            row += 2
            
            # 工程量数据
            ws[f'A{row}'] = "工程量数据详情"
            ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
            row += 1
            
            # 工程量表头
            quantity_headers = ['工程量项目', '数量', '单位', '备注']
            for col, header in enumerate(quantity_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = cell.font.copy(bold=True)
            row += 1
            
            # 工程量数据
            quantity_items = [
                ('钢筋总吨数', quantities.get('钢筋总吨数', 0), '吨'),
                ('塔吊租赁工程量', quantities.get('塔吊租赁工程量', 0), '台班'),
                ('吊索具数量', quantities.get('吊索具数量', 0), '套'),
                ('套筒数量', quantities.get('套筒数量', 0), '个'),
            ]
            
            for item_name, quantity, unit in quantity_items:
                ws[f'A{row}'] = item_name
                ws[f'B{row}'] = f"{quantity:.2f}"
                ws[f'C{row}'] = unit
                ws[f'D{row}'] = "用户输入或智能估算"
                row += 1
            
            if measures_cost_value > 0:
                ws[f'A{row}'] = "措施费"
                ws[f'B{row}'] = f"{measures_cost_value:,.2f}"
                ws[f'C{row}'] = "元"
                ws[f'D{row}'] = "额外费用"
                row += 1
            row += 1
            
            # 预测结果
            ws[f'A{row}'] = "预测结果分析"
            ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
            row += 1
            
            # 结果表头
            result_headers = ['预测方法', '预测总价(元)', '可信度', '备注']
            for col, header in enumerate(result_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = cell.font.copy(bold=True)
            row += 1
            
# 预测结果数据 - 适配新的状态结构
            prediction_status = report_data.get('预测方法状态', {})
            
            # AI预测结果处理
            main_ai_result = report_data.get('主要AI预测', {})
            ai_status = main_ai_result.get('status', 'unknown')
            ai_value = main_ai_result.get('value')
            
            if ai_status == 'success' and isinstance(ai_value, (int, float)):
                ws[f'A{row}'] = "AI预测（机器学习集成）"
                ws[f'B{row}'] = f"{ai_value:,.0f}"
                ws[f'C{row}'] = "高"
                ws[f'D{row}'] = "多算法集成结果，预测成功"
                row += 1
            elif ai_status == 'display_disabled':
                ws[f'A{row}'] = "AI预测（机器学习集成）"
                ws[f'B{row}'] = "显示被禁用"
                ws[f'C{row}'] = "不适用"
                ws[f'D{row}'] = main_ai_result.get('message', '显示权限被禁用')
                row += 1
            elif ai_status == 'error':
                ws[f'A{row}'] = "AI预测（机器学习集成）"
                ws[f'B{row}'] = "预测失败"
                ws[f'C{row}'] = "低"
                ws[f'D{row}'] = f"错误: {main_ai_result.get('message', '未知错误')}"
                row += 1
            else:
                ws[f'A{row}'] = "AI预测（机器学习集成）"
                ws[f'B{row}'] = "不可用"
                ws[f'C{row}'] = "无"
                ws[f'D{row}'] = f"状态: {ai_status}, 消息: {main_ai_result.get('message', '未知状态')}"
                row += 1
            
            # 比率法预测结果处理
            main_ratio_result = report_data.get('主要比率法预测', {})
            ratio_status = main_ratio_result.get('status', 'unknown')
            ratio_value = main_ratio_result.get('value')
            
            if ratio_status == 'success' and isinstance(ratio_value, (int, float)):
                ws[f'A{row}'] = "比率法预测"
                ws[f'B{row}'] = f"{ratio_value:,.0f}"
                ws[f'C{row}'] = "中"
                ws[f'D{row}'] = "传统估算方法，预测成功"
                row += 1
            elif ratio_status == 'display_disabled':
                ws[f'A{row}'] = "比率法预测"
                ws[f'B{row}'] = "显示被禁用"
                ws[f'C{row}'] = "不适用"
                ws[f'D{row}'] = main_ratio_result.get('message', '显示权限被禁用')
                row += 1
            elif ratio_status == 'error':
                ws[f'A{row}'] = "比率法预测"
                ws[f'B{row}'] = "预测失败"
                ws[f'C{row}'] = "低"
                ws[f'D{row}'] = f"错误: {main_ratio_result.get('message', '未知错误')}"
                row += 1
            else:
                ws[f'A{row}'] = "比率法预测"
                ws[f'B{row}'] = "不可用"
                ws[f'C{row}'] = "无"
                ws[f'D{row}'] = f"状态: {ratio_status}, 消息: {main_ratio_result.get('message', '未知状态')}"
                row += 1
            row += 1
            
            # ===== 新增：最佳模型分析 =====
            if ml_predictions and ratio_prediction:
                ws[f'A{row}'] = "最佳模型分析"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
                row += 1
                
                # 找出最佳模型（最接近比率法的模型）
                optimal_model_key = None
                algorithm_names = {
                    'LinearRegression': '线性回归算法',
                    'DecisionTree': '决策树算法',
                    'RandomForest': '随机森林算法',
                    'SupportVectorRegression': '支持向量机算法',
                    'NeuralNetwork': '神经网络算法',
                    "岭回归 (RidgeCV)": "岭回归算法"
                }
                min_diff_to_ratio = float('inf')  # 添加这行：初始化变量
                for model_key_iter, model_pred_val in ml_predictions.items():
                    if (model_key_iter != '集成平均预测' and 
                        model_pred_val is not None and 
                        isinstance(model_pred_val, (int, float))):
                            diff = abs(model_pred_val - ratio_prediction)
                            if diff < min_diff_to_ratio:
                                min_diff_to_ratio = diff
                                optimal_model_key = model_key_iter
                
                if optimal_model_key and ml_predictions.get(optimal_model_key) is not None:
                    model_pred_val = ml_predictions[optimal_model_key]
                    # 添加类型检查，确保预测值是数值类型
                    if isinstance(model_pred_val, (int, float)):
                        deviation_val = model_pred_val - ratio_prediction if ratio_prediction else 0
                        deviation_pct = (deviation_val / ratio_prediction * 100) if ratio_prediction != 0 else 0
                        
                        ws[f'A{row}'] = "最佳单一模型:"
                        ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True)
                        ws[f'B{row}'] = f"{algorithm_names.get(optimal_model_key, optimal_model_key)}"
                        row += 1
                        
                        ws[f'A{row}'] = "最佳模型预测价格:"
                        ws[f'B{row}'] = f"{model_pred_val:,.0f} 元"
                        row += 1
                        
                        ws[f'A{row}'] = "与比率法偏差:"
                        ws[f'B{row}'] = f"{deviation_val:,.0f} 元 ({deviation_pct:.1f}%)"
                        row += 1
                        
                        ws[f'A{row}'] = "模型特点:"
                        ws[f'B{row}'] = "基于历史项目数据训练，最接近传统比率法预测"
                        row += 1
                    else:
                        # 如果最佳模型被停用，跳过计算或设置默认值
                        deviation_val = 0
                        deviation_pct = 0
                row += 1
           # ===== 新增：算法执行状态分析 =====
            algorithm_status = report_data.get('算法执行状态', {})
            if algorithm_status:
                ws[f'A{row}'] = "算法执行状态分析"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
                row += 1
                
                ws[f'A{row}'] = "AI算法可用性:"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True)
                can_execute = algorithm_status.get('can_execute_ai', False)
                ws[f'B{row}'] = "可用" if can_execute else "不可用"
                ws[f'B{row}'].font = ws[f'B{row}'].font.copy(bold=True)
                row += 1
                
                ws[f'A{row}'] = "启用算法数量:"
                enabled_count = algorithm_status.get('enabled_count', 0)
                total_count = algorithm_status.get('total_count', 0)
                ws[f'B{row}'] = f"{enabled_count}/{total_count}"
                row += 1
                
                ws[f'A{row}'] = "算法状态消息:"
                ws[f'B{row}'] = algorithm_status.get('message', '无消息')
                row += 1
                
                ws[f'A{row}'] = "可用算法列表:"
                enabled_algorithms = algorithm_status.get('enabled_algorithms', [])
                ws[f'B{row}'] = ', '.join(enabled_algorithms) if enabled_algorithms else '无可用算法'
                row += 1
                row += 1
            
            # ===== 新增：预测方法详细状态 =====
            method_status = report_data.get('预测方法状态', {})
            if method_status:
                ws[f'A{row}'] = "预测方法详细状态"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
                row += 1
                
                # 状态表头
                status_headers = ['预测方法', '最终状态', '可执行', '可显示', '执行消息', '显示消息']
                for col, header in enumerate(status_headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = cell.font.copy(bold=True)
                row += 1
                
                # 状态详情
                status_mapping = {
                    'fully_available': '完全可用',
                    'execute_only': '可执行但显示禁用',
                    'display_error': '显示启用但执行失败',
                    'fully_disabled': '完全禁用',
                    'unknown': '状态未知'
                }
                
                for method_key, method_info in method_status.items():
                    method_name = method_info.get('name', method_key)
                    final_status = method_info.get('final_status', 'unknown')
                    can_execute = method_info.get('can_execute', False)
                    can_display = method_info.get('can_display', False)
                    execution_msg = method_info.get('execution_message', '')
                    display_msg = method_info.get('display_message', '')
                    
                    ws[f'A{row}'] = method_name
                    ws[f'B{row}'] = status_mapping.get(final_status, final_status)
                    ws[f'C{row}'] = "是" if can_execute else "否"
                    ws[f'D{row}'] = "是" if can_display else "否"
                    ws[f'E{row}'] = execution_msg[:50] + "..." if len(execution_msg) > 50 else execution_msg
                    ws[f'F{row}'] = display_msg[:50] + "..." if len(display_msg) > 50 else display_msg
                    row += 1
                row += 1
            
            # ===== 新增：综合状态汇总 =====
            prediction_status_summary = format_prediction_status_for_export(report_data.get('预测结果', {}))
            if prediction_status_summary:
                ws[f'A{row}'] = "综合状态汇总"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
                row += 1
                
                ws[f'A{row}'] = "状态概览:"
                ws[f'B{row}'] = prediction_status_summary.get('状态概览', '无信息')
                row += 1
                
                ws[f'A{row}'] = "完全可用方法数:"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True)
                ws[f'B{row}'] = str(prediction_status_summary.get('可用方法数量', 0))
                ws[f'B{row}'].font = ws[f'B{row}'].font.copy(bold=True)
                row += 1
                
                ws[f'A{row}'] = "算法执行能力:"
                ws[f'B{row}'] = prediction_status_summary.get('算法执行能力', '未知')
                row += 1
                
                ws[f'A{row}'] = "系统建议:"
                if prediction_status_summary.get('可用方法数量', 0) == 4:
                    suggestion = "系统运行正常，所有预测方法都可用，可进行全面分析"
                elif prediction_status_summary.get('可用方法数量', 0) > 0:
                    suggestion = "系统部分正常，建议检查相关配置以启用更多预测方法"
                else:
                    suggestion = "系统存在问题，建议检查算法配置和综合指标设置"
                ws[f'B{row}'] = suggestion
                row += 1
                row += 1
           
            # ===== 新增：预测可信度评估 =====
            if ensemble_prediction and ratio_prediction:
                ws[f'A{row}'] = "预测可信度评估"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
                row += 1
                
                # 计算差异百分比
                diff_pct = abs(ensemble_prediction - ratio_prediction) / max(ensemble_prediction, ratio_prediction) * 100 if max(ensemble_prediction, ratio_prediction) != 0 else 0
                
                if diff_pct < 10:
                    confidence = "高"
                    recommendation = "两种方法预测总价接近，建议采用集成平均预测值。"
                elif diff_pct < 20:
                    confidence = "中"
                    recommendation = "两种方法预测总价存在一定差异，建议结合项目特点综合判断。"
                else:
                    confidence = "低"
                    recommendation = "两种方法预测总价差异较大，建议重新检查输入数据或寻求专家意见。"
                
                ws[f'A{row}'] = "机器学习集成预测:"
                ws[f'B{row}'] = f"{ensemble_prediction:,.0f} 元"
                row += 1
                
                ws[f'A{row}'] = "比率法预测:"
                ws[f'B{row}'] = f"{ratio_prediction:,.0f} 元"
                row += 1
                
                ws[f'A{row}'] = "预测差异:"
                ws[f'B{row}'] = f"{diff_pct:.1f}%"
                row += 1
                
                ws[f'A{row}'] = "可信度等级:"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True)
                ws[f'B{row}'] = confidence
                ws[f'B{row}'].font = ws[f'B{row}'].font.copy(bold=True)
                row += 1
                
                ws[f'A{row}'] = "专业建议:"
                ws[f'B{row}'] = recommendation
                row += 1
                row += 1
            
            # ===== 新增：预测总价区间建议 =====
            if ensemble_prediction and ratio_prediction:
                ws[f'A{row}'] = "预测总价区间建议"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
                row += 1
                
                mid_value = (ensemble_prediction + ratio_prediction) / 2
                min_value = min(ensemble_prediction, ratio_prediction)
                max_value = max(ensemble_prediction, ratio_prediction)
                buffer_10 = mid_value * 1.1
                buffer_15 = mid_value * 1.15
                
                ws[f'A{row}'] = "预测总价中值:"
                ws[f'B{row}'] = f"{mid_value:,.0f} 元"
                row += 1
                
                ws[f'A{row}'] = "预测总价范围:"
                ws[f'B{row}'] = f"{min_value:,.0f} - {max_value:,.0f} 元"
                row += 1
                
                ws[f'A{row}'] = "建议预算(含10%缓冲):"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True)
                ws[f'B{row}'] = f"{buffer_10:,.0f} 元"
                ws[f'B{row}'].font = ws[f'B{row}'].font.copy(bold=True)
                row += 1
                
                ws[f'A{row}'] = "保守预算(含15%缓冲):"
                ws[f'B{row}'] = f"{buffer_15:,.0f} 元"
                row += 1
                
                ws[f'A{row}'] = "预算说明:"
                ws[f'B{row}'] = "建议在预测总价基础上增加10-15%的风险缓冲"
                row += 1
                row += 1
            
            # 成本构成分析（如果有估算的各项成本数据）
            chart_data = []
            if estimated_costs and any(cost > 0 for cost in estimated_costs.values()):
                ws[f'A{row}'] = "成本构成分析"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
                row += 1
                
                # 成本构成表头
                cost_headers = ['成本项目', '预测成本(元)', '占总成本(%)', '备注']
                for col, header in enumerate(cost_headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = cell.font.copy(bold=True)
                row += 1
                
                chart_start_row = row
                total_estimated_cost = sum(estimated_costs.values())
                
                for cost_name, cost in estimated_costs.items():
                    if cost > 0:
                        percentage = (cost / total_estimated_cost * 100) if total_estimated_cost > 0 else 0
                        ws[f'A{row}'] = cost_name
                        ws[f'B{row}'] = f"{cost:,.2f}"
                        ws[f'C{row}'] = f"{percentage:.1f}%"
                        ws[f'D{row}'] = "ML模型估算"
                        chart_data.append([cost_name, cost])
                        row += 1
                
                # 添加措施费到成本构成
                if measures_cost_value > 0:
                    percentage = (measures_cost_value / total_cost * 100) if total_cost > 0 else 0
                    ws[f'A{row}'] = "措施费"
                    ws[f'B{row}'] = f"{measures_cost_value:,.2f}"
                    ws[f'C{row}'] = f"{percentage:.1f}%"
                    ws[f'D{row}'] = "额外费用"
                    chart_data.append(["措施费", measures_cost_value])
                    row += 1
                
                chart_end_row = row - 1
                row += 2
                
                # 创建成本构成饼图
                if chart_data and len(chart_data) > 1:
                    ws[f'A{row}'] = "成本构成分析图表"
                    ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
                    row += 1
                    
                    # 准备图表数据区域
                    chart_data_start = row
                    ws[f'A{row}'] = "成本项目"
                    ws[f'B{row}'] = "预测成本"
                    row += 1
                    
                    for cost_name, cost in chart_data:
                        ws[f'A{row}'] = cost_name
                        ws[f'B{row}'] = cost
                        row += 1
                    
                    chart_data_end = row - 1
                    
                    try:
                        # 创建饼图
                        chart = PieChart()
                        chart.title = "钢筋笼模式成本构成分析"
                        
                        labels = Reference(ws, min_col=1, min_row=chart_data_start+1, max_row=chart_data_end)
                        data = Reference(ws, min_col=2, min_row=chart_data_start, max_row=chart_data_end)
                        
                        chart.add_data(data, titles_from_data=True)
                        chart.set_categories(labels)
                        
                        # 设置图表样式
                        chart.width = 20
                        chart.height = 15
                        chart.dataLabels = DataLabelList()
                        chart.dataLabels.showPercent = True
                        chart.dataLabels.showVal = False
                        chart.dataLabels.showCatName = True
                        chart.dataLabels.position = 'bestFit'
                        
                        # 添加图表到工作表
                        ws.add_chart(chart, f"H{chart_data_start}")
                        
                    except Exception as chart_error:
                        logger.warning(f"创建钢筋笼模式图表时出错，跳过图表生成: {chart_error}")
                        ws[f'H{chart_data_start}'] = "图表生成失败，请检查数据完整性"
            
            # 调整列宽
            for col_num in range(1, ws.max_column + 1):
                column_letter = get_column_letter(col_num)
                max_length = 0
                for row_num in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    if hasattr(cell, 'value') and cell.value is not None and str(cell.value).strip():
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                adjusted_width = max(min(max_length + 2, 50), 10)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存到临时文件并读取
            excel_data = None
            temp_file_path = None
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                    temp_file_path = tmp_file.name
                    wb.save(tmp_file.name)
                
                wb.close()
                
                with open(temp_file_path, 'rb') as f:
                    excel_data = f.read()
                    
            finally:
                if temp_file_path and os.path.exists(temp_file_path):
                    try:
                        os.unlink(temp_file_path)
                    except Exception as e:
                        logger.warning(f"删除临时文件时出错: {e}")
            
            if excel_data is None:
                raise Exception("无法读取Excel文件内容")
            
            # 生成文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"钢筋笼模式预测报告_{timestamp}.xlsx"

            return dcc.send_bytes(excel_data, filename)
            
        except Exception as e:
            logger.error(f"导出钢筋笼模式Excel报告或保存数据失败: {e}", exc_info=True)
            raise PreventUpdate


    @app.callback(
        Output("download-steel-lining-report", "data"),
        Input("export-steel-lining-report-btn", "n_clicks"),
        State("steel-lining-report-data", "data"),
        prevent_initial_call=True
    )
    def export_steel_lining_excel(n_clicks, report_data):
        """导出钢衬里模式报告为Excel（合并表格+扇形图）"""
        if not n_clicks or not report_data:
            raise PreventUpdate      
        try:
            from openpyxl import Workbook
            from openpyxl.chart import PieChart, Reference
            from openpyxl.chart.label import DataLabelList
            from openpyxl.utils import get_column_letter
            import tempfile
            import os
            
            # 数据完整性检查
            if not report_data:
                raise Exception("报告数据为空，无法生成Excel文件")
            
            quantities = report_data.get('工程量数据', {})
            ml_prediction_results = report_data.get('预测结果', {})
            measures_cost_value = report_data.get('措施费', 0)
            
            # 获取预测结果
            ml_predictions = ml_prediction_results.get("机器学习预测结果", {}) if ml_prediction_results else {}
            ratio_prediction = ml_prediction_results.get("比率法预测总价") if ml_prediction_results else None
            estimated_costs = ml_prediction_results.get("估算的各项成本 (用于ML的特征)", {}) if ml_prediction_results else {}
            
            # 计算总预测价格
            ensemble_prediction = ml_predictions.get('集成平均预测') if ml_predictions else None
            total_cost = ensemble_prediction or ratio_prediction or 0
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "钢衬里模式预测报告"
            
            # 设置标题
            ws['A1'] = "钢衬里施工模式智能预测分析报告"
            ws['A1'].font = ws['A1'].font.copy(bold=True, size=16)
            ws.merge_cells('A1:E1')
            
            # 基本信息
            row = 3
            ws[f'A{row}'] = "报告概述"
            ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
            row += 1
            
            ws[f'A{row}'] = "施工模式:"
            ws[f'B{row}'] = report_data.get('模式', '钢衬里施工模式')
            row += 1
            ws[f'A{row}'] = "生成时间:"
            ws[f'B{row}'] = report_data.get('生成时间', '')
            row += 1
            ws[f'A{row}'] = "预测总价(元):"
            ws[f'B{row}'] = f"{total_cost:,.0f}"
            row += 2
            
            # 工程量数据
            ws[f'A{row}'] = "工程量数据详情"
            ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
            row += 1
            
            # 工程量表头
            quantity_headers = ['工程量项目', '数量', '单位', '备注']
            for col, header in enumerate(quantity_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = cell.font.copy(bold=True)
            row += 1
            
            # 工程量数据
            quantity_items = [
                ('拼装场地总工程量', quantities.get('拼装场地总工程量', 0), '立方米'),
                ('制作胎具总工程量', quantities.get('制作胎具总工程量', 0), '吨'),
                ('钢支墩埋件总工程量', quantities.get('钢支墩埋件总工程量', 0), '吨'),
                ('扶壁柱总工程量', quantities.get('扶壁柱总工程量', 0), '吨'),
                ('走道板操作平台总工程量', quantities.get('走道板操作平台总工程量', 0), '吨'),
                ('钢网梁总工程量', quantities.get('钢网梁总工程量', 0), '吨'),
            ]
            
            for item_name, quantity, unit in quantity_items:
                ws[f'A{row}'] = item_name
                ws[f'B{row}'] = f"{quantity:.2f}"
                ws[f'C{row}'] = unit
                ws[f'D{row}'] = "用户输入或智能估算"
                row += 1
            
            if measures_cost_value > 0:
                ws[f'A{row}'] = "措施费"
                ws[f'B{row}'] = f"{measures_cost_value:,.2f}"
                ws[f'C{row}'] = "元"
                ws[f'D{row}'] = "额外费用"
                row += 1
            row += 1
            
            # 预测结果
            ws[f'A{row}'] = "预测结果分析"
            ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
            row += 1
            
            # 结果表头
            result_headers = ['预测方法', '预测总价(元)', '可信度', '备注']
            for col, header in enumerate(result_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = cell.font.copy(bold=True)
            row += 1
            
# 预测结果数据 - 适配新的状态结构（钢衬里模式）
            prediction_status = report_data.get('预测方法状态', {})
            
            # AI预测结果处理
            main_ai_result = report_data.get('主要AI预测', {})
            ai_status = main_ai_result.get('status', 'unknown')
            ai_value = main_ai_result.get('value')
            
            if ai_status == 'success' and isinstance(ai_value, (int, float)):
                ws[f'A{row}'] = "AI预测（机器学习集成）"
                ws[f'B{row}'] = f"{ai_value:,.0f}"
                ws[f'C{row}'] = "高"
                ws[f'D{row}'] = "钢衬里模式多算法集成结果，预测成功"
                row += 1
            elif ai_status == 'display_disabled':
                ws[f'A{row}'] = "AI预测（机器学习集成）"
                ws[f'B{row}'] = "显示被禁用"
                ws[f'C{row}'] = "不适用"
                ws[f'D{row}'] = main_ai_result.get('message', '钢衬里模式AI预测显示权限被禁用')
                row += 1
            elif ai_status == 'error':
                ws[f'A{row}'] = "AI预测（机器学习集成）"
                ws[f'B{row}'] = "预测失败"
                ws[f'C{row}'] = "低"
                ws[f'D{row}'] = f"钢衬里模式AI预测错误: {main_ai_result.get('message', '未知错误')}"
                row += 1
            else:
                ws[f'A{row}'] = "AI预测（机器学习集成）"
                ws[f'B{row}'] = "不可用"
                ws[f'C{row}'] = "无"
                ws[f'D{row}'] = f"钢衬里模式AI预测状态: {ai_status}, 消息: {main_ai_result.get('message', '未知状态')}"
                row += 1
            
            # 比率法预测结果处理
            main_ratio_result = report_data.get('主要比率法预测', {})
            ratio_status = main_ratio_result.get('status', 'unknown')
            ratio_value = main_ratio_result.get('value')
            
            if ratio_status == 'success' and isinstance(ratio_value, (int, float)):
                ws[f'A{row}'] = "比率法预测"
                ws[f'B{row}'] = f"{ratio_value:,.0f}"
                ws[f'C{row}'] = "中"
                ws[f'D{row}'] = "钢衬里模式传统估算方法，预测成功"
                row += 1
            elif ratio_status == 'display_disabled':
                ws[f'A{row}'] = "比率法预测"
                ws[f'B{row}'] = "显示被禁用"
                ws[f'C{row}'] = "不适用"
                ws[f'D{row}'] = main_ratio_result.get('message', '钢衬里模式比率法预测显示权限被禁用')
                row += 1
            elif ratio_status == 'error':
                ws[f'A{row}'] = "比率法预测"
                ws[f'B{row}'] = "预测失败"
                ws[f'C{row}'] = "低"
                ws[f'D{row}'] = f"钢衬里模式比率法预测错误: {main_ratio_result.get('message', '未知错误')}"
                row += 1
            else:
                ws[f'A{row}'] = "比率法预测"
                ws[f'B{row}'] = "不可用"
                ws[f'C{row}'] = "无"
                ws[f'D{row}'] = f"钢衬里模式比率法预测状态: {ratio_status}, 消息: {main_ratio_result.get('message', '未知状态')}"
                row += 1
            row += 1
            
            # ===== 新增：钢衬里模式算法执行状态分析 =====
            algorithm_status = report_data.get('算法执行状态', {})
            if algorithm_status:
                ws[f'A{row}'] = "钢衬里模式算法执行状态分析"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
                row += 1
                
                ws[f'A{row}'] = "AI算法可用性:"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True)
                can_execute = algorithm_status.get('can_execute_ai', False)
                ws[f'B{row}'] = "可用" if can_execute else "不可用"
                ws[f'B{row}'].font = ws[f'B{row}'].font.copy(bold=True)
                row += 1
                
                ws[f'A{row}'] = "启用算法数量:"
                enabled_count = algorithm_status.get('enabled_count', 0)
                total_count = algorithm_status.get('total_count', 0)
                ws[f'B{row}'] = f"{enabled_count}/{total_count}"
                row += 1
                
                ws[f'A{row}'] = "算法状态消息:"
                ws[f'B{row}'] = algorithm_status.get('message', '无消息')
                row += 1
                
                ws[f'A{row}'] = "可用算法列表:"
                enabled_algorithms = algorithm_status.get('enabled_algorithms', [])
                ws[f'B{row}'] = ', '.join(enabled_algorithms) if enabled_algorithms else '钢衬里模式无可用算法'
                row += 1
                
                ws[f'A{row}'] = "钢衬里模式特殊说明:"
                ws[f'B{row}'] = "钢衬里模式使用与钢筋笼模式相同的算法配置，但数据特征不同"
                row += 1
                row += 1
            
            # ===== 新增：钢衬里模式预测方法详细状态 =====
            method_status = report_data.get('预测方法状态', {})
            if method_status:
                ws[f'A{row}'] = "钢衬里模式预测方法详细状态"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
                row += 1
                
                # 状态表头
                status_headers = ['预测方法', '最终状态', '可执行', '可显示', '执行消息', '显示消息']
                for col, header in enumerate(status_headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = cell.font.copy(bold=True)
                row += 1
                
                # 状态详情
                status_mapping = {
                    'fully_available': '完全可用',
                    'execute_only': '可执行但显示禁用',
                    'display_error': '显示启用但执行失败',
                    'fully_disabled': '完全禁用',
                    'unknown': '状态未知'
                }
                
                for method_key, method_info in method_status.items():
                    method_name = method_info.get('name', method_key)
                    final_status = method_info.get('final_status', 'unknown')
                    can_execute = method_info.get('can_execute', False)
                    can_display = method_info.get('can_display', False)
                    execution_msg = method_info.get('execution_message', '')
                    display_msg = method_info.get('display_message', '')
                    
                    ws[f'A{row}'] = f"钢衬里-{method_name}"
                    ws[f'B{row}'] = status_mapping.get(final_status, final_status)
                    ws[f'C{row}'] = "是" if can_execute else "否"
                    ws[f'D{row}'] = "是" if can_display else "否"
                    ws[f'E{row}'] = execution_msg[:50] + "..." if len(execution_msg) > 50 else execution_msg
                    ws[f'F{row}'] = display_msg[:50] + "..." if len(display_msg) > 50 else display_msg
                    row += 1
                row += 1
            
            # ===== 新增：钢衬里模式综合状态汇总 =====
            prediction_status_summary = format_prediction_status_for_export(report_data.get('预测结果', {}))
            if prediction_status_summary:
                ws[f'A{row}'] = "钢衬里模式综合状态汇总"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
                row += 1
                
                ws[f'A{row}'] = "状态概览:"
                ws[f'B{row}'] = prediction_status_summary.get('状态概览', '无信息')
                row += 1
                
                ws[f'A{row}'] = "完全可用方法数:"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True)
                ws[f'B{row}'] = str(prediction_status_summary.get('可用方法数量', 0))
                ws[f'B{row}'].font = ws[f'B{row}'].font.copy(bold=True)
                row += 1
                
                ws[f'A{row}'] = "算法执行能力:"
                ws[f'B{row}'] = prediction_status_summary.get('算法执行能力', '未知')
                row += 1
                
                ws[f'A{row}'] = "钢衬里模式特征:"
                ws[f'B{row}'] = "基于拼装场地、制作胎具、钢支墩埋件、扶壁柱、走道板平台、钢网梁等六大要素"
                row += 1
                
                ws[f'A{row}'] = "系统建议:"
                if prediction_status_summary.get('可用方法数量', 0) == 4:
                    suggestion = "钢衬里模式系统运行正常，所有预测方法都可用，可进行全面分析"
                elif prediction_status_summary.get('可用方法数量', 0) > 0:
                    suggestion = "钢衬里模式系统部分正常，建议检查相关配置以启用更多预测方法"
                else:
                    suggestion = "钢衬里模式系统存在问题，建议检查算法配置和综合指标设置"
                ws[f'B{row}'] = suggestion
                row += 1
                
                ws[f'A{row}'] = "数据来源说明:"
                ws[f'B{row}'] = "钢衬里模式使用 key_factors_2 和 price_baseline_2 表的历史数据"
                row += 1
                row += 1

            # ===== 钢衬里模式特有的成本构成分析 =====
            if estimated_costs and any(cost > 0 for cost in estimated_costs.values()):
                ws[f'A{row}'] = "钢衬里模式成本构成详情"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
                row += 1
                
                # 钢衬里模式特有的成本项分析
                steel_lining_cost_items = [
                    "拼装场地费用", "制作胎具费用", "钢支墩、埋件费用", 
                    "扶壁柱费用", "走道板及操作平台费用", "钢网架费用"
                ]
                
                ws[f'A{row}'] = "成本项目类型:"
                ws[f'B{row}'] = "钢衬里施工六大费用要素"
                row += 1
                
                ws[f'A{row}'] = "主要成本构成:"
                cost_breakdown = []
                total_estimated_cost = sum(estimated_costs.values())
                
                for item in steel_lining_cost_items:
                    if item in estimated_costs and estimated_costs[item] > 0:
                        percentage = (estimated_costs[item] / total_estimated_cost * 100) if total_estimated_cost > 0 else 0
                        cost_breakdown.append(f"{item}: {percentage:.1f}%")
                
                ws[f'B{row}'] = "; ".join(cost_breakdown) if cost_breakdown else "成本构成数据不完整"
                row += 1
                
                ws[f'A{row}'] = "成本分析建议:"
                if cost_breakdown:
                    # 找出占比最大的成本项
                    max_cost_item = max(steel_lining_cost_items, 
                                      key=lambda x: estimated_costs.get(x, 0), 
                                      default="未知")
                    max_cost_value = estimated_costs.get(max_cost_item, 0)
                    max_percentage = (max_cost_value / total_estimated_cost * 100) if total_estimated_cost > 0 else 0
                    
                    if max_percentage > 40:
                        suggestion = f"建议重点关注{max_cost_item}的成本控制，该项占总成本{max_percentage:.1f}%"
                    elif max_percentage > 25:
                        suggestion = f"{max_cost_item}是主要成本项(占{max_percentage:.1f}%)，需适当关注"
                    else:
                        suggestion = "钢衬里模式各成本项分布相对均衡，建议整体优化"
                else:
                    suggestion = "成本分析数据不足，建议完善工程量输入"
                    
                ws[f'B{row}'] = suggestion
                row += 1
                row += 1

            
            
            # ===== 新增：最佳模型分析 =====
            if ml_predictions and ratio_prediction:
                ws[f'A{row}'] = "最佳模型分析"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
                row += 1
                
                # 找出最佳模型（最接近比率法的模型）
                optimal_model_key = None
                algorithm_names = {
                    'LinearRegression': '线性回归算法',
                    'DecisionTree': '决策树算法',
                    'RandomForest': '随机森林算法',
                    'SupportVectorRegression': '支持向量机算法',
                    'NeuralNetwork': '神经网络算法',
                    "岭回归 (RidgeCV)": "岭回归算法"
                }
                min_diff_to_ratio = float('inf')  # 添加这行：初始化变量
                for model_key_iter, model_pred_val in ml_predictions.items():
                    if (model_key_iter != '集成平均预测' and 
                        model_pred_val is not None and 
                        isinstance(model_pred_val, (int, float))):  # 添加类型检查
                        diff = abs(model_pred_val - ratio_prediction)
                        if diff < min_diff_to_ratio:
                            min_diff_to_ratio = diff
                            optimal_model_key = model_key_iter
                
                if optimal_model_key and ml_predictions.get(optimal_model_key) is not None:
                    model_pred_val = ml_predictions[optimal_model_key]
                    if isinstance(model_pred_val, (int, float)):
                        deviation_val = model_pred_val - ratio_prediction if ratio_prediction else 0
                        deviation_pct = (deviation_val / ratio_prediction * 100) if ratio_prediction != 0 else 0
                        
                        ws[f'A{row}'] = "最佳单一模型:"
                        ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True)
                        ws[f'B{row}'] = f"{algorithm_names.get(optimal_model_key, optimal_model_key)}"
                        row += 1
                        
                        ws[f'A{row}'] = "最佳模型预测价格:"
                        ws[f'B{row}'] = f"{model_pred_val:,.0f} 元"
                        row += 1
                        
                        ws[f'A{row}'] = "与比率法偏差:"
                        ws[f'B{row}'] = f"{deviation_val:,.0f} 元 ({deviation_pct:.1f}%)"
                        row += 1
                        
                        ws[f'A{row}'] = "模型特点:"
                        ws[f'B{row}'] = "基于历史项目数据训练，最接近传统比率法预测"
                        row += 1
                row += 1
            
            # ===== 新增：预测可信度评估 =====
            if ensemble_prediction and ratio_prediction:
                ws[f'A{row}'] = "预测可信度评估"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
                row += 1
                
                # 计算差异百分比
                diff_pct = abs(ensemble_prediction - ratio_prediction) / max(ensemble_prediction, ratio_prediction) * 100 if max(ensemble_prediction, ratio_prediction) != 0 else 0
                
                if diff_pct < 10:
                    confidence = "高"
                    recommendation = "两种方法预测总价接近，建议采用集成平均预测值。"
                elif diff_pct < 20:
                    confidence = "中"
                    recommendation = "两种方法预测总价存在一定差异，建议结合项目特点综合判断。"
                else:
                    confidence = "低"
                    recommendation = "两种方法预测总价差异较大，建议重新检查输入数据或寻求专家意见。"
                
                ws[f'A{row}'] = "机器学习集成预测:"
                ws[f'B{row}'] = f"{ensemble_prediction:,.0f} 元"
                row += 1
                
                ws[f'A{row}'] = "比率法预测:"
                ws[f'B{row}'] = f"{ratio_prediction:,.0f} 元"
                row += 1
                
                ws[f'A{row}'] = "预测差异:"
                ws[f'B{row}'] = f"{diff_pct:.1f}%"
                row += 1
                
                ws[f'A{row}'] = "可信度等级:"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True)
                ws[f'B{row}'] = confidence
                ws[f'B{row}'].font = ws[f'B{row}'].font.copy(bold=True)
                row += 1
                
                ws[f'A{row}'] = "专业建议:"
                ws[f'B{row}'] = recommendation
                row += 1
                row += 1
            
            # ===== 新增：预测总价区间建议 =====
            if ensemble_prediction and ratio_prediction:
                ws[f'A{row}'] = "预测总价区间建议"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
                row += 1
                
                mid_value = (ensemble_prediction + ratio_prediction) / 2
                min_value = min(ensemble_prediction, ratio_prediction)
                max_value = max(ensemble_prediction, ratio_prediction)
                buffer_10 = mid_value * 1.1
                buffer_15 = mid_value * 1.15
                
                ws[f'A{row}'] = "预测总价中值:"
                ws[f'B{row}'] = f"{mid_value:,.0f} 元"
                row += 1
                
                ws[f'A{row}'] = "预测总价范围:"
                ws[f'B{row}'] = f"{min_value:,.0f} - {max_value:,.0f} 元"
                row += 1
                
                ws[f'A{row}'] = "建议预算(含10%缓冲):"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True)
                ws[f'B{row}'] = f"{buffer_10:,.0f} 元"
                ws[f'B{row}'].font = ws[f'B{row}'].font.copy(bold=True)
                row += 1
                
                ws[f'A{row}'] = "保守预算(含15%缓冲):"
                ws[f'B{row}'] = f"{buffer_15:,.0f} 元"
                row += 1
                
                ws[f'A{row}'] = "预算说明:"
                ws[f'B{row}'] = "建议在预测总价基础上增加10-15%的风险缓冲"
                row += 1
                row += 1
            
            # 成本构成分析（如果有估算的各项成本数据）
            chart_data = []
            if estimated_costs and any(cost > 0 for cost in estimated_costs.values()):
                ws[f'A{row}'] = "成本构成分析"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
                row += 1
                
                # 成本构成表头
                cost_headers = ['成本项目', '预测成本(元)', '占总成本(%)', '备注']
                for col, header in enumerate(cost_headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = cell.font.copy(bold=True)
                row += 1
                
                chart_start_row = row
                total_estimated_cost = sum(estimated_costs.values())
                
                for cost_name, cost in estimated_costs.items():
                    if cost > 0:
                        percentage = (cost / total_estimated_cost * 100) if total_estimated_cost > 0 else 0
                        ws[f'A{row}'] = cost_name
                        ws[f'B{row}'] = f"{cost:,.2f}"
                        ws[f'C{row}'] = f"{percentage:.1f}%"
                        ws[f'D{row}'] = "ML模型估算"
                        chart_data.append([cost_name, cost])
                        row += 1
                
                # 添加措施费到成本构成
                if measures_cost_value > 0:
                    percentage = (measures_cost_value / total_cost * 100) if total_cost > 0 else 0
                    ws[f'A{row}'] = "措施费"
                    ws[f'B{row}'] = f"{measures_cost_value:,.2f}"
                    ws[f'C{row}'] = f"{percentage:.1f}%"
                    ws[f'D{row}'] = "额外费用"
                    chart_data.append(["措施费", measures_cost_value])
                    row += 1
                
                chart_end_row = row - 1
                row += 2
                
                # 创建成本构成饼图
                if chart_data and len(chart_data) > 1:
                    ws[f'A{row}'] = "成本构成分析图表"
                    ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
                    row += 1
                    
                    # 准备图表数据区域
                    chart_data_start = row
                    ws[f'A{row}'] = "成本项目"
                    ws[f'B{row}'] = "预测成本"
                    row += 1
                    
                    for cost_name, cost in chart_data:
                        ws[f'A{row}'] = cost_name
                        ws[f'B{row}'] = cost
                        row += 1
                    
                    chart_data_end = row - 1
                    
                    try:
                        # 创建钢衬里模式专用饼图
                        chart = PieChart()
                        chart.title = "钢衬里模式成本构成分析"
                        
                        labels = Reference(ws, min_col=1, min_row=chart_data_start+1, max_row=chart_data_end)
                        data = Reference(ws, min_col=2, min_row=chart_data_start, max_row=chart_data_end)
                        
                        chart.add_data(data, titles_from_data=True)
                        chart.set_categories(labels)
                        
                        # 钢衬里模式图表样式设置
                        chart.width = 20
                        chart.height = 15
                        chart.dataLabels = DataLabelList()
                        chart.dataLabels.showPercent = True
                        chart.dataLabels.showVal = False
                        chart.dataLabels.showCatName = True
                        chart.dataLabels.position = 'bestFit'
                        
                        # 添加图表到工作表
                        ws.add_chart(chart, f"H{chart_data_start}")
                        
                        # 添加钢衬里模式图表说明
                        ws[f"H{chart_data_start + 20}"] = "图表说明：钢衬里模式基于六大费用要素的成本分布"
                        
                    except Exception as chart_error:
                        logger.warning(f"创建钢衬里模式图表时出错，跳过图表生成: {chart_error}")
                        ws[f'H{chart_data_start}'] = "钢衬里模式图表生成失败，请检查数据完整性"
                        ws[f'H{chart_data_start + 1}'] = "可能原因：成本数据不完整或Excel组件限制"
            
            # 调整列宽
            for col_num in range(1, ws.max_column + 1):
                column_letter = get_column_letter(col_num)
                max_length = 0
                for row_num in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    if hasattr(cell, 'value') and cell.value is not None and str(cell.value).strip():
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                adjusted_width = max(min(max_length + 2, 50), 10)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存到临时文件并读取
            excel_data = None
            temp_file_path = None
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                    temp_file_path = tmp_file.name
                    wb.save(tmp_file.name)
                
                wb.close()
                
                with open(temp_file_path, 'rb') as f:
                    excel_data = f.read()
                    
            finally:
                if temp_file_path and os.path.exists(temp_file_path):
                    try:
                        os.unlink(temp_file_path)
                    except Exception as e:
                        logger.warning(f"删除临时文件时出错: {e}")
            
            if excel_data is None:
                raise Exception("无法读取Excel文件内容")
            
            # 生成钢衬里模式专用文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"钢衬里模式预测报告_{timestamp}.xlsx"

            return dcc.send_bytes(excel_data, filename)
            
        except Exception as e:
            logger.error(f"导出钢衬里模式Excel报告或保存数据失败: {e}", exc_info=True)
            raise PreventUpdate
        
    @app.callback(
        Output("download-custom-mode-report", "data"),
        Input("export-custom-mode-report-btn", "n_clicks"),
        State("custom-mode-report-data", "data"),
        prevent_initial_call=True
    )
    def export_custom_mode_excel(n_clicks, report_data):
        """导出自定义模式报告为Excel（合并表格+扇形图）"""
        if not n_clicks or not report_data:
            raise PreventUpdate      
        try:
            from openpyxl import Workbook
            from openpyxl.chart import PieChart, Reference
            from openpyxl.chart.label import DataLabelList
            from openpyxl.utils import get_column_letter
            import tempfile
            import os
            
            # 数据完整性检查
            if not report_data:
                raise Exception("报告数据为空，无法生成Excel文件")
            
            total_prediction = report_data.get('预测结果', {})
            total_cost = total_prediction.get('total_predicted_cost', 0)
            input_params = report_data.get('输入参数', [])
            param_costs = total_prediction.get('param_costs', {})
            
            # 检查关键数据
            if total_cost <= 0:
                logger.warning("预测总价为0或负数，可能影响报告质量")
            
            if not input_params:
                logger.warning("输入参数为空，报告内容可能不完整")
            
            if not param_costs:
                logger.warning("参数成本数据为空，无法生成成本分析")
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "自定义模式预测报告"
            
            # 设置标题
            ws['A1'] = "自定义模式智能预测分析报告"
            ws['A1'].font = ws['A1'].font.copy(bold=True, size=16)
            ws.merge_cells('A1:E1')
            
            # 基本信息
            row = 3
            ws[f'A{row}'] = "报告概述"
            ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
            row += 1
            
            ws[f'A{row}'] = "施工模式:"
            ws[f'B{row}'] = report_data.get('模式', '自定义模式')
            row += 1
            ws[f'A{row}'] = "生成时间:"
            ws[f'B{row}'] = report_data.get('生成时间', '')
            row += 1
            
            total_prediction = report_data.get('预测结果', {})
            total_cost = total_prediction.get('total_predicted_cost', 0)
            ws[f'A{row}'] = "预测总价(元):"
            ws[f'B{row}'] = f"{total_cost:,.0f}"
            row += 2
            
            # 输入参数
            ws[f'A{row}'] = "输入参数详情"
            ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
            row += 1
            
            # 参数表头
            param_headers = ['参数名称', '工程量', '工程量占比(%)', '价格量(元)', '价格占比(%)', '关键因素']
            for col, header in enumerate(param_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = cell.font.copy(bold=True)
            row += 1
            
            # 参数数据
            if input_params:
                for param in input_params:
                    ws[f'A{row}'] = param.get('name', '未知参数')
                    ws[f'B{row}'] = f"{param.get('quantity', 0):.2f}"
                    ws[f'C{row}'] = f"{param.get('quantity_ratio', 0):.2f}"
                    ws[f'D{row}'] = f"{param.get('price_amount', 0):,.2f}"
                    ws[f'E{row}'] = f"{param.get('price_ratio', 0):.2f}"
                    ws[f'F{row}'] = param.get('key_factor', '')
                    row += 1
            else:
                # 如果没有输入参数，添加提示信息
                ws[f'A{row}'] = "暂无输入参数"
                ws[f'B{row}'] = "0.00"
                ws[f'C{row}'] = "0.00"
                ws[f'D{row}'] = "0.00"
                ws[f'E{row}'] = "0.00"
                ws[f'F{row}'] = "请检查参数输入"
                row += 1
            row += 1
            
            # ===== 新增：自定义模式预测方法说明 =====
            ws[f'A{row}'] = "预测方法说明"
            ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
            row += 1
            
            ws[f'A{row}'] = "预测模式:"
            ws[f'B{row}'] = "自定义参数智能估算"
            row += 1
            
            ws[f'A{row}'] = "计算方法:"
            ws[f'B{row}'] = "基于用户输入参数和占比关系进行成本估算"
            row += 1
            
            ws[f'A{row}'] = "预测总价:"
            ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True)
            ws[f'B{row}'] = f"{total_cost:,.0f} 元"
            ws[f'B{row}'].font = ws[f'B{row}'].font.copy(bold=True)
            row += 1
            
            ws[f'A{row}'] = "可信度评估:"
            if total_cost > 0 and input_params and len(input_params) >= 2:
                confidence = "中等"
                suggestion = "基于用户自定义参数，建议结合项目实际情况验证"
            elif total_cost > 0:
                confidence = "较低"
                suggestion = "参数较少，建议增加更多参数提高预测准确性"
            else:
                confidence = "低"
                suggestion = "预测总价异常，请检查参数输入的合理性"
            
            ws[f'B{row}'] = confidence
            row += 1
            
            ws[f'A{row}'] = "专业建议:"
            ws[f'B{row}'] = suggestion
            row += 2
            
            # ===== 新增：自定义模式预算建议 =====
            if total_cost > 0:
                ws[f'A{row}'] = "预算建议"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
                row += 1
                
                buffer_10 = total_cost * 1.1
                buffer_15 = total_cost * 1.15
                buffer_20 = total_cost * 1.2
                
                ws[f'A{row}'] = "基准预算:"
                ws[f'B{row}'] = f"{total_cost:,.0f} 元"
                row += 1
                
                ws[f'A{row}'] = "建议预算(含10%缓冲):"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True)
                ws[f'B{row}'] = f"{buffer_10:,.0f} 元"
                ws[f'B{row}'].font = ws[f'B{row}'].font.copy(bold=True)
                row += 1
                
                ws[f'A{row}'] = "保守预算(含15%缓冲):"
                ws[f'B{row}'] = f"{buffer_15:,.0f} 元"
                row += 1
                
                ws[f'A{row}'] = "风险预算(含20%缓冲):"
                ws[f'B{row}'] = f"{buffer_20:,.0f} 元"
                row += 1
                
                ws[f'A{row}'] = "缓冲说明:"
                ws[f'B{row}'] = "自定义模式建议增加15-20%缓冲以应对不确定性"
                row += 2
            
            # 预测结果
            ws[f'A{row}'] = "预测结果分析"
            ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
            row += 1
            
            # 结果表头
            result_headers = ['参数名称', '预测成本(元)', '占总成本(%)', '备注']
            for col, header in enumerate(result_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = cell.font.copy(bold=True)
            row += 1
            
            # 结果数据
            chart_data = []  # 初始化图表数据列表
            chart_start_row = row  # 记录图表数据开始行
            
            if param_costs and any(cost > 0 for cost in param_costs.values()):
                for param_name, cost in param_costs.items():
                    if cost > 0:
                        percentage = (cost / total_cost * 100) if total_cost > 0 else 0
                        ws[f'A{row}'] = param_name
                        ws[f'B{row}'] = f"{cost:,.2f}"
                        ws[f'C{row}'] = f"{percentage:.1f}%"
                        ws[f'D{row}'] = "智能估算"
                        chart_data.append([param_name, cost])
                        row += 1
            else:
                # 如果没有成本数据，添加提示信息
                ws[f'A{row}'] = "暂无成本预测数据"
                ws[f'B{row}'] = "0.00"
                ws[f'C{row}'] = "0.0%"
                ws[f'D{row}'] = "请检查参数配置"
                row += 1
            
            # 总计行
            ws[f'A{row}'] = "总计"
            ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True)
            ws[f'B{row}'] = f"{total_cost:,.0f}"
            ws[f'B{row}'].font = ws[f'B{row}'].font.copy(bold=True)
            ws[f'C{row}'] = "100.0%"
            ws[f'C{row}'].font = ws[f'C{row}'].font.copy(bold=True)
            ws[f'D{row}'] = "预测总价"
            ws[f'D{row}'].font = ws[f'D{row}'].font.copy(bold=True)
            chart_end_row = row - 1
            row += 2
            
            # 创建成本构成饼图 - 只有在有有效数据时才创建
            if chart_data and len(chart_data) > 1 and total_cost > 0:
                ws[f'A{row}'] = "成本构成分析图表"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
                row += 1
                
                # 准备图表数据区域
                chart_data_start = row
                ws[f'A{row}'] = "参数名称"
                ws[f'B{row}'] = "预测成本"
                row += 1
                
                for param_name, cost in chart_data:
                    ws[f'A{row}'] = param_name
                    ws[f'B{row}'] = cost
                    row += 1
                
                chart_data_end = row - 1
                
                try:
                    # 创建饼图
                    chart = PieChart()
                    chart.title = "成本构成分析"
                    
                    labels = Reference(ws, min_col=1, min_row=chart_data_start+1, max_row=chart_data_end)
                    data = Reference(ws, min_col=2, min_row=chart_data_start, max_row=chart_data_end)
                    
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(labels)
                    
                    # 设置图表样式 - 放大图表并优化标签
                    chart.width = 20  # 放大宽度
                    chart.height = 15  # 放大高度
                    chart.dataLabels = DataLabelList()
                    chart.dataLabels.showPercent = True
                    chart.dataLabels.showVal = False  # 不显示具体数值，避免重叠
                    chart.dataLabels.showCatName = True  # 显示类别名称
                    chart.dataLabels.position = 'bestFit'  # 自动最佳位置
                    
                    # 添加图表到工作表
                    ws.add_chart(chart, f"H{chart_data_start}")
                    
                except Exception as chart_error:
                    logger.warning(f"创建图表时出错，跳过图表生成: {chart_error}")
                    # 添加提示信息
                    ws[f'H{chart_data_start}'] = "图表生成失败，请检查数据完整性"
                
                # 添加智能分析建议
                row += 2
                ws[f'A{row}'] = "智能分析建议"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
                row += 1
                
                # 找出成本最高的参数
                if param_costs:
                    try:
                        max_cost_param = max(param_costs.items(), key=lambda x: x[1] if x[1] > 0 else 0)
                        max_cost_percentage = (max_cost_param[1] / total_cost * 100) if total_cost > 0 else 0
                        
                        ws[f'A{row}'] = "主要成本项:"
                        ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True)
                        ws[f'B{row}'] = f"{max_cost_param[0]} (占{max_cost_percentage:.1f}%)"
                        row += 1
                        
                        ws[f'A{row}'] = "成本优化建议:"
                        ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True)
                        if max_cost_percentage > 50:
                            suggestion = "主要成本项占比较高，建议重点关注该项目的成本控制"
                        elif max_cost_percentage > 30:
                            suggestion = "成本分布相对集中，建议平衡各项成本投入"
                        else:
                            suggestion = "成本分布较为均衡，建议保持当前配置"
                        ws[f'B{row}'] = suggestion
                        row += 1
                        
                        # 自定义模式特有的建议
                        ws[f'A{row}'] = "自定义模式建议:"
                        ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True)
                        ws[f'B{row}'] = "建议结合实际项目经验，对关键参数进行细化和验证"
                        row += 1
                        
                    except Exception as analysis_error:
                        logger.warning(f"生成分析建议时出错: {analysis_error}")
                        ws[f'A{row}'] = "分析建议生成失败，请检查数据完整性"
                        row += 1
            else:
                # 如果没有足够的数据创建图表，添加说明
                ws[f'A{row}'] = "数据不足，无法生成成本构成图表"
                ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=14)
                row += 1
                ws[f'A{row}'] = "建议："
                ws[f'B{row}'] = "1. 检查输入参数是否完整"
                row += 1
                ws[f'A{row}'] = ""
                ws[f'B{row}'] = "2. 确保至少有2个有效的成本项目"
                row += 1
                ws[f'A{row}'] = ""
                ws[f'B{row}'] = "3. 检查预测总价是否大于0"
                row += 1
            
            # 调整列宽 - 修复这里的问题
            for col_num in range(1, ws.max_column + 1):
                column_letter = get_column_letter(col_num)  # 直接使用 get_column_letter 函数
                max_length = 0
                for row_num in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    # 检查单元格是否是合并单元格的一部分
                    if hasattr(cell, 'value') and cell.value is not None and str(cell.value).strip():
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                # 设置最小宽度为10，最大宽度为50
                adjusted_width = max(min(max_length + 2, 50), 10)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存到临时文件并读取
            excel_data = None
            temp_file_path = None
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                    temp_file_path = tmp_file.name
                    wb.save(tmp_file.name)
                
                # 确保工作簿已关闭
                wb.close()
                
                # 读取文件内容
                with open(temp_file_path, 'rb') as f:
                    excel_data = f.read()
                    
            finally:
                # 安全删除临时文件
                if temp_file_path and os.path.exists(temp_file_path):
                    try:
                        os.unlink(temp_file_path)
                    except PermissionError:
                        # 如果无法删除，记录警告但不影响主要功能
                        logger.warning(f"无法删除临时文件: {temp_file_path}")
                    except Exception as e:
                        logger.warning(f"删除临时文件时出错: {e}")
            
            if excel_data is None:
                raise Exception("无法读取Excel文件内容")
            
            # 生成文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"自定义模式预测报告_{timestamp}.xlsx"
        
            return dcc.send_bytes(excel_data, filename)
            
        except Exception as e:
            logger.error(f"导出自定义模式Excel报告或保存数据失败: {e}", exc_info=True)
            raise PreventUpdate
        


    # ========== 新增：保存数据到数据库的回调函数 ==========
    
    @app.callback(
        Output("steel-cage-save-feedback", "children"),
        Input("save-steel-cage-data-btn", "n_clicks"),
        State("steel-cage-report-data", "data"),
        prevent_initial_call=True
    )
    def save_steel_cage_data_to_database(n_clicks, report_data):
        """保存钢筋笼模式数据到数据库"""
        if not n_clicks or not report_data:
            raise PreventUpdate
        
        try:
            # 调用已有的保存函数
            save_success = save_report_to_database(report_data, "steel_cage")
            
            if save_success:
                return dbc.Alert([
                    html.I(className="fas fa-check-circle me-2"),
                    "钢筋笼模式预测数据已成功保存到数据库！"
                ], color="success", duration=4000)
            else:
                return dbc.Alert([
                    html.I(className="fas fa-exclamation-triangle me-2"),
                    "保存失败：数据库操作异常，请检查数据库连接状态。"
                ], color="danger", duration=6000)
                
        except Exception as e:
            logger.error(f"钢筋笼模式数据保存异常: {e}", exc_info=True)
            return dbc.Alert([
                html.I(className="fas fa-exclamation-triangle me-2"),
                f"保存过程中发生错误：{str(e)}"
            ], color="danger", duration=6000)

    @app.callback(
        Output("steel-lining-save-feedback", "children"),
        Input("save-steel-lining-data-btn", "n_clicks"),
        State("steel-lining-report-data", "data"),
        prevent_initial_call=True
    )
    def save_steel_lining_data_to_database(n_clicks, report_data):
        """保存钢衬里模式数据到数据库"""
        if not n_clicks or not report_data:
            raise PreventUpdate
        
        try:
            # 调用已有的保存函数
            save_success = save_report_to_database(report_data, "steel_lining")
            
            if save_success:
                return dbc.Alert([
                    html.I(className="fas fa-check-circle me-2"),
                    "钢衬里模式预测数据已成功保存到数据库！"
                ], color="success", duration=4000)
            else:
                return dbc.Alert([
                    html.I(className="fas fa-exclamation-triangle me-2"),
                    "保存失败：数据库操作异常，请检查数据库连接状态。"
                ], color="danger", duration=6000)
                
        except Exception as e:
            logger.error(f"钢衬里模式数据保存异常: {e}", exc_info=True)
            return dbc.Alert([
                html.I(className="fas fa-exclamation-triangle me-2"),
                f"保存过程中发生错误：{str(e)}"
            ], color="danger", duration=6000)

    @app.callback(
        Output("custom-mode-save-feedback", "children"),
        Input("save-custom-mode-data-btn", "n_clicks"),
        State("custom-mode-report-data", "data"),
        prevent_initial_call=True
    )
    def save_custom_mode_data_to_database(n_clicks, report_data):
        """保存自定义模式数据到数据库"""
        if not n_clicks or not report_data:
            raise PreventUpdate
        
        try:
            # 调用已有的保存函数
            save_success = save_report_to_database(report_data, "custom_mode")
            
            if save_success:
                return dbc.Alert([
                    html.I(className="fas fa-check-circle me-2"),
                    "自定义模式预测数据已成功保存到数据库！"
                ], color="success", duration=4000)
            else:
                return dbc.Alert([
                    html.I(className="fas fa-exclamation-triangle me-2"),
                    "保存失败：数据库操作异常，请检查数据库连接状态。"
                ], color="danger", duration=6000)
                
        except Exception as e:
            logger.error(f"自定义模式数据保存异常: {e}", exc_info=True)
            return dbc.Alert([
                html.I(className="fas fa-exclamation-triangle me-2"),
                f"保存过程中发生错误：{str(e)}"
            ], color="danger", duration=6000)
    


    # 钢筋笼模式字段状态更新回调
    @app.callback(
        [Output(f"{field_id}", "disabled") for field_id in [
            "tower-crane-category-param", "tower-crane-1500", "heavy-tower-crane",
            "steel-production-category-param", 
            "lifting-equipment-category-param", "steel-wire-a36-1500", "shackle-55t", "basket-bolt-3128",
            "sleeve-category-param", "straight-threaded-sleeve", "cone-steel-sleeve", "module-vertical-connection",
            "steel-tonnage-category-param", "steel-price-category-param"
        ]] + 
        [Output("steel-cage-field-status-store", "data")],
        [Input("steel-reinforcement-parameter-modal", "is_open")],
        prevent_initial_call=True
    )
    def update_steel_cage_field_status(is_open):
        """更新钢筋笼模式字段的启用/禁用状态"""
        if not is_open:
            # 模态窗口关闭时，返回默认状态
            field_count = 14  # 上面字段的数量
            return [False] * field_count + [{}]
        
        # 获取字段状态
        field_status = get_field_status_for_mode('steel_cage')
        
        # 准备字段禁用状态列表
        field_ids = [
            "tower-crane-category-param", "tower-crane-1500", "heavy-tower-crane",
            "steel-production-category-param",
            "lifting-equipment-category-param", "steel-wire-a36-1500", "shackle-55t", "basket-bolt-3128", 
            "sleeve-category-param", "straight-threaded-sleeve", "cone-steel-sleeve", "module-vertical-connection",
            "steel-tonnage-category-param", "steel-price-category-param"
        ]
        
        disabled_states = []
        for field_id in field_ids:
            field_info = field_status.get(field_id, {'status': 'enabled'})
            disabled_states.append(field_info['status'] == 'disabled')
        
        logger.info(f"钢筋笼模式字段状态更新: {sum(disabled_states)} 个字段被禁用")
        
        return disabled_states + [field_status]

    # 钢衬里模式字段状态更新回调
    @app.callback(
        [Output(f"{field_id}", "disabled") for field_id in [
            "assembly-site-category-param", "fixture-making-category-param",
            "steel-support-embedded-category-param", "steel-support-concrete-chiseling", 
            "steel-support-concrete-backfill", "steel-support-installation", "steel-support-depreciation",
            "buttress-category-param", "buttress-installation", "buttress-removal", "buttress-component-depreciation",
            "walkway-platform-category-param", "walkway-platform-manufacturing", "walkway-platform-erection", "walkway-platform-removal",
            "steel-grid-beam-category-param", "steel-grid-manufacturing", "steel-grid-installation", "steel-grid-removal",
            "steel-lining-measures-category-param"
        ]] +
        [Output("steel-lining-field-status-store", "data")],
        [Input("steel-lining-parameter-modal2", "is_open")],
        prevent_initial_call=True
    )
    def update_steel_lining_field_status(is_open):
        """更新钢衬里模式字段的启用/禁用状态"""
        if not is_open:
            # 模态窗口关闭时，返回默认状态
            field_count = 20  # 上面字段的数量
            return [False] * field_count + [{}]
        
        # 获取字段状态
        field_status = get_field_status_for_mode('steel_lining')
        
        # 准备字段禁用状态列表
        field_ids = [
            "assembly-site-category-param", "fixture-making-category-param",
            "steel-support-embedded-category-param", "steel-support-concrete-chiseling",
            "steel-support-concrete-backfill", "steel-support-installation", "steel-support-depreciation",
            "buttress-category-param", "buttress-installation", "buttress-removal", "buttress-component-depreciation", 
            "walkway-platform-category-param", "walkway-platform-manufacturing", "walkway-platform-erection", "walkway-platform-removal",
            "steel-grid-beam-category-param", "steel-grid-manufacturing", "steel-grid-installation", "steel-grid-removal",
            "steel-lining-measures-category-param"
        ]
        
        disabled_states = []
        for field_id in field_ids:
            field_info = field_status.get(field_id, {'status': 'enabled'})
            disabled_states.append(field_info['status'] == 'disabled')
        
        logger.info(f"钢衬里模式字段状态更新: {sum(disabled_states)} 个字段被禁用")
        
        return disabled_states + [field_status]




    def get_indicator_status_from_db():
        """
        从basic_indicators表查询所有指标的状态 - 增强版本
        
        Returns:
            dict: {指标名称: 状态} 的字典，状态为 'enabled' 或 'disabled'
        """
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            
            # 查询所有指标的名称和状态
            query = """
            SELECT name, status FROM basic_indicators 
            WHERE status IN ('enabled', 'disabled')
            ORDER BY name
            """
            
            cursor.execute(query)
            results = cursor.fetchall()
            
            # 构建状态字典
            status_dict = {}
            for row in results:
                indicator_name = row['name']
                status = row['status']
                status_dict[indicator_name] = status
                
            logger.info(f"成功查询到 {len(status_dict)} 个指标的状态信息")
            
            # 添加详细的调试信息
            for indicator_name, status in status_dict.items():
                logger.info(f"指标 '{indicator_name}' 状态: {status}")
                
            return status_dict
            
        except mysql.connector.Error as e:
            logger.error(f"查询指标状态时MySQL错误: {e}")
            return {}
        except Exception as e:
            logger.error(f"查询指标状态时发生错误: {e}")
            return {}
        finally:
            if conn:
                conn.close()

    def get_field_status_for_mode(mode):
        """
        获取指定模式下所有字段的状态信息 - 增强版本
        
        Args:
            mode (str): 模式 ('steel_cage' 或 'steel_lining')
        
        Returns:
            dict: {字段ID: {'status': 'enabled'/'disabled', 'indicator_name': '指标名称', 'message': '提示信息'}}
        """
        logger.info(f"开始获取模式 {mode} 的字段状态")
        
        # 获取数据库中的指标状态
        indicator_status = get_indicator_status_from_db()
        
        # 获取该模式的映射关系
        mode_mapping = INDICATOR_FIELD_MAPPING.get(mode, {})
        
        logger.info(f"模式 {mode} 的映射关系包含 {len(mode_mapping)} 个指标")
        
        field_status = {}
        
        for indicator_name, config in mode_mapping.items():
            # 获取该指标在数据库中的状态
            db_status = indicator_status.get(indicator_name, 'enabled')  # 默认启用
            
            logger.info(f"指标 '{indicator_name}' 在数据库中的状态: {db_status}")
            
            # 为该指标下的所有字段设置状态
            fields = config.get('fields', [])
            section_title = config.get('section_title', indicator_name)
            
            for field_id in fields:
                if db_status == 'disabled':
                    field_status[field_id] = {
                        'status': 'disabled',
                        'indicator_name': indicator_name,
                        'section_title': section_title,
                        'message': f'已禁用 - 若启用请到数据管理模块打开「{indicator_name}」指标'
                    }
                    logger.info(f"字段 {field_id} 被设置为禁用状态（对应指标: {indicator_name}）")
                else:
                    field_status[field_id] = {
                        'status': 'enabled',
                        'indicator_name': indicator_name,
                        'section_title': section_title,
                        'message': ''
                    }
                    logger.info(f"字段 {field_id} 被设置为启用状态（对应指标: {indicator_name}）")
        
        logger.info(f"为模式 {mode} 生成了 {len(field_status)} 个字段的状态信息")
        
        # 添加详细的字段状态日志
        for field_id, status_info in field_status.items():
            logger.info(f"字段状态详情 - {field_id}: {status_info}")
        # ===== 新增：添加算法状态影响的额外检查 =====
        try:
            # 检查算法执行能力对字段状态的影响
            from .data import CostPredictionSystem
            temp_system = CostPredictionSystem(mode=mode)
            temp_system.load_algorithm_configs()
            
            algo_capability = temp_system.check_algorithm_execution_capability()
            
            # 如果所有算法都禁用，给相关字段添加额外警告信息
            if not algo_capability.get('can_execute_ai', True):
                for field_id, status_info in field_status.items():
                    if status_info['status'] == 'enabled':
                        # 为启用的字段添加算法状态警告
                        status_info['algorithm_warning'] = {
                            'has_warning': True,
                            'message': f"注意：{algo_capability.get('message', 'AI算法不可用')}",
                            'suggestion': '请检查算法配置以确保预测功能正常'
                        }
                    
        except Exception as e:
            logger.warning(f"检查算法状态对字段的影响时出错: {e}")
            # 不影响主要功能，继续执行
                   
        return field_status
def collect_user_quantities(tower_crane_qty_category, steel_production_qty_category,
                           lifting_equipment_qty_category, sleeve_qty_category,
                           steel_tonnage_category, measures_qty_category,
                           tower_crane_1500, heavy_tower_crane, steel_wire_a36,
                           shackle_55t, basket_bolt, straight_threaded_sleeve,
                           cone_steel_sleeve, module_vertical_connection, safe_float):
    """
    收集和整理用户输入的工程量数据
    逻辑：优先使用分类标题输入，如果没有则汇总细分参数
    """
    quantities = {}
    
    # 1. 塔吊租赁工程量
    if tower_crane_qty_category is not None and str(tower_crane_qty_category).strip() != "":
        quantities['塔吊租赁工程量'] = safe_float(tower_crane_qty_category)
    else:
        sub_total = safe_float(tower_crane_1500) + safe_float(heavy_tower_crane)
        quantities['塔吊租赁工程量'] = sub_total
    
    # 2. 钢筋总吨数
    if steel_tonnage_category is not None and str(steel_tonnage_category).strip() != "":
        quantities['钢筋总吨数'] = safe_float(steel_tonnage_category)
    else:
        if steel_production_qty_category is not None and str(steel_production_qty_category).strip() != "":
            quantities['钢筋总吨数'] = safe_float(steel_production_qty_category)
        else:
            quantities['钢筋总吨数'] = 0.0
    
    # 3. 吊索具数量
    if lifting_equipment_qty_category is not None and str(lifting_equipment_qty_category).strip() != "":
        quantities['吊索具数量'] = safe_float(lifting_equipment_qty_category)
    else:
        sub_total = (safe_float(steel_wire_a36) + safe_float(shackle_55t) + safe_float(basket_bolt))
        quantities['吊索具数量'] = sub_total
    
    # 4. 套筒数量
    if sleeve_qty_category is not None and str(sleeve_qty_category).strip() != "":
        quantities['套筒数量'] = safe_float(sleeve_qty_category)
    else:
        sub_total = (safe_float(straight_threaded_sleeve) + safe_float(cone_steel_sleeve) + 
                    safe_float(module_vertical_connection))
        quantities['套筒数量'] = sub_total
    
    # 5. 措施费相关工程量（可选）
    if measures_qty_category is not None and str(measures_qty_category).strip() != "":
        quantities['措施费工程量'] = safe_float(measures_qty_category)
    else:
        quantities['措施费工程量'] = 0.0
    
    return quantities

def prepare_ml_inputs(quantities):
    """
    准备机器学习预测系统的输入参数
    将收集到的工程量数据转换为预测系统所需的格式
    """
    user_inputs = {}

    if quantities.get('钢筋总吨数', 0) > 0:
        user_inputs['钢筋总吨数'] = quantities['钢筋总吨数']
    
    if quantities.get('塔吊租赁工程量', 0) > 0:
        user_inputs['塔吊租赁工程量'] = quantities['塔吊租赁工程量']
    
    if quantities.get('吊索具数量', 0) > 0:
        user_inputs['吊索具数量'] = quantities['吊索具数量']
    
    if quantities.get('套筒数量', 0) > 0:
        user_inputs['套筒数量'] = quantities['套筒数量']
    
    user_inputs = {k: v for k, v in user_inputs.items() if v > 0}
    
    return user_inputs


def create_prediction_results_table(quantities, ml_prediction_results, measures_cost_value):
    """创建钢筋笼预测结果表格显示 - 显示具体算法名称版本"""
    
    title_section = html.Div([
        html.H3("钢筋笼施工智能预测分析报告", className="text-center text-primary mb-4"),
        html.Hr()
    ])
    
    input_summary_rows = [
        html.Tr([html.Td("钢筋总吨数"), html.Td(f"{quantities.get('钢筋总吨数', 0):.2f}"), html.Td("吨")]),
        html.Tr([html.Td("塔吊租赁工程量"), html.Td(f"{quantities.get('塔吊租赁工程量', 0):.2f}"), html.Td("台班")]),
        html.Tr([html.Td("吊索具数量"), html.Td(f"{quantities.get('吊索具数量', 0):.0f}"), html.Td("套")]),
        html.Tr([html.Td("套筒数量"), html.Td(f"{quantities.get('套筒数量', 0):.0f}"), html.Td("个")])
    ]
    if measures_cost_value > 0:
        input_summary_rows.append(
             html.Tr([html.Td("措施费工程量", style={'fontWeight': 'bold'}), html.Td(f"{measures_cost_value:.2f}", style={'fontWeight': 'bold'}), html.Td("元", style={'fontWeight': 'bold'})])
        )

    input_summary = html.Div([
        html.H5("📋 项目工程量输入汇总", className="text-info mb-3"),
        dbc.Row([
            dbc.Col(dbc.Card([
                dbc.CardHeader("📊 数据概览"),
                dbc.CardBody([
                    html.P(f"工程量类型: {len(input_summary_rows)}项", className="mb-1"),
                    html.P(f"有效输入: {sum(1 for k, v in quantities.items() if v > 0 and k != '措施费工程量') + (1 if measures_cost_value > 0 else 0)}项", className="mb-1"),
                    html.P(f"预测状态: {'✅ 可执行' if any(v > 0 for k,v in quantities.items() if k != '措施费工程量') else '❌ 数据不足'}", className="mb-0")
                ])
            ]), width=4)
        ], className="mb-4")
    ])

    # 预测方法状态总览
    status_summary_section = create_prediction_status_summary(ml_prediction_results, mode='steel_cage')

    # 【新增】最佳算法参数信息部分
    best_algorithm_section = create_best_algorithm_params_section(ml_prediction_results)
    
    prediction_section = html.Div()

    if ml_prediction_results and "error" not in ml_prediction_results:
        # 获取状态信息
        method_status = ml_prediction_results.get("预测方法状态", {})
        
        # 获取四种独立的预测结果
        ai_raw = ml_prediction_results.get("AI预测-原始值")
        ai_final = ml_prediction_results.get("AI预测-最终值") 
        ratio_raw = ml_prediction_results.get("比率法-原始值")
        ratio_final = ml_prediction_results.get("比率法-最终值")
        
        matched_cluster = ml_prediction_results.get("匹配到的规则来源", "未知")
        estimated_quantities_from_ml = ml_prediction_results.get("估算的工程量", {})
        estimated_costs_from_ml = ml_prediction_results.get("估算的各项成本 (用于ML的特征)", {})

        prediction_table_rows = []

        # 匹配规则来源
        prediction_table_rows.append(
            html.Tr([
                html.Td("🎯 匹配规则来源", className="fw-bold"),
                html.Td(matched_cluster, className="text-info"),
                html.Td("-"),html.Td("-"),html.Td("-"),
                html.Td("基于历史项目聚类分析匹配最相似的项目类型")
            ])
        )

        # 智能工程量补全
        if estimated_quantities_from_ml:
            prediction_table_rows.append(
                html.Tr([html.Td("📊 智能工程量补全 (机器学习模型估算)", className="fw-bold", colSpan=6)], className="table-info")
            )
            for param, qty in estimated_quantities_from_ml.items():
                unit = get_quantity_unit(param)
                prediction_table_rows.append(
                    html.Tr([html.Td(f"  └ {param}"), html.Td(f"{qty:.2f}"), html.Td(unit), html.Td("-"),html.Td("-"), html.Td("基于历史数据比例关系智能推算")])
                )

        # 各项成本预测
        if estimated_costs_from_ml:
            prediction_table_rows.append(
                html.Tr([html.Td("💰 各项成本预测 (机器学习模型估算)", className="fw-bold", colSpan=6)], className="table-warning")
            )
            cost_items = ['塔吊租赁费', '钢筋生产线费用', '吊索具费用', '套筒费用']
            for item in cost_items:
                if item in estimated_costs_from_ml:
                    prediction_table_rows.append(
                        html.Tr([html.Td(f"  └ {item}"), html.Td(f"{estimated_costs_from_ml[item]:,.2f}"), html.Td("元"), html.Td("-"),html.Td("-"), html.Td("基于聚类规则和历史单价计算")])
                    )

        # 关键新增：解析最佳算法信息
        best_algorithm_info = get_best_algorithm_info(ml_prediction_results)
        
        # 四种独立的预测方法显示 - 显示具体算法名称
        prediction_methods = [
            {
                'key': 'ml_prediction_raw',
                'title': f'🤖 {best_algorithm_info["raw"]["display_name"]}',
                'data': ai_raw,
                'description': f'{best_algorithm_info["raw"]["description"]} (不含措施费)',
                'row_class': 'table-primary',
                'algorithm_details': best_algorithm_info["raw"]
            },
            {
                'key': 'ml_prediction_final', 
                'title': f'🤖 {best_algorithm_info["final"]["display_name"]}',
                'data': ai_final,
                'description': f'{best_algorithm_info["final"]["description"]} (含措施费 ¥{measures_cost_value:,.2f})',
                'row_class': 'table-warning',
                'algorithm_details': best_algorithm_info["final"]
            },
            {
                'key': 'ratio_method_raw',
                'title': '📈 比率法-原始值', 
                'data': ratio_raw,
                'description': '比率法预测，不含措施费',
                'row_class': 'table-info'
            },
            {
                'key': 'ratio_method_final',
                'title': '📈 比率法-最终值',
                'data': ratio_final, 
                'description': f'比率法预测，已包含措施费 ¥{measures_cost_value:,.2f}',
                'row_class': 'table-success'
            }
        ]

        # 处理每个方法的状态和显示
        for method in prediction_methods:
            method_data = method['data']
            method_title = method['title']
            method_description = method['description']
            method_row_class = method['row_class']
            method_key = method['key']
            
            # 获取该方法的详细状态信息
            method_status_info = method_status.get(method_key, {})
            final_status = method_status_info.get('final_status', 'unknown')
            
            # 根据最终状态决定如何显示
            if final_status == 'fully_available':
                # 完全可用 - 显示预测结果
                if method_key.startswith('ml_') and isinstance(method_data, dict) and method_data.get('集成平均预测'):
                    # AI预测方法：显示集成结果
                    algorithm_details = method.get('algorithm_details', {})
                    confidence_info = algorithm_details.get('confidence', '高可信度')
                    
                    prediction_table_rows.append(
                        html.Tr([
                            html.Td([
                                html.Div(method_title, className="fw-bold"),
                                html.Small([
                                html.Span("🏆 最佳预测算法 | ", className="text-success fw-bold", style={'fontSize': '11px'}),
                                html.Span(f"算法可信度: {confidence_info}", className="text-muted", style={'fontSize': '11px'})
                            ], className="d-block")
                            ], style={'color': 'black'}),
                            html.Td(f"{method_data['集成平均预测']:,.0f}", className="fw-bold", style={'color': 'black'}),
                            html.Td("元", className="fw-bold", style={'color': 'black'}),
                            html.Td("-", className="fw-bold", style={'color': 'black'}),
                            html.Td("-", className="fw-bold", style={'color': 'black'}),
                            html.Td([
                                html.Div(method_description, className="mb-1"),
                                html.Small(f"基于算法: {algorithm_details.get('algorithm_name', '未知算法')}", className="text-info", style={'fontSize': '11px'})
                            ], style={'color': 'black'})
                        ], className=method_row_class)
                    )
                elif method_key.startswith('ratio_') and isinstance(method_data, (int, float)):
                    # 比率法：直接显示数值
                    prediction_table_rows.append(
                        html.Tr([
                            html.Td(method_title, className="fw-bold", style={'color': 'black'}),
                            html.Td(f"{method_data:,.0f}", className="fw-bold", style={'color': 'black'}),
                            html.Td("元", className="fw-bold", style={'color': 'black'}),
                            html.Td("-", className="fw-bold", style={'color': 'black'}),
                            html.Td("-", className="fw-bold", style={'color': 'black'}),
                            html.Td(method_description, className="fw-bold", style={'color': 'black'})
                        ], className=method_row_class)
                    )
                else:
                    # 数据格式异常
                    prediction_table_rows.append(
                        html.Tr([
                            html.Td(f"⚠️ {method_title} (数据异常)", style={'color': '#ffc107', 'fontWeight': 'bold'}),
                            html.Td("数据格式错误", style={'color': '#ffc107'}),
                            html.Td("元", style={'color': '#6c757d'}),
                            html.Td("-", style={'color': '#6c757d'}),
                            html.Td("-", style={'color': '#6c757d'}),
                            html.Td(f"预期数据格式异常: {type(method_data)}", style={'color': '#6c757d'})
                        ], style={'backgroundColor': '#fffbf0'})
                    )
                    
            elif final_status == 'execute_only':
                # 可以执行但显示被禁用
                prediction_table_rows.append(
                    html.Tr([
                        html.Td(f"🚫 {method_title} (显示已禁用)", style={'color': '#ffc107', 'fontWeight': 'bold'}),
                        html.Td("🚫 显示被禁用", style={'color': '#ffc107', 'fontWeight': 'bold'}),
                        html.Td("元", style={'color': '#6c757d'}),
                        html.Td("-", style={'color': '#6c757d'}),
                        html.Td("-", style={'color': '#6c757d'}),
                        html.Td(method_status_info.get('display_message', '显示权限已禁用'), style={'color': '#6c757d'})
                    ], style={'backgroundColor': '#fffbf0'})
                )
                
            elif final_status == 'display_error':
                # 想显示但无法执行
                prediction_table_rows.append(
                    html.Tr([
                        html.Td(f"❌ {method_title} (执行失败)", style={'color': '#dc3545', 'fontWeight': 'bold'}),
                        html.Td("执行失败", style={'color': '#dc3545', 'fontWeight': 'bold'}),
                        html.Td("元", style={'color': '#6c757d'}),
                        html.Td("-", style={'color': '#6c757d'}),
                        html.Td("-", style={'color': '#6c757d'}),
                        html.Td(method_status_info.get('execution_message', '执行条件不满足'), style={'color': '#6c757d'})
                    ], style={'backgroundColor': '#fff5f5'})
                )
                
            else:  # fully_disabled 或其他状态
                # 完全禁用或未知状态
                prediction_table_rows.append(
                    html.Tr([
                        html.Td(f"🚫 {method_title} (完全禁用)", style={'color': '#6c757d', 'fontWeight': 'bold'}),
                        html.Td("🚫 完全禁用", style={'color': '#6c757d', 'fontWeight': 'bold'}),
                        html.Td("元", style={'color': '#6c757d'}),
                        html.Td("-", style={'color': '#6c757d'}),
                        html.Td("-", style={'color': '#6c757d'}),
                        html.Td(f"状态: {final_status}", style={'color': '#6c757d'})
                    ], style={'backgroundColor': '#f8f9fa', 'opacity': '0.7'})
                )
        
        prediction_section = html.Div([
            html.H5("智能预测分析结果", className="text-primary mb-3"),
            dbc.Table([
                html.Thead(html.Tr([
                    html.Th("预测项目", style={"width": "25%"}),
                    html.Th("预测总价(元)", style={"width": "15%"}), 
                    html.Th("单位", style={"width": "5%"}),
                    html.Th("与比例法偏差(元)", style={"width": "15%"}), 
                    html.Th("与比例法偏差(%)", style={"width": "15%"}),
                    html.Th("算法说明", style={"width": "25%"})
                ])),
                html.Tbody(prediction_table_rows)
            ], bordered=True, hover=True, striped=True, size="sm", className="mb-4"),
            create_independent_prediction_summary(method_status, ai_raw, ai_final, ratio_raw, ratio_final, measures_cost_value)
        ])
        
    elif ml_prediction_results and "error" in ml_prediction_results:
        prediction_section = html.Div([
            html.H5("智能预测结果", className="text-warning mb-3"),
            dbc.Alert([
                html.I(className="fas fa-exclamation-triangle me-2"),
                html.Strong("⚠️ 预测失败: "),
                html.Span(ml_prediction_results['error'])
            ], color="warning", className="mb-4")
        ])
    
    instruction_section = html.Div([
        html.Hr(),
        html.H6("📖 技术说明", className="text-muted mb-3"),
        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardHeader("🔬 算法原理"), dbc.CardBody(html.Ul([html.Li("K-Means聚类：自动识别项目类型"), html.Li("多算法集成：提高预测准确性"), html.Li("交叉验证：四种独立方法相互验证")], className="small mb-0"))]), width=4),
            dbc.Col(dbc.Card([dbc.CardHeader("📊 数据来源"), dbc.CardBody(html.Ul([html.Li("历史项目数据库：真实施工记录"), html.Li("成本规律提取：自动学习价格模式"), html.Li("持续优化：新数据不断完善模型")], className="small mb-0"))]), width=4),
            dbc.Col(dbc.Card([dbc.CardHeader("💡 使用建议"), dbc.CardBody(html.Ul([html.Li("独立方法：每种方法可单独参考"), html.Li("灵活选择：根据项目需要选择合适方法"), html.Li("综合判断：多方法结果综合考虑")], className="small mb-0"))]), width=4)
        ])
    ])
    
    return html.Div([title_section, input_summary, status_summary_section, best_algorithm_section, prediction_section, instruction_section], className="container-fluid")


def get_best_algorithm_info(ml_prediction_results):
    """
    解析预测结果，获取最佳算法信息
    
    Args:
        ml_prediction_results: 预测结果字典
        
    Returns:
        dict: 包含原始值和最终值的最佳算法信息
    """
    # 算法显示名称映射
    algorithm_display_names = {
        "岭回归 (RidgeCV)": "岭回归算法预测",
        "决策树 (Decision Tree)": "决策树算法预测", 
        "随机森林 (Random Forest)": "随机森林算法预测",
        "支持向量回归 (SVR)": "支持向量机算法预测",
        "神经网络 (MLPRegressor)": "神经网络算法预测"
    }
    
    # 获取比率法预测值用于计算偏差
    ratio_raw = ml_prediction_results.get("比率法-原始值", 0)
    
    # 分析原始值的最佳算法
    ai_raw_data = ml_prediction_results.get("AI预测-原始值", {})
    raw_best_algo = find_best_single_algorithm(ai_raw_data, ratio_raw)
    
    # 分析最终值的最佳算法
    ai_final_data = ml_prediction_results.get("AI预测-最终值", {})
    final_best_algo = find_best_single_algorithm(ai_final_data, ratio_raw)
    
    return {
        "raw": {
            "display_name": algorithm_display_names.get(raw_best_algo["algorithm"], "集成算法预测"),
            "algorithm_name": raw_best_algo["algorithm"],
            "confidence": raw_best_algo["confidence"],
            "description": f"基于{raw_best_algo['algorithm']}的预测结果"
        },
        "final": {
            "display_name": algorithm_display_names.get(final_best_algo["algorithm"], "集成算法预测"),
            "algorithm_name": final_best_algo["algorithm"], 
            "confidence": final_best_algo["confidence"],
            "description": f"基于{final_best_algo['algorithm']}的预测结果"
        }
    }


def find_best_single_algorithm(ai_prediction_data, ratio_value):
    """
    从AI预测数据中找出表现最佳的单一算法
    
    Args:
        ai_prediction_data: AI预测数据字典
        ratio_value: 比率法预测值
        
    Returns:
        dict: 最佳算法信息
    """
    if not isinstance(ai_prediction_data, dict) or not ratio_value:
        return {
            "algorithm": "集成平均",
            "confidence": "中等",
            "deviation": 0
        }
    
    best_algorithm = "集成平均"
    min_deviation = float('inf')
    best_confidence = "中等"
    
    # 遍历所有算法预测结果
    for algo_name, prediction_value in ai_prediction_data.items():
        if (algo_name != '集成平均预测' and 
            isinstance(prediction_value, (int, float)) and 
            prediction_value > 0):
            
            # 计算与比率法的偏差
            deviation = abs(float(prediction_value) - float(ratio_value)) if isinstance(prediction_value, (int, float)) and isinstance(ratio_value, (int, float)) else float('inf')
            
            if deviation < min_deviation:
                min_deviation = deviation
                best_algorithm = algo_name
                
                # 根据偏差程度确定可信度
                deviation_pct = (deviation / ratio_value * 100) if ratio_value > 0 else 100
                if deviation_pct < 5:
                    best_confidence = "极高可信度"
                elif deviation_pct < 10:
                    best_confidence = "高可信度"
                elif deviation_pct < 20:
                    best_confidence = "中等可信度"
                else:
                    best_confidence = "低可信度"
    
    return {
        "algorithm": best_algorithm,
        "confidence": best_confidence,
        "deviation": min_deviation
    }

def create_independent_prediction_summary(method_status, ai_raw, ai_final, ratio_raw, ratio_final, measures_cost_value, mode='steel_cage'):
    """创建独立预测方法的汇总信息 - 只保留预测结果汇总框"""
    
    enabled_results = []
    
    methods = [
        ('ml_prediction_raw', 'AI预测-原始值', ai_raw),
        ('ml_prediction_final', 'AI预测-最终值', ai_final), 
        ('ratio_method_raw', '比率法-原始值', ratio_raw),
        ('ratio_method_final', '比率法-最终值', ratio_final)
    ]
    
    mode_name = "钢筋笼" if mode == 'steel_cage' else "钢衬里"
    
    # 收集有效的预测结果
    for method_key, method_name, method_data in methods:
        method_info = method_status.get(method_key, {})
        
        # 检查是否使用新的状态结构
        if isinstance(method_info, dict) and 'final_status' in method_info:
            final_status = method_info.get('final_status', 'unknown')
            
            if final_status in ['fully_available', 'execute_only']:
                result_value = extract_prediction_value(method_data, method_key)
                if result_value is not None:
                    enabled_results.append({
                        'name': method_name,
                        'value': result_value,
                        'status': final_status
                    })
        else:
            # 兼容旧的状态结构
            if method_info.get('enabled', False):
                result_value = extract_prediction_value(method_data, method_key)
                if result_value is not None:
                    enabled_results.append({
                        'name': method_name,
                        'value': result_value,
                        'status': 'legacy_enabled'
                    })

    # 只创建预测结果汇总卡片
    if len(enabled_results) > 0:
        # 过滤出有效的数值结果
        valid_results = [r for r in enabled_results if isinstance(r.get('value'), (int, float))]
        
        if len(valid_results) > 0:
            min_value = min(r['value'] for r in valid_results)
            max_value = max(r['value'] for r in valid_results)
            avg_value = sum(r['value'] for r in valid_results) / len(valid_results)
            
            summary_card = dbc.Col(dbc.Card([
                dbc.CardHeader(f"📈 {mode_name}预测结果汇总"),
                dbc.CardBody([
                    html.P(f"预测范围: {min_value:,.0f} - {max_value:,.0f} 元", className="mb-1"),
                    html.P(f"平均预测: {avg_value:,.0f} 元", className="mb-1"),
                    html.P(f"建议预算(含10%缓冲): {avg_value * 1.1:,.0f} 元", className="mb-1", style={'fontWeight': 'bold'}),
                    html.Small(f"基于 {len(valid_results)} 种有效预测结果", className="text-muted")
                ])
            ], color="info", outline=True), width=6)
            
            return html.Div([
                html.Hr(),
                dbc.Row([summary_card], className="mb-3")
            ])
    
    # 如果没有有效结果，显示提示信息
    summary_card = dbc.Col(dbc.Card([
        dbc.CardHeader(f"📈 {mode_name}预测结果汇总"),
        dbc.CardBody([
            html.P("暂无有效的预测结果", className="mb-2"),
            html.Small("请启用至少一种预测方法并确保算法正常运行", className="text-muted")
        ])
    ], color="secondary", outline=True), width=6)
    
    return html.Div([
        html.Hr(),
        dbc.Row([summary_card], className="mb-3")
    ])


def get_algorithm_configs_from_database(mode):
    """
    从数据库获取算法配置信息 - 调试版本
    
    Args:
        mode (str): 模式名称 ('steel_cage' 或 'steel_lining')
        
    Returns:
        dict: 包含算法配置信息的字典
    """
    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        
        # 先检查表是否存在
        cursor.execute("SHOW TABLES LIKE 'algorithm_configs'")
        table_exists = cursor.fetchone()
        
        if not table_exists:
            logger.warning("algorithm_configs 表不存在")
            # 如果表不存在，返回默认值
            return create_default_algorithm_info()
        
        # 检查表结构
        cursor.execute("DESCRIBE algorithm_configs")
        columns = cursor.fetchall()
        logger.info(f"algorithm_configs 表结构: {[col['Field'] for col in columns]}")
        
        # 查看所有数据（调试用）
        cursor.execute("SELECT * FROM algorithm_configs LIMIT 10")
        all_data = cursor.fetchall()
        logger.info(f"algorithm_configs 表数据示例: {all_data}")
        
        # 根据模式确定查询条件 - 改进查询逻辑
        if mode == 'steel_cage':
            mode_conditions = ["钢筋笼", "steel_cage", "both", "all"]
        elif mode == 'steel_lining':
            mode_conditions = ["钢衬里", "steel_lining", "both", "all"]
        else:
            mode_conditions = [mode, "both", "all"]
        
        # 构建查询条件
        placeholders = ", ".join(["%s"] * len(mode_conditions))
        query = f"""
        SELECT id, algorithm_name, status
        FROM algorithm_configs 
        WHERE mode_type IN ({placeholders}) OR mode_type IS NULL
        ORDER BY id
        """
        
        cursor.execute(query, mode_conditions)
        results = cursor.fetchall()
        
        logger.info(f"查询到的算法配置 ({mode}): {results}")
        
        # 如果没有匹配的数据，尝试查询所有数据
        if not results:
            logger.warning(f"没有找到 {mode} 模式的算法配置，尝试查询所有算法")
            cursor.execute("SELECT id, algorithm_name, status FROM algorithm_configs")
            results = cursor.fetchall()
            logger.info(f"所有算法配置: {results}")
        
        # 统计算法状态
        total_count = len(results)
        enabled_count = 0
        enabled_algorithms = []
        disabled_algorithms = []
        
        for row in results:
            algorithm_name = row['algorithm_name']
            status = row['status']
            
            if status == 'enabled':
                enabled_count += 1
                enabled_algorithms.append(algorithm_name)
            else:
                disabled_algorithms.append(algorithm_name)
        
        logger.info(f"算法统计结果 - 总数: {total_count}, 启用: {enabled_count}, 启用列表: {enabled_algorithms}")
        
        return {
            'total_count': total_count,
            'enabled_count': enabled_count,
            'disabled_count': total_count - enabled_count,
            'enabled_algorithms': enabled_algorithms,
            'disabled_algorithms': disabled_algorithms,
            'can_execute_ai': enabled_count > 0,
            'algorithm_details': results
        }
        
    except mysql.connector.Error as e:
        logger.error(f"MySQL查询算法配置失败: {e}")
        return create_default_algorithm_info()
    except Exception as e:
        logger.error(f"获取算法配置时发生错误: {e}", exc_info=True)
        return create_default_algorithm_info()
    finally:
        if conn:
            conn.close()


def create_default_algorithm_info():
    """创建默认的算法信息（当数据库查询失败时使用）"""
    # 提供一个默认的算法配置
    default_algorithms = ["岭回归", "决策树", "随机森林", "支持向量机", "神经网络"]
    
    return {
        'total_count': 5,
        'enabled_count': 1,  # 假设只有1个启用
        'disabled_count': 4,
        'enabled_algorithms': ["岭回归"],  # 假设岭回归启用
        'disabled_algorithms': ["决策树", "随机森林", "支持向量机", "神经网络"],
        'can_execute_ai': True,
        'algorithm_details': []
    }




def extract_prediction_value(method_data, method_key):
    """
    从预测数据中提取数值结果的辅助函数
    
    Args:
        method_data: 预测方法的数据
        method_key (str): 预测方法的键名
        
    Returns:
        float or None: 提取的数值结果
    """
    if method_data is None:
        return None
        
    if isinstance(method_data, (int, float)):
        return method_data
        
    if isinstance(method_data, dict):
        # 检查是否是错误或状态信息
        if 'error' in method_data or 'status' in method_data:
            return None
            
        # 对于AI预测，尝试提取集成平均预测
        if method_key.startswith('ml_') and '集成平均预测' in method_data:
            ensemble_pred = method_data['集成平均预测']
            if isinstance(ensemble_pred, (int, float)):
                return ensemble_pred
                
    return None

def create_prediction_status_summary(ml_prediction_results, mode='steel_cage'):
    """创建预测方法状态总览 - 基于新的状态结构"""
    
    # 从预测结果中提取状态信息（新的数据结构）
    algorithm_status = ml_prediction_results.get('算法执行状态', {}) if ml_prediction_results else {}
    prediction_method_status = ml_prediction_results.get('预测方法状态', {}) if ml_prediction_results else {}
    display_permission_status = ml_prediction_results.get('显示权限状态', {}) if ml_prediction_results else {}
    
    # 如果没有获取到状态信息，尝试创建临时系统实例
    if not algorithm_status or not prediction_method_status:
        try:
            temp_system = CostPredictionSystem(mode=mode)
            temp_system.load_algorithm_configs()
            temp_system.load_comprehensive_indicators_status()
            
            algorithm_status = temp_system.check_algorithm_execution_capability()
            
            # 获取四种预测方法的状态
            methods = ['ml_prediction_raw', 'ml_prediction_final', 'ratio_method_raw', 'ratio_method_final']
            for method_key in methods:
                combined_status = temp_system.get_combined_prediction_status(method_key)
                prediction_method_status[method_key] = combined_status
                
        except Exception as e:
            logger.error(f"创建临时系统获取状态失败: {e}")
            # 使用默认状态
            algorithm_status = {'can_execute_ai': True, 'message': '状态未知', 'enabled_count': 0}
            prediction_method_status = {}
    
    # 构建状态显示项
    status_items = []
    mode_name = "钢筋笼" if mode == 'steel_cage' else "钢衬里"
    
    # 显示算法执行状态
    if algorithm_status.get('can_execute_ai', False):
        enabled_count = algorithm_status.get('enabled_count', 0)
        status_items.append(
            html.Span([
                f"✅ AI算法可用: {enabled_count}个算法启用"
            ], className="me-3 badge bg-success")
        )
    else:
        status_items.append(
            html.Span([
                f"🚫 AI算法不可用: {algorithm_status.get('message', '未知原因')}"
            ], className="me-3 badge bg-danger")
        )
    
    # 显示各预测方法的最终状态
    method_display_names = {
        'ml_prediction_raw': 'AI预测-原始值',
        'ml_prediction_final': 'AI预测-最终值',
        'ratio_method_raw': '比率法-原始值',
        'ratio_method_final': '比率法-最终值'
    }
    
    for method_key, display_name in method_display_names.items():
        method_status = prediction_method_status.get(method_key, {})
        final_status = method_status.get('final_status', 'unknown')
        
        if final_status == 'fully_available':
            status_items.append(
                html.Span([f"✅ {display_name}: 可用"], className="me-3 badge bg-success")
            )
        elif final_status == 'execute_only':
            status_items.append(
                html.Span([f"⚠️ {display_name}: 可执行但显示被禁用"], className="me-3 badge bg-warning")
            )
        elif final_status == 'display_error':
            status_items.append(
                html.Span([f"❌ {display_name}: 显示启用但无法执行"], className="me-3 badge bg-danger")
            )
        else:  # fully_disabled or unknown
            status_items.append(
                html.Span([f"🚫 {display_name}: 不可用"], className="me-3 badge bg-secondary")
            )
    
    # 确定整体状态消息
    available_methods = sum(1 for method_status in prediction_method_status.values() 
                          if method_status.get('final_status') == 'fully_available')
    total_methods = len(method_display_names)
    
    if available_methods == 0:
        status_color = "danger"
        main_message = f"所有{mode_name}预测方法都不可用，请检查算法配置和综合指标设置"
    elif available_methods == total_methods:
        status_color = "success"
        main_message = f"所有{mode_name}预测方法都可用，可以进行全面的预测分析对比"
    else:
        status_color = "warning"
        main_message = f"部分{mode_name}预测方法可用 ({available_methods}/{total_methods})，建议检查相关配置以获得更准确的分析结果"
    
    # 构建返回的HTML组件
    return html.Div([
        html.Div([
            html.H6(f"📊 {mode_name}模式预测方法状态", className="mb-2"),
            html.P(main_message, className="mb-2"),
            html.Div(status_items, className="mb-0")
        ], className=f"alert alert-{status_color}", style={
            'borderRadius': '10px',
            'border': f'2px solid var(--bs-{status_color})'
        })
    ], className="mb-4")

def create_enhanced_prediction_summary_with_status(ml_predictions, ratio_prediction_value, method_status):
    """创建增强版预测结果总结 - 支持状态控制"""
    ml_enabled = method_status.get('ml_prediction_raw', {}).get('enabled', True)
    ratio_enabled = method_status.get('ratio_method_raw', {}).get('enabled', True)
    
    if not ml_enabled and not ratio_enabled:
        return html.Div([
            dbc.Alert([
                html.I(className="fas fa-exclamation-triangle me-2"),
                html.Strong("⚠️ 所有预测方法都已禁用"),
                html.Br(),
                "请前往数据管理模块启用相关预测方法指标"
            ], color="danger")
        ])
    
    if not ml_enabled or not ratio_enabled:
        missing_method = "机器学习预测" if not ml_enabled else "比率法预测"
        return html.Div([
            dbc.Alert([
                html.I(className="fas fa-info-circle me-2"),
                html.Strong(f"部分预测方法不可用：{missing_method}已禁用"),
                html.Br(),
                "建议启用完整的预测方法以获得更准确的分析结果"
            ], color="warning")
        ])
    
    # 如果两种方法都启用，使用原来的逻辑
    return create_enhanced_prediction_summary(ml_predictions, ratio_prediction_value)


def create_enhanced_prediction_summary(ml_predictions, ratio_prediction_value):
    """创建增强版预测结果总结"""
    if not ml_predictions and not ratio_prediction_value:
        return html.Div()
    
    summary_content = []
    
    if ml_predictions and ml_predictions.get('集成平均预测') and ratio_prediction_value:
        ml_pred = ml_predictions['集成平均预测']
        
        diff_pct = abs(ml_pred - ratio_prediction_value) / max(ml_pred, ratio_prediction_value) * 100 if max(ml_pred, ratio_prediction_value) != 0 else 0

        if diff_pct < 10:
            confidence = "高"
            color = "success"
            icon = "✅"
            recommendation = "两种方法预测总价接近，建议采用集成平均预测值。"
        elif diff_pct < 20:
            confidence = "中"
            color = "warning" 
            icon = "⚠️"
            recommendation = "两种方法预测总价存在一定差异，建议结合项目特点综合判断。"
        else:
            confidence = "低"
            color = "danger"
            icon = "❌"
            recommendation = "两种方法预测总价差异较大，建议重新检查输入数据或寻求专家意见。"
        
        summary_content.append(
            dbc.Row([
                dbc.Col(dbc.Card([
                    dbc.CardHeader(html.H6(f"{icon} 预测可信度评估: {confidence}", className="mb-0")),
                    dbc.CardBody([
                        html.P(f"🤖 机器学习集成平均预测总价: {ml_pred:,.0f} 元", className="mb-1"),
                        html.P(f"📈 比率法预测总价: {ratio_prediction_value:,.0f} 元", className="mb-1"),
                        html.P(f"📊 预测总价差异: {diff_pct:.1f}%", className="mb-2"),
                        html.Hr(),
                        html.P(recommendation, className="small text-muted mb-0")
                    ])
                ], color=color, outline=True), width=6),
                dbc.Col(dbc.Card([
                    dbc.CardHeader(html.H6("📈 预测总价区间建议", className="mb-0")),
                    dbc.CardBody([
                        html.P(f"预测总价中值: {(ml_pred + ratio_prediction_value) / 2:,.0f} 元", className="mb-1"),
                        html.P(f"预测总价范围: {min(ml_pred, ratio_prediction_value):,.0f} - {max(ml_pred, ratio_prediction_value):,.0f} 元", className="mb-1"),
                        html.P(f"建议预算 (含10%缓冲): {(ml_pred + ratio_prediction_value) / 2 * 1.1:,.0f} 元", className="mb-2"),
                        html.Hr(),
                        html.P("建议在预测总价基础上增加10%的风险缓冲。", className="small text-muted mb-0")
                    ])
                ], color="info", outline=True), width=6)
            ], className="mb-3")
        )
    
    return html.Div(summary_content)

def get_quantity_unit(param_name):
    unit_mapping = {
        '钢筋总吨数': '吨', '塔吊租赁工程量': '台班',
        '吊索具数量': '套', '套筒数量': '个',
        '措施费工程量': '元',
        '拼装场地总工程量': '立方米',
        '制作胎具总工程量': '吨',
        '钢支墩埋件总工程量': '吨',
        '扶壁柱总工程量': '吨',
        '走道板操作平台总工程量': '吨',
        '钢网梁总工程量': '吨',
        '钢支墩埋件混凝土剔凿总工程量': '立方米',
        '钢支墩埋件混凝土回填总工程量': '立方米',
        '钢支墩埋件安装总工程量': '吨',
        '钢支墩埋件制作总工程量': '吨',
        '扶壁柱安装总工程量': '吨',
        '扶壁柱拆除总工程量': '吨',
        '扶壁柱构件使用折旧总工程量': '吨',
        '走道板操作平台制作总工程量': '吨',
        '走道板操作平台搭设总工程量': '吨',
        '走道板操作平台拆除总工程量': '吨',
        '钢网架制作总工程量': '吨',
        '钢网架安装总工程量': '吨',
        '钢网架拆除总工程量': '吨',
    }
    return unit_mapping.get(param_name, '')




def perform_custom_mode_calculation(selected_params):
    """执行自定义模式的预测计算逻辑"""
    try:
        # 数据预处理
        processed_params = []
        total_known_quantity = 0
        total_known_price = 0
        total_quantity_ratio = 0
        total_price_ratio = 0
        
        for param in selected_params:
            processed_param = {
                'name': param.get('param_name', '未知参数'),
                'quantity': safe_float(param.get('param_quantity')),
                'quantity_ratio': safe_float(param.get('quantity_ratio')),
                'price_amount': safe_float(param.get('price_amount')),
                'price_ratio': safe_float(param.get('price_ratio')),
                'key_factor': param.get('key_factor', ''),
            }
            processed_params.append(processed_param)
            
            # 统计已知数据
            if processed_param['quantity'] > 0:
                total_known_quantity += processed_param['quantity']
            if processed_param['price_amount'] > 0:
                total_known_price += processed_param['price_amount']
            if processed_param['quantity_ratio'] > 0:
                total_quantity_ratio += processed_param['quantity_ratio']
            if processed_param['price_ratio'] > 0:
                total_price_ratio += processed_param['price_ratio']
        
        if len(processed_params) == 0:
            return {"error": "没有有效的参数用于计算"}
        
        # Step 1: 估算工程量
        estimated_quantities = estimate_custom_quantities(processed_params, total_known_quantity, total_quantity_ratio)
        
        # Step 2: 估算价格
        estimated_prices = estimate_custom_prices(processed_params, total_known_price, total_price_ratio)
        
        # Step 3: 计算预测总价
        total_prediction = calculate_custom_total_prediction(estimated_quantities, estimated_prices)
        
        return {
            "input_params": processed_params,
            "estimated_quantities": estimated_quantities,
            "estimated_prices": estimated_prices, 
            "total_prediction": total_prediction
        }
        
    except Exception as e:
        logger.error(f"自定义模式计算逻辑异常: {e}", exc_info=True)
        return {"error": f"计算逻辑异常: {str(e)}"}


def estimate_custom_quantities(processed_params, total_known_quantity, total_quantity_ratio):
    """估算自定义参数的工程量"""
    estimated_quantities = {}
    
    # 如果有工程量占比，需要估算基准总工程量
    if total_quantity_ratio > 0 and total_known_quantity > 0:
        # 假设已知工程量占总工程量的比例为其对应的占比
        # 这里使用一个简化的逻辑：基准总工程量 = 已知总工程量 / (已知占比/100)
        estimated_total_quantity = total_known_quantity * 100 / total_quantity_ratio if total_quantity_ratio > 0 else total_known_quantity
    else:
        estimated_total_quantity = total_known_quantity
    
    for param in processed_params:
        param_name = param['name']
        
        if param['quantity'] > 0:
            # 使用用户直接输入的工程量
            estimated_quantities[param_name] = param['quantity']
        elif param['quantity_ratio'] > 0:
            # 根据占比估算工程量
            estimated_quantities[param_name] = estimated_total_quantity * param['quantity_ratio'] / 100
        else:
            # 没有工程量信息，设为0
            estimated_quantities[param_name] = 0.0
    
    return estimated_quantities


def estimate_custom_prices(processed_params, total_known_price, total_price_ratio):
    """估算自定义参数的价格"""
    estimated_prices = {}
    
    # 如果有价格占比，需要估算基准总价格
    if total_price_ratio > 0 and total_known_price > 0:
        # 基准总价格 = 已知总价格 / (已知占比/100)
        estimated_total_price = total_known_price * 100 / total_price_ratio if total_price_ratio > 0 else total_known_price
    else:
        estimated_total_price = total_known_price
    
    for param in processed_params:
        param_name = param['name']
        
        if param['price_amount'] > 0:
            # 使用用户直接输入的价格
            estimated_prices[param_name] = param['price_amount']
        elif param['price_ratio'] > 0:
            # 根据占比估算价格
            estimated_prices[param_name] = estimated_total_price * param['price_ratio'] / 100
        else:
            # 没有价格信息，设为0
            estimated_prices[param_name] = 0.0
    
    return estimated_prices


def calculate_custom_total_prediction(estimated_quantities, estimated_prices):
    """计算自定义模式的预测总价"""
    total_cost = 0.0
    param_costs = {}
    
    # 获取所有参数名称
    all_params = set(list(estimated_quantities.keys()) + list(estimated_prices.keys()))
    
    for param_name in all_params:
        quantity = estimated_quantities.get(param_name, 0.0)
        price = estimated_prices.get(param_name, 0.0)
        
        # 如果价格已经是总价，直接使用；否则按工程量×单价计算
        if price > 0:
            param_cost = price  # 假设输入的价格量就是该参数的总成本
        else:
            param_cost = 0.0
            
        param_costs[param_name] = param_cost
        total_cost += param_cost
    
    return {
        "total_predicted_cost": total_cost,
        "param_costs": param_costs
    }


def create_custom_mode_results_table(input_params, estimated_quantities, estimated_prices, total_prediction):
    """创建自定义模式预测结果表格"""
    
    title_section = html.Div([
        html.H3("自定义模式智能预测分析报告", className="text-center text-primary mb-4"),
        html.Hr()
    ])
    
    # 输入参数汇总
    input_summary_rows = []
    for param in input_params:
        param_name = param['name']
        quantity = estimated_quantities.get(param_name, 0.0)
        price = estimated_prices.get(param_name, 0.0)
        
        input_summary_rows.append(
            html.Tr([
                html.Td(param_name),
                html.Td(f"{quantity:.2f}"),
                html.Td(f"{price:,.2f}"),
                html.Td(param.get('key_factor', '-'))
            ])
        )
    
    input_summary = html.Div([
        html.H5("📋 参数输入与估算汇总", className="text-info mb-3"),
        dbc.Table([
            html.Thead(html.Tr([
                html.Th("参数名称"),
                html.Th("估算工程量"),
                html.Th("估算价格(元)"), 
                html.Th("关键因素")
            ])),
            html.Tbody(input_summary_rows)
        ], bordered=True, hover=True, striped=True, className="mb-4")
    ])
    
    # 预测结果部分
    prediction_rows = []
    param_costs = total_prediction.get("param_costs", {})


    # 【在这里添加算法状态检查】
    # 如果自定义模式也使用机器学习预测，需要显示算法状态
    if hasattr(total_prediction, 'get') and total_prediction.get("ml_predictions"):
        ml_predictions = total_prediction.get("ml_predictions", {})
        
        # 显示算法状态（启用和停用）
        for model_name, prediction_value in ml_predictions.items():
            if model_name in ['集成平均预测', '_algorithm_status']:
                continue
            
            # 检查是否为停用算法
            if isinstance(prediction_value, dict) and prediction_value.get('status') == 'disabled':
                prediction_rows.append(
                    html.Tr([
                        html.Td(f"🚫 {model_name} (已停用)"),
                        html.Td("-"),
                        html.Td("-"),
                        html.Td("若需启用请到数据管理模块")
                    ])
                )
    # 【新代码结束】 
    # 
    #    
    for param_name, cost in param_costs.items():
        if cost > 0:
            prediction_rows.append(
                html.Tr([
                    html.Td(f"💰 {param_name}"),
                    html.Td(f"{cost:,.2f}"),
                    html.Td("元"),
                    html.Td("基于用户输入和智能估算")
                ])
            )
    
    # 总价预测
    total_cost = total_prediction.get("total_predicted_cost", 0.0)
    prediction_rows.append(
        html.Tr([
            html.Td("🎯 预测总价", className="fw-bold", style={'color': 'white'}),
            html.Td(f"{total_cost:,.0f}", className="fw-bold", style={'color': 'white'}),
            html.Td("元", className="fw-bold", style={'color': 'white'}),
            html.Td("所有参数成本之和", className="fw-bold", style={'color': 'white'})
        ], className="table-primary", style={'backgroundColor': PRIMARY_COLOR})
    )
    
    prediction_section = html.Div([
        html.H5("🎯 智能预测分析结果", className="text-primary mb-3"),
        dbc.Table([
            html.Thead(html.Tr([
                html.Th("预测项目", style={"width": "30%"}),
                html.Th("预测价格", style={"width": "20%"}),
                html.Th("单位", style={"width": "10%"}),
                html.Th("说明", style={"width": "40%"})
            ])),
            html.Tbody(prediction_rows)
        ], bordered=True, hover=True, striped=True, size="sm", className="mb-4")
    ])
    
    # 计算说明
    calculation_summary = html.Div([
        html.H6("💡 计算逻辑说明", className="text-muted mb-3"),
        dbc.Row([
            dbc.Col(dbc.Card([
                dbc.CardHeader("📊 工程量估算"),
                dbc.CardBody(html.Ul([
                    html.Li("优先使用用户直接输入的工程量数值"),
                    html.Li("根据工程量占比推算未知参数的工程量"),
                    html.Li("基于已知参数建立比例关系")
                ], className="small mb-0"))
            ]), width=4),
            dbc.Col(dbc.Card([
                dbc.CardHeader("💰 价格估算"),
                dbc.CardBody(html.Ul([
                    html.Li("优先使用用户直接输入的价格数值"),
                    html.Li("根据价格占比推算未知参数的价格"),
                    html.Li("智能分析参数间的成本关系")
                ], className="small mb-0"))
            ]), width=4),
            dbc.Col(dbc.Card([
                dbc.CardHeader("🎯 预测汇总"),
                dbc.CardBody(html.Ul([
                    html.Li("将所有参数的预测成本相加"),
                    html.Li("生成项目总体预测价格"),
                    html.Li("提供详细的计算过程透明度")
                ], className="small mb-0"))
            ]), width=4)
        ])
    ])
    
    return html.Div([title_section, input_summary, prediction_section, calculation_summary], 
                   className="container-fluid")

def create_detailed_error_report(exception, mode_name, error_context):
    """
    创建详细的错误报告
    
    Args:
        exception (Exception): 异常对象
        mode_name (str): 模式名称
        error_context (str): 错误上下文
        
    Returns:
        dict: 详细的错误报告
    """
    import traceback
    
    error_report = {
        "错误时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "模式": mode_name,
        "错误上下文": error_context,
        "异常类型": type(exception).__name__,
        "错误消息": str(exception),
        "完整错误": traceback.format_exc(),
        "可能原因": [],
        "建议解决方案": [],
        "严重级别": "未知"
    }
    
    # 根据异常类型分析可能原因和解决方案
    exception_type = type(exception).__name__
    error_message = str(exception).lower()
    
    if "mysql" in error_message or "connection" in error_message:
        error_report["可能原因"] = [
            "MySQL数据库连接失败",
            "数据库服务未启动",
            "数据库配置信息错误",
            "网络连接问题"
        ]
        error_report["建议解决方案"] = [
            "检查MySQL服务是否正常运行",
            "验证数据库连接配置（主机、用户名、密码）",
            "确认数据库名称是否正确",
            "检查防火墙设置"
        ]
        error_report["严重级别"] = "高"
        
    elif "table" in error_message or "column" in error_message:
        error_report["可能原因"] = [
            f"{mode_name}模式相关数据表不存在或结构异常",
            "数据库表结构与代码不匹配",
            "数据迁移未完成"
        ]
        error_report["建议解决方案"] = [
            f"检查{mode_name}模式相关数据表是否存在",
            "验证数据表结构是否符合系统要求",
            "重新初始化数据库或导入数据",
            "联系系统管理员检查数据库状态"
        ]
        error_report["严重级别"] = "高"
        
    elif "algorithm" in error_message or "model" in error_message:
        error_report["可能原因"] = [
            f"{mode_name}模式算法配置异常",
            "machine learning模型训练失败",
            "算法状态配置错误"
        ]
        error_report["建议解决方案"] = [
            "检查算法配置表中的状态设置",
            "验证训练数据的完整性",
            "重新训练机器学习模型",
            "检查算法依赖库是否正常"
        ]
        error_report["严重级别"] = "中"
        
    elif "indicator" in error_message or "comprehensive" in error_message:
        error_report["可能原因"] = [
            "综合指标配置异常",
            f"{mode_name}模式综合指标状态错误",
            "指标映射关系配置问题"
        ]
        error_report["建议解决方案"] = [
            "检查comprehensive_indicators表配置",
            "验证指标编码和状态设置",
            "重新配置指标映射关系",
            "确认模式对应的指标是否完整"
        ]
        error_report["严重级别"] = "中"
        
    elif "permission" in error_message or "access" in error_message:
        error_report["可能原因"] = [
            "数据库访问权限不足",
            "文件系统权限限制",
            "用户权限配置问题"
        ]
        error_report["建议解决方案"] = [
            "检查数据库用户权限",
            "验证文件系统访问权限",
            "联系系统管理员调整权限设置"
        ]
        error_report["严重级别"] = "中"
        
    else:
        error_report["可能原因"] = [
            "系统内部逻辑错误",
            "数据格式异常",
            "配置参数错误",
            "依赖组件故障"
        ]
        error_report["建议解决方案"] = [
            "检查输入数据格式是否正确",
            "验证系统配置文件",
            "重启应用程序",
            "查看详细错误日志",
            "联系技术支持"
        ]
        error_report["严重级别"] = "中"
    
    return error_report


def create_error_display_component(error_report):
    """
    创建错误显示组件
    
    Args:
        error_report (dict): 详细错误报告
        
    Returns:
        html.Div: 错误显示组件
    """
    severity_colors = {
        "高": "danger",
        "中": "warning", 
        "低": "info",
        "未知": "secondary"
    }
    
    severity = error_report.get("严重级别", "未知")
    alert_color = severity_colors.get(severity, "danger")
    
    # 构建错误显示内容
    error_content = [
        html.H5([
            html.I(className="fas fa-exclamation-triangle me-2"),
            f"{error_report.get('模式', '')}模式预测失败"
        ], className="text-danger"),
        
        html.P([
            html.Strong("错误时间: "),
            error_report.get("错误时间", "未知")
        ], className="mb-2"),
        
        html.P([
            html.Strong("错误类型: "),
            error_report.get("异常类型", "未知"),
            html.Span(f" (严重级别: {severity})", className=f"badge bg-{alert_color} ms-2")
        ], className="mb-2"),
        
        html.P([
            html.Strong("错误描述: "),
            error_report.get("错误消息", "未知错误")
        ], className="mb-3")
    ]
    
    # 添加可能原因
    if error_report.get("可能原因"):
        error_content.extend([
            html.H6("🔍 可能原因:", className="text-warning mt-3"),
            html.Ul([
                html.Li(reason) for reason in error_report["可能原因"]
            ], className="mb-3")
        ])
    
    # 添加解决方案
    if error_report.get("建议解决方案"):
        error_content.extend([
            html.H6("💡 建议解决方案:", className="text-info"),
            html.Ol([
                html.Li(solution) for solution in error_report["建议解决方案"]
            ], className="mb-3")
        ])
    
    # 添加技术详情（可折叠）
    error_content.extend([
        html.Hr(),
        dbc.Accordion([
            dbc.AccordionItem([
                html.Pre(error_report.get("完整错误", "无详细错误信息"), 
                        style={
                            'background-color': '#f8f9fa',
                            'padding': '10px',
                            'border-radius': '5px',
                            'font-size': '12px',
                            'max-height': '300px',
                            'overflow-y': 'auto'
                        })
            ], title="📋 技术详情（供开发人员参考）")
        ], start_collapsed=True, className="mt-3")
    ])
    
    return html.Div([
        dbc.Alert(error_content, color=alert_color, className="mb-4")
    ])


def create_system_status_check_component(mode):
    """
    创建系统状态检查组件
    
    Args:
        mode (str): 模式名称
        
    Returns:
        html.Div: 状态检查组件
    """
    try:
        from .data import CostPredictionSystem
        
        # 初始化系统进行状态检查
        temp_system = CostPredictionSystem(mode=mode)
        
        # 检查各个组件状态
        status_checks = {
            "数据库连接": check_database_connection(),
            "算法配置": check_algorithm_configs(temp_system),
            "综合指标": check_comprehensive_indicators(temp_system),
            "历史数据": check_historical_data(temp_system),
            "预测能力": check_prediction_capability(temp_system)
        }
        
        # 构建状态显示
        status_items = []
        overall_status = "正常"
        
        for check_name, check_result in status_checks.items():
            if check_result["status"] == "正常":
                status_items.append(
                    html.Li([
                        html.I(className="fas fa-check-circle text-success me-2"),
                        html.Strong(check_name + ": "),
                        check_result["message"]
                    ])
                )
            elif check_result["status"] == "警告":
                status_items.append(
                    html.Li([
                        html.I(className="fas fa-exclamation-triangle text-warning me-2"),
                        html.Strong(check_name + ": "),
                        check_result["message"]
                    ])
                )
                if overall_status == "正常":
                    overall_status = "警告"
            else:
                status_items.append(
                    html.Li([
                        html.I(className="fas fa-times-circle text-danger me-2"),
                        html.Strong(check_name + ": "),
                        check_result["message"]
                    ])
                )
                overall_status = "异常"
        
        status_color = {"正常": "success", "警告": "warning", "异常": "danger"}[overall_status]
        
        return html.Div([
            dbc.Alert([
                html.H6(f"📊 {mode}模式系统状态检查", className="mb-3"),
                html.P(f"整体状态: {overall_status}", className=f"text-{status_color} fw-bold"),
                html.Ul(status_items, className="mb-0")
            ], color=status_color, className="mt-3")
        ])
        
    except Exception as e:
        return html.Div([
            dbc.Alert([
                html.H6("⚠️ 系统状态检查失败", className="text-danger"),
                html.P(f"无法执行系统状态检查: {str(e)}")
            ], color="danger")
        ])


def check_database_connection():
    """检查数据库连接状态"""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT 1")
        cursor.fetchone()
        conn.close()
        return {"status": "正常", "message": "数据库连接正常"}
    except Exception as e:
        return {"status": "异常", "message": f"数据库连接失败: {str(e)}"}


def check_algorithm_configs(system):
    """检查算法配置状态"""
    try:
        system.load_algorithm_configs()
        capability = system.check_algorithm_execution_capability()
        
        if capability.get('can_execute_ai', False):
            enabled_count = capability.get('enabled_count', 0)
            total_count = capability.get('total_count', 0)
            if enabled_count == total_count:
                return {"status": "正常", "message": f"所有算法正常({enabled_count}/{total_count})"}
            else:
                return {"status": "警告", "message": f"部分算法可用({enabled_count}/{total_count})"}
        else:
            return {"status": "异常", "message": "所有算法不可用"}
    except Exception as e:
        return {"status": "异常", "message": f"算法配置检查失败: {str(e)}"}


def check_comprehensive_indicators(system):
    """检查综合指标状态"""
    try:
        system.load_comprehensive_indicators_status()
        methods = ['ml_prediction_raw', 'ml_prediction_final', 'ratio_method_raw', 'ratio_method_final']
        
        available_count = 0
        for method in methods:
            perm = system.check_result_display_permission(method)
            if perm.get('can_display', False):
                available_count += 1
        
        if available_count == len(methods):
            return {"status": "正常", "message": f"所有预测方法可显示({available_count}/{len(methods)})"}
        elif available_count > 0:
            return {"status": "警告", "message": f"部分预测方法可显示({available_count}/{len(methods)})"}
        else:
            return {"status": "异常", "message": "所有预测方法显示被禁用"}
    except Exception as e:
        return {"status": "异常", "message": f"综合指标检查失败: {str(e)}"}


def check_historical_data(system):
    """检查历史数据状态"""
    try:
        if system.load_data_from_database(system.mode):
            if system.df_historical is not None and len(system.df_historical) > 0:
                data_count = len(system.df_historical)
                return {"status": "正常", "message": f"历史数据正常({data_count}条记录)"}
            else:
                return {"status": "异常", "message": "历史数据为空"}
        else:
            return {"status": "异常", "message": "历史数据加载失败"}
    except Exception as e:
        return {"status": "异常", "message": f"历史数据检查失败: {str(e)}"}


def check_prediction_capability(system):
    """检查预测能力状态"""
    try:
        if system.train_system():
            ai_capable = system.can_execute_ai_prediction()
            ratio_capable = system.can_execute_ratio_prediction()
            
            if ai_capable and ratio_capable:
                return {"status": "正常", "message": "AI预测和比率法预测都可用"}
            elif ai_capable or ratio_capable:
                available_method = "AI预测" if ai_capable else "比率法预测"
                return {"status": "警告", "message": f"只有{available_method}可用"}
            else:
                return {"status": "异常", "message": "所有预测方法都不可用"}
        else:
            return {"status": "异常", "message": "预测系统训练失败"}
    except Exception as e:
        return {"status": "异常", "message": f"预测能力检查失败: {str(e)}"}
    